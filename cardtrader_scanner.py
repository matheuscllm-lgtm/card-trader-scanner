#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║     CardTrader Arbitrage Scanner — Pokémon TCG Singles             ║
║                                                                      ║
║  Compara preços de singles (EN, Near Mint, não-graded) no           ║
║  cardtrader.com vs preço TCG Player (market price).                 ║
║  Gera planilha .xlsx com alertas de arbitragem (default ≥ 30%).     ║
╚══════════════════════════════════════════════════════════════════════╝

Por que CardTrader em vez de MYP?
    MYP é Brasil (compra em BRL) → o scanner MYP busca cartas baratas no
    Brasil para exportar aos EUA. Já o CardTrader é marketplace europeu
    (Itália) → agregador de sellers da UE com preços em EUR. Oportunidade:
    cartas valorizadas no mercado US que estão desatualizadas na UE.
    Setup inverso, mesma tese.

Modelo operacional (2026-06-06, decisão do operador — SUPERSEDE o × 1.06):
    A margem reportada é BRUTA. Nenhuma taxa é embutida pelo scanner:

        margem bruta = (preço de referência TCG − preço do site CT) / TCG

    O operador calcula Hub fee, frete, cartão e IOF POR FORA, manualmente.
    Por isso `--hub-fee` agora tem DEFAULT 0.0 e `--shipping-brl` segue 0.0.
    (Histórico: até v2.11 o scanner embutia 6% de Hub fee — `custo = preço × 1.06`.
    Quem quiser reproduzir aquele comportamento passa `--hub-fee 0.06`.)

Uso:
    python cardtrader_scanner.py                          # Scan padrão
    python cardtrader_scanner.py --sets sv1 sv2 sv3       # Sets específicos
    python cardtrader_scanner.py --threshold 35           # Margem mín 35%
    python cardtrader_scanner.py --min-price-usd 15       # Preço mín USD
    python cardtrader_scanner.py --dry-run                # Usa só cache
    python cardtrader_scanner.py --provider justtcg       # Troca pricing API

Env vars (arquivo .env na mesma pasta):
    CT_JWT=<JWT do CardTrader: Settings → API Access → Create New Token>
    POKEMONTCG_API_KEY=<opcional, grátis em pokemontcg.io/dev>
    JUSTTCG_API_KEY=<opcional, se usar provider justtcg>

Requisitos:
    pip install requests openpyxl python-dotenv pyyaml

Autor: Elizandra / Claude
Data: 2026-04-20 (v1.0) | 2026-04-29 (v2.1) | 2026-05-12 (v2.2 + v2.3)
      | 2026-05-16 (v2.5) | 2026-05-18 (v2.7/v2.8) | 2026-05-19 (v2.9/v2.10)
      | 2026-06-02 (v2.11) | 2026-06-06 (v2.12) | 2026-06-15 (v2.14/v2.15)
      | 2026-06-20 (v2.17/v2.18) | 2026-06-21 (v2.19)
Versão: v2.19
    (= changelog inline mais recente. Manter em sync ao adicionar novos blocos
     de changelog abaixo.)

Changelog v2.19 (2026-06-21 — validação per-blueprint casa condição NM + reverse/variante):
  - [CORREÇÃO] validate_per_blueprint casava o listing per-blueprint SÓ pelo
    username do vendedor, IGNORANDO condição e reverse. Um vendedor com uma cópia
    Poor barata (ou a versão não-reverse) "contaminava" a validação. Caso real
    Shiftry hl: o scan achou a NM reverse (~R$140) referenciada a reverseHolofoil
    $42.95, mas a validação casava a Poor não-reverse do MESMO vendedor (R$46,46)
    → comparava preço de carta Poor vs referência reverse → falso "79%" (o
    ground-truth real: holo NM R$86,15 vs $20,27 ≈ 18%; reverse NM R$139,70 vs
    $42,94 ≈ 37% — nunca 79%). Isso RESSUSCITAVA falsos positivos DEPOIS do fix
    v2.18, na etapa que deveria ser o ground-truth.
  - [FIX] o match agora exige condition == listing.condition (Near Mint, garantido
    pelo filtro NM-only) E reverse igual (pokemon_reverse/mtg_foil/foil == listing.foil,
    simétrico com o parse do scan v2.18); entre os que batem, pega o MAIS BARATO.
    Se nenhum bate → STALE honesto (a cópia exata que o scan viu sumiu), em vez
    de casar a cópia errada.
  - [LIMPEZA] removido script órfão local validate_vintage_candidates.py (gitignored,
    one-off) que carregava o mesmo padrão de match buggy.
  - INALTERADO: tiers de markup, recálculo de margem real, contrato STALE→NÃO no
    postprocess, threshold fração, margem bruta.
  - Testes: tests/test_per_blueprint_variant_match.py (4 casos: reverse, holo
    padrão, mais-barato-entre-iguais, STALE). Suíte 110/110 verde.

Changelog v2.18 (2026-06-20 — fim da inflação de holo rare vintage; pricing por raridade+reverse):
  - [CORREÇÃO-RAIZ] Cartas Pokémon SEMPRE chegavam foil=False: o parse do listing
    lia props["mtg_foil"]/["foil"] (campos MTG, inexistentes em Pokémon) e IGNORAVA
    `pokemon_reverse`. Consequência: TODA "Holo Rare" PADRÃO caía no ramo foil=False
    e, sem variante `normal` na TCGPlayer (holo rare não tem printagem normal),
    escorregava pra `reverseHolofoil` — variante reverse, mais rara e CARA →
    margem inflada sistemática. Exemplos reais (pokemontcg.io 2026-06-20):
    Gengar Legendary Collection holofoil $146.89 vs reverse $1599.99 (10×!);
    Claydol EX Hidden Legends $14.47 vs $44.95 (+210%); Shiftry hl $19.92 vs
    $42.95 (+116%). A varredura vintage de 2026-06-19 era quase toda falso
    positivo por causa disso.
  - [FIX] (1) o parse lê `pokemon_reverse` → reverse-holo chega como foil=True.
    (2) seleção de variante ciente de raridade (novo helper _rarity_is_holo +
    param `rarity` em market_price_usd, alimentado por l.rarity no call site):
      • foil=True (reverse)        → reverseHolofoil primeiro;
      • foil=False + raridade holo → holofoil/unlimitedHolofoil (printagem holo
        base), reverseHolofoil POR ÚLTIMO;
      • foil=False + não-holo      → normal, reverse por último (era 2º).
    Verificado no data model do CardTrader: pokemon_rarity="Holo Rare" +
    pokemon_reverse=false = holo padrão (→ holofoil); reverse=true (→ reverse).
  - [FLAG] "Variante Baixa Confiança" NÃO dispara mais pra holo rare (holofoil é
    a variante CORRETA, não um casamento errado); só p/ carta NÃO-holo que casou
    preço holo (o caso Gengar não-foil → $1600).
  - INALTERADO: margem bruta, threshold fração, validação per-blueprint, skip-list,
    overrides de timeout por set, filtro TG##, --skip-backcatalog.
  - Testes: scripts/test_variant_disambiguation.py reescrito (semântica
    rarity+reverse; +regressões Shiftry/Gengar LC/Pidgey). Suíte 106/106 verde.

Changelog v2.17 (2026-06-20 — flag --skip-backcatalog; foco em sets recentes):
  - Nova flag --skip-backcatalog: restringe o scan às coleções modernas/curadas
    (PRIORITY_SET_CODES), pulando o back-catalog. Lição da auditoria 2026-06-08:
    back-catalog = mercado eficiente = ~0 deal acionável; o gap mora em
    lançamentos novos. Mais útil com --all-sets (corta ~832 → ~30 sets);
    combinado com --sets, intersecta com a lista do usuário.
  - Novo helper PURO filter_modern_sets(expansions, priority_codes=PRIORITY_SET_CODES):
    mantém só codes na lista priority, preserva ordem, case-insensitive, tolera
    code ausente/None. Testável offline (tests/test_skip_backcatalog.py, 11 casos).
  - INALTERADO: margem bruta, threshold fração, --hub-fee 0.0, validação
    per-blueprint, skip-list, overrides de timeout por set, filtro TG##.
  - Obs: v2.16 foi postprocess-only (tabela de entrega obrigatória); por isso o
    header do scanner pula de v2.15 → v2.17.

Changelog v2.15 (2026-06-15 — overrides de timeout por SET; fim do "churn" vintage):
  - [TIMEOUT vintage] Novo mapa code-level SET_TIMEOUT_OVERRIDES (código CT →
    segundos): df=1200s (20min), ds/n1/n4=1080s (18min). Sets vintage pesados
    NÃO cabiam no per-set-timeout default (8min) → estouravam SEMPRE → iam pra
    skip-list → eram pulados → nunca escaneados por completo ("churn"). Agora
    ganham fôlego sem o operador lembrar de passar --per-set-timeout à mão.
  - [RESOLVER] Novo effective_set_timeout_s(exp_code, default_s): o override só
    ELEVA o teto (max(default, override)); --per-set-timeout global maior ainda
    vence; --per-set-timeout 0 (desligado) deixa TUDO sem timeout (override não
    reativa); set sem override = comportamento histórico; case-insensitive.
  - scan_expansion resolve o timeout efetivo UMA vez no topo e o usa em todos os
    checks (pré-blueprints, pré-listings, loop de pricing) + na deadline_ts que
    governa retries/429-sleeps. _check_set_timeout ganhou param opcional
    timeout_s (None = retrocompat: cai no default da instância).
  - n2 (Neo Discovery) NÃO recebe override: é no-coverage genuíno (pokemontcg.io
    ~zero referência), tratado pelo cap de --max-consecutive-misses, não timeout.
  - INALTERADO: flag --per-set-timeout (default 8min); skip-list; cap de misses;
    mass-pricing-abort; margem bruta; threshold fração; filtro TG##.
  - Testes: tests/test_set_timeout_overrides.py (11 casos).

Changelog v2.14 (2026-06-15 — correção + robustez; resultados corretos e sem bug):
  - [CORREÇÃO timeout] _get._sleep_capped: quando o backoff pedido NÃO cabia no
    tempo restante (ex.: 429 Retry-After 60s com 3s de deadline), o retorno
    `time.monotonic() > deadline_ts` podia dar False por uma fração de segundo,
    o loop re-tentava e MASCARAVA o timeout. Agora um backoff TRUNCADO sinaliza
    estouro de forma determinística. (Teste pré-existente test_per_set_timeout_
    initial_calls estava VERMELHO no main por causa disso.)
  - [CORREÇÃO silent-failure pricing] pokemontcg.io: 429 (rate-limit) e 5xx
    eram tratados como "carta não encontrada" (miss) → sumiam deals em silêncio,
    envenenavam o contador de misses (abort no_coverage falso) e inflavam a taxa
    de falha. Novo _ptcg_get faz retry bounded em 429/5xx e, se persistir,
    LEVANTA requests.Timeout (→ contado como pricing_failure VISÍVEL, não miss).
    4xx ≠ 429 segue como sem-match legítimo.
  - [ROBUSTEZ concorrência] Run-guard: o scanner agora RECUSA iniciar se já há
    outra instância no mesmo --state-dir (lockfile scanner_run.lock). Evita o
    incidente de contenção de lock SQLite (pricing → ~9s/listing) que envenenava
    a skip-list. Escape hatch: --allow-concurrent.
  - [CORREÇÃO recovery FX] O checkpoint JSONL agora grava usd_brl/eur_brl no
    header; recover_from_checkpoint.py reconstrói a célula Stats usd_brl_rate.
    Antes o XLSX recuperado tinha FX=0.0 → a coluna "CT US$" da tabela de
    entrega do postprocess ficava vazia. (Classificação COMPRA/REVISAR NÃO era
    afetada — usa colunas em BRL.)
  - [SINAL variante] Nova coluna "Variante Baixa Confiança": "Sim" quando o
    listing é NÃO-foil mas o único preço veio de variante metálica premium
    (holofoil/unlimitedHolofoil). Só FLAG — não muda preço nem decisão.
    reverseHolofoil NÃO dispara (fallback intencional p/ Holo Rare vintage).

Changelog v2.12 (2026-06-06 — margem BRUTA pura; remove Hub fee do cálculo):
  - Decisão do operador (cross-scanner): scanner reporta APENAS margem bruta
    `(tcg − preço_site) / tcg`, threshold 30%, SEM nenhuma taxa embutida.
    Hub fee/frete/cartão/IOF passam a ser calculados FORA, pelo operador.
    SUPERSEDE a fórmula `× 1.06` (v2.3).
  - --hub-fee default 0.06 → 0.0 (constante HUB_FEE_RATE mantida só como
    referência histórica / opt-in via --hub-fee 0.06).
  - Construtor Scanner: hub_fee_rate default 0.06 → 0.0 (paridade programática).
  - --shipping-brl segue default 0.0; piso de preço $10 MANTIDO (é filtro).
  - --threshold segue fração 0.30 (=30%) — convenção inalterada.

Changelog v2.11 (2026-06-02 — scan completo + parciais ao vivo; integrado ao
main no saneamento 2026-06-05):
  - Flag --all-sets: escaneia TODAS as ~832 expansões (ignora --sets e config),
    pro scan COMPLETO (weekly).
  - PRIORITY_SET_CODES: no modo --all-sets, sets SV/curados são escaneados
    PRIMEIRO (catálogo CT vem old→new; protege os modernos do corte por timeout).
  - Companion `scripts/checkpoint_to_partial.py`: lê o .checkpoint.jsonl DURANTE
    o scan e emite PARTIAL_DEALS.md + .csv (revisão incremental; stdlib puro).
  - Validado em scan real de ~4h (2026-06-02, GH Actions, 832 exp).

Changelog v2.3 (2026-05-12 — alinhamento Hub fee scanner ↔ postprocess):
  - HUB_FEE_RATE = 0.06 promovido a constante; CLI flag --hub-fee
  - validate_per_blueprint agora aplica live_brl × (1 + hub_fee) como custo
    real no recalc de real_margin_pct / real_net_margin_pct / real_lucro_brl
  - Antes scanner era ~6pp otimista vs realidade (apenas postprocess aplicava
    a taxa). Agora margem REAL bate com BUY NOW/REJECT do relatório.
  - hub_fee_rate exportado na aba Stats do XLSX
  - Auto-conversão de --hub-fee > 1.0 (operador pode passar `6` ou `0.06`)

Changelog v2.2 (2026-05-12 — auditoria 9 bugs C/H/M + alinhamento modelo
consolidação):
  - C1: tier markup 30-45% reclassificado como "Alto markup" / VALIDATED_MARKUP
    (era PRICE_CHANGED descartado como erro)
  - C2: clean_collector_number aplicado no fallback bp.version
  - H1: --threshold > 1.0 auto-converte com warning (UX trap "25 vs 0.25")
  - H2: pricing provider agora aceita `foil=` e prioriza variante correta
    (reverseHolofoil vs normal); cache key inclui foil flag
  - M1: filter de validation_status sempre roda quando validate_top > 0
  - M2: opps sempre re-ordenados por real_net_margin_pct após validação
  - M3: SHIPPING_EUR_* promovidas a constantes; flag --shipping-brl
  - M4: log + counter pra moedas exóticas (GBP/JPY) dropadas
  - M5: log de supranumerários (collector_number > set.printedTotal) detecta
    SIR/SAR/HR pra revisão manual
  - Modelo de consolidação: frete default = 0 (cartas consolidadas em
    deposito Hub depot, frete diluído per-card ~R$0.30 → desprezível).
    Override via --shipping-brl X.

Changelog v2.1 (2026-04-29):
  - validate_per_blueprint: validação per-blueprint com tier markup
  - filtro anti-Trainer Gallery (_TRAINER_GALLERY_RE)
  - --validate-top, --min-net-margin CLI flags
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sqlite3
import sys
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterator, Optional

import portalocker
import requests
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── v2.3.2 fix (bug-hunt 2026-05-17): força UTF-8 no I/O ─────────────────────
# Paridade com postprocess (commit 26c27bd). openpyxl.Workbook.save() pega
# encoding do interpreter locale via algum caminho interno. Em Windows com
# locale pt-BR, sys.stdout.encoding default é cp1252 → headers UTF-8 são
# gravados como bytes UTF-8 mas LIDOS como latin-1, produzindo mojibake
# (`Decis�o`, `Pre�o CT`). PYTHONIOENCODING=utf-8 resolve, mas só quando o
# invocador seta a var (wrappers PS setam; scripts ad-hoc esquecem).
# Garantia defensiva aqui (defense-in-depth):
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    # Python <3.7 ou stdout não-reconfigurável (raro). Fallback silencioso.
    pass

# ══════════════════════════════════════════════════════════════════════
# LOGGING — saída para stdout E arquivo (auditoria pós-execução)
# ══════════════════════════════════════════════════════════════════════

# Log file path:
#   - CT_LOG_FILE env var (preferido — set pelo wrapper PS antes de invocar)
#   - default cardtrader_scanner.log no cwd
# Por que via env e nao via --log-file CLI?
#   logging.basicConfig roda no IMPORT (antes de argparse). Env var resolve
#   o problema de ordem sem refactor maior.
_LOG_FILE = os.environ.get("CT_LOG_FILE", "cardtrader_scanner.log")
_log_handlers = [
    logging.StreamHandler(sys.stdout),
    logging.FileHandler(_LOG_FILE, encoding="utf-8"),
]
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=_log_handlers,
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════
# CONFIG — valores default. CLI ou config.yaml sobrescrevem.
# Por que constantes no topo? Uma mudança aqui afeta o scanner inteiro;
# fica fácil ajustar sem ler o código todo.
# ══════════════════════════════════════════════════════════════════════
CT_BASE = "https://api.cardtrader.com/api/v2"
POKEMONTCG_BASE = "https://api.pokemontcg.io/v2"
JUSTTCG_BASE = "https://api.justtcg.com/v1"
FX_BASE = "https://api.frankfurter.app/latest"  # BCE — grátis, sem auth

# Game ID no CardTrader. Pokemon = 5 (MTG=1, Yu-Gi-Oh=4, One Piece=15, etc).
# Descobrível via GET /games se quebrar no futuro.
CT_POKEMON_GAME_ID = 5

# Defaults do scanner (sobrescritos por CLI)
MARGIN_THRESHOLD = 0.30          # 30% margem mínima (requisito do usuário)
MIN_PRICE_USD = 10.0             # filtro extra do usuário
LANGUAGE_FILTER = "en"           # apenas inglês
CONDITION_FILTER = "Near Mint"   # apenas NM
EXCLUDE_GRADED = True            # exclui PSA/BGS/CGC
REQUEST_DELAY_CT = 0.15          # CT tem limite 10/s → 0.15s deixa folga
REQUEST_DELAY_PRICING = 0.1      # pokemontcg.io permite 20k/dia (~0.23/s)
TIMEOUT = 30

# Frete base por tier de seller (EUR — CT é europeu por origem).
# Convertido pra BRL via FX no Scanner._estimate_shipping_brl.
# 2026-05-12 M3 fix: promovido de magic number a constantes.
SHIPPING_EUR_HUB = 5.0           # Hub seller / zero_fee — CT centraliza envio
SHIPPING_EUR_PROFESSIONAL = 10.0 # Professional seller — envia do país do seller
SHIPPING_EUR_PRIVATE = 12.0      # Private seller — envio mais caro

# Hub fee médio sobre preço do site (per-blueprint LIVE).
# v2.12 (2026-06-06 — decisão do operador): margem reportada agora é BRUTA.
# O DEFAULT efetivo aplicado é 0.0 (--hub-fee default 0.0): custo = preço do
# site, sem nenhuma taxa embutida. O operador calcula Hub fee/frete/cartão/IOF
# por fora, manualmente. Isso SUPERSEDE a fórmula antiga `× 1.06`.
# A constante abaixo permanece apenas como valor de referência histórico/opt-in:
# passe `--hub-fee 0.06` para reembutir os 6% (custo = live_brl × (1 + hub_fee)).
HUB_FEE_RATE = 0.06              # referência histórica — NÃO é mais o default (v2.12)

# Paths
SCRIPT_DIR = Path(__file__).resolve().parent
CACHE_DB = SCRIPT_DIR / "cache.db"
ENV_FILE = SCRIPT_DIR / ".env"
CONFIG_FILE = SCRIPT_DIR / "config.yaml"

# v2.4: per-set timeout + auto skip-list.
# Run 25898522951 (2026-05-15) cancelou em 30:09 com 5/10 sets processados
# pq o timeout-minutes do GH Actions matou o job, mas o problema mais antigo
# (run 25838570927 de 2026-05-14) foi um set unico travado 24m53s sem progresso.
# Solucao: wall-clock timeout per-set + persistir sets travados em skip-list.
SKIP_LIST_FILE = SCRIPT_DIR / "scanner_skip_list.json"
DEFAULT_PER_SET_TIMEOUT_MIN = 8  # conservador: pior caso medido foi 7min/set

# v2.14 (2026-06-15): overrides de timeout por SET (vintage churn fix).
# Problema: alguns sets vintage pesados (muitas listings + pokemontcg.io lento)
# NÃO cabem no per-set-timeout default (8min) e estouram SEMPRE. O timeout joga
# o set na skip-list; runs futuros o pulam; ele nunca mais é escaneado por
# completo. Vira "churn": entra na skip-list, é pulado, volta a entrar. O
# operador teria que lembrar de passar `--per-set-timeout 20` manualmente toda
# vez — frágil.
# Solução: um mapa code-level de overrides (segundos) pra sets confirmados que
# precisam de mais fôlego. É código (não config que o operador mantém à mão)
# porque são fatos estáveis sobre sets específicos — pertencem ao lado das
# outras constantes (regex TG, TTLs). O override só ELEVA o teto: se o operador
# passar um `--per-set-timeout` global ainda maior, esse vence (max). Se passar
# 0 (desativado), tudo fica sem timeout. Chaves = códigos CardTrader.
# Casos confirmados na investigação do ciclo vintage 2026-06-15:
#   df (EX Dragon Frontiers) — ~19min p/ 100% das 79 listings; 8min cortava sempre.
#   ds (EX Delta Species), n1 (Neo Genesis), n4 (Neo Destiny) — cobertura PARCIAL
#       cortada por timeout de 12min; escaneariam mais com budget maior.
# NÃO inclui n2 (Neo Discovery): aquele é NO-COVERAGE genuíno (pokemontcg.io ~zero
# referência → ~40 misses consecutivos), tratado pelo --max-consecutive-misses,
# não por timeout. Entrada de n2 na skip-list por miss-cap é defensável.
SET_TIMEOUT_OVERRIDES: dict[str, int] = {
    "df": 1200,  # EX Dragon Frontiers — 20min (medido ~19min p/ 100%)
    "ds": 1080,  # EX Delta Species — 18min (folga sobre os 12min parciais)
    "n1": 1080,  # Neo Genesis — 18min
    "n4": 1080,  # Neo Destiny — 18min
}


def effective_set_timeout_s(exp_code: str, default_s: float) -> float:
    """Resolve o timeout efetivo (segundos) pra um set.

    Regras:
      - default_s == 0 (timeout global desativado) → 0 (sem timeout pra ninguém,
        incluindo sets com override; o operador desligou de propósito).
      - set com override → max(default_s, override) — o override só ELEVA o teto;
        um --per-set-timeout global ainda maior continua vencendo.
      - set sem override → default_s (comportamento histórico).
    Código normalizado (lower/strip) pra casar com os códigos CardTrader.
    """
    if not default_s:
        return 0.0
    code = (exp_code or "").strip().lower()
    override = SET_TIMEOUT_OVERRIDES.get(code)
    if override is None:
        return default_s
    return float(max(default_s, override))

# TTL do cache. Blueprints praticamente não mudam (cartas impressas uma vez).
# Preços mudam diário → refresh.
BLUEPRINT_TTL = timedelta(days=30)
PRICE_TTL = timedelta(hours=24)
FX_TTL = timedelta(hours=12)


# Regex para extrair o número do collector de strings sujas tipo
# "Special Illustration Rare | 161/131" → "161". Capturamos o último grupo N/N
# (algumas cartas trazem códigos de promo antes).
_NUMBER_IN_VERSION = re.compile(r"(\d+)\s*/\s*\d+")

# v2.1 — Filtro anti-Trainer Gallery. Cartas com collector_number "TG##" do
# subset Trainer Gallery (era SWSH 2021-2023) produzem falsos positivos massivos
# no pricing pokemontcg.io: a base retorna preços inflados em 5-10× pra essas
# cartas, possivelmente porque bate com variantes secret-rare alt-art com mesmo
# nome. Resultado: 35 de 38 oportunidades em scan SWSH (2026-04-29 madrugada)
# eram falsos positivos com margens "90%+ líq" estatisticamente impossíveis.
# Solução: skipar listings cujo collector_number bate `^TG\d+`.
_TRAINER_GALLERY_RE = re.compile(r"^TG\d+", re.IGNORECASE)


# ══════════════════════════════════════════════════════════════════════
# CHASE TIER — classificação de raridade (v2.13, frente "chase cards")
# ══════════════════════════════════════════════════════════════════════
# Espelha CHASE_TIER_PATTERNS do postprocess (mesma taxonomia oficial
# PokemonTCG) pra manter coerência entre scanner e relatório. Substring-aware,
# case-insensitive. Mantenha as duas listas em sincronia se editar.
#   TOP    = raridades top (SIR/SAR/Hyper/Secret/Illustration Rare)
#   MID    = alt/full art, rainbow/gold, double/ultra rare
#   MODEST = holo/reverse holo, promos
#   BULK   = common/uncommon (sem liquidez mesmo barato)
CHASE_TIER_PATTERNS = {
    "TOP": [
        "special illustration rare", "sir", "illustration rare",
        "special art rare", "sar", "hyper rare", "ultra hyper rare",
        "secret rare", "rara secreta",
    ],
    "MID": [
        "full art", "alt art", "alternate art", "alternative art",
        "rainbow rare", "gold rare", "trainer gallery", "double rare",
        "rara hiper", "ultra rare",
    ],
    "MODEST": [
        "holo rare", "reverse holo", "reverse foil", "promo",
        "rare holo",
    ],
    "BULK": [
        "common", "comum", "uncommon", "incomum",
    ],
}
# Ordem de prioridade (alto→baixo) pra ranquear e filtrar --chase-only.
CHASE_TIER_ORDER = {"TOP": 3, "MID": 2, "MODEST": 1, "BULK": 0, "": -1}
# Tiers considerados "chase" pelo filtro --chase-only.
CHASE_ONLY_TIERS = {"TOP", "MID"}


def classify_chase_tier(rarity: Optional[str]) -> str:
    """Classifica o chase tier (TOP/MID/MODEST/BULK) a partir da rarity.

    `rarity` pode vir da rarity oficial pokemontcg.io (preferida) ou da rarity
    do blueprint CT (fallback). Sem rarity → "" (desconhecido), tratado como
    não-chase pelo filtro. A versão completa com proxies (número supranumerário,
    markup tier) vive no postprocess — aqui mantemos só o caminho por rarity,
    que é o que o scanner tem em mãos durante o scan.
    """
    if not rarity:
        return ""
    r = str(rarity).lower().strip()
    for tier, patterns in CHASE_TIER_PATTERNS.items():
        if any(p in r for p in patterns):
            return tier
    # Rarity presente mas não mapeada → MODEST (conservador).
    return "MODEST"


# ══════════════════════════════════════════════════════════════════════
# VALORIZAÇÃO — score heurístico honesto (v2.13, frente "médio/longo prazo")
# ══════════════════════════════════════════════════════════════════════
# IMPORTANTE (honestidade de dados): este score NÃO usa preço histórico nem
# dados de tendência — as APIs atuais (CardTrader + pokemontcg.io) NÃO fornecem
# série temporal de preços. O que temos é: (1) data de lançamento do set
# (pokemontcg.io set.releaseDate) → idade; (2) rarity oficial → chase tier;
# (3) preço de mercado atual. O score combina SÓ esses sinais derivávies, e
# cada componente é explicado em texto na coluna "Valorização — Notas".
# Não inventamos "alta histórica" nem "% do topo" porque a fonte não dá isso.
#
# Heurística (0-100), três componentes objetivos:
#   • Raridade (0-45):  TOP=45, MID=30, MODEST=12, BULK=0, desconhecida=8
#   • Maturidade (0-35): sets com 1.5-4 anos pontuam mais (já passaram do
#     dip pós-lançamento e ainda não saíram totalmente de circulação). Recém-
#     lançado (<6m) e muito antigo (>6 anos) pontuam menos — heurística simples,
#     não previsão. Set sem data → 10 (neutro).
#   • Preço-âncora (0-20): cartas de market mais alto têm mais "chão" e
#     histórico de retenção de valor; <$15→4, $15-40→10, $40-100→16, >$100→20.
# Limitações ficam documentadas na própria nota de cada linha.
def _valorization_age_component(set_release_date: Optional[str],
                                now: Optional[datetime] = None) -> tuple[int, str]:
    """Componente de maturidade (0-35) + nota textual. Data no formato
    'YYYY/MM/DD' (pokemontcg.io). Sem data → neutro."""
    if not set_release_date:
        return 10, "idade do set desconhecida (sem data na fonte)"
    now = now or datetime.now()
    parsed = None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(set_release_date, fmt)
            break
        except (ValueError, TypeError):
            continue
    if parsed is None:
        return 10, "idade do set desconhecida (data ilegível)"
    age_days = (now - parsed).days
    age_years = age_days / 365.25
    if age_years < 0.5:
        return 12, f"set recém-lançado (~{age_years:.1f} anos) — preço ainda instável"
    if age_years < 1.5:
        return 22, f"set jovem (~{age_years:.1f} anos) — saindo do dip pós-lançamento"
    if age_years <= 4.0:
        return 35, f"set maduro (~{age_years:.1f} anos) — janela clássica de valorização"
    if age_years <= 6.0:
        return 24, f"set antigo (~{age_years:.1f} anos) — fora de circulação, demanda residual"
    return 16, f"set muito antigo (~{age_years:.1f} anos) — vintage, líquido só em chase"


def compute_valorization(rarity: Optional[str],
                         set_release_date: Optional[str],
                         tcg_market_usd: float,
                         now: Optional[datetime] = None) -> tuple[int, str]:
    """Score de valorização médio/longo prazo (0-100) + nota explicativa.

    Honestidade: combina SÓ sinais que as APIs fornecem (rarity, idade do set,
    preço atual). NÃO usa preço histórico (a fonte não tem). É uma HEURÍSTICA
    de triagem, não previsão. Operador decide capital.
    """
    tier = classify_chase_tier(rarity)
    rarity_pts = {"TOP": 45, "MID": 30, "MODEST": 12, "BULK": 0}.get(tier, 8)
    rarity_note = {
        "TOP": "raridade top (chase forte)",
        "MID": "raridade média-alta (alt/ultra/rainbow)",
        "MODEST": "raridade modesta (holo/promo)",
        "BULK": "bulk (common/uncommon — baixo potencial)",
    }.get(tier, "raridade desconhecida")

    age_pts, age_note = _valorization_age_component(set_release_date, now=now)

    if tcg_market_usd < 15:
        price_pts, price_note = 4, "preço-âncora baixo (<$15)"
    elif tcg_market_usd < 40:
        price_pts, price_note = 10, "preço-âncora médio ($15-40)"
    elif tcg_market_usd < 100:
        price_pts, price_note = 16, "preço-âncora alto ($40-100)"
    else:
        price_pts, price_note = 20, "preço-âncora premium (>$100)"

    score = rarity_pts + age_pts + price_pts
    note = f"{rarity_note}; {age_note}; {price_note} (heurística — sem série histórica)"
    return score, note


def clean_collector_number(raw: str) -> str:
    """Normaliza um collector_number pra uso em query no pokemontcg.io.
    Entrada pode vir limpa ("007") ou com rarity concatenada ("SIR | 161/131").
    Saída: só o dígito, sem zeros à esquerda ("161", "7").
    """
    if not raw:
        return ""
    m = _NUMBER_IN_VERSION.search(raw)
    if m:
        return m.group(1).lstrip("0") or "0"
    # fallback: separador "/" sem total, ou só o número
    head = raw.split("/")[0].strip()
    # mantém só dígitos se houver sujeira ("TG01", "SWSH123")
    digits = "".join(c for c in head if c.isdigit())
    return digits.lstrip("0") or digits or head


# ══════════════════════════════════════════════════════════════════════
# MODELS — dataclasses em vez de dicts soltos. Por quê?
#   1. Autocompletar no IDE
#   2. Typos viram erro em vez de KeyError em produção
#   3. Serialização fácil para .xlsx/.json
# ══════════════════════════════════════════════════════════════════════
@dataclass
class Listing:
    """Uma oferta ativa em CardTrader (um seller + uma carta + 1 condição)."""
    product_id: int                  # ID do listing no CT
    blueprint_id: int                # ID da carta-molde
    card_name: str
    set_code: str                    # ex: "sv1", "sv3pt5"
    set_name: str
    collector_number: str            # ex: "125/165"
    condition: str                   # "Near Mint", "Slightly Played"...
    language: str                    # "en", "it"...
    price_cents: int                 # centavos no currency original do listing
    price_currency: str              # moeda original do listing ("BRL", "EUR", "USD")
    price_brl: float                 # preço normalizado em BRL (moeda interna)
    quantity: int
    foil: bool
    graded: bool
    seller_username: str
    seller_can_sell_via_hub: bool    # envio centralizado (+rápido, +seguro)
    seller_user_type: str            # "private", "professional", "zero_fee"
    cardtrader_url: str
    # v2.5 (2026-05-16): rarity persistida pra Chase Tier no postprocess v2+.
    # Fonte: blueprint CT (preferido) ou properties_hash. String vazia ""
    # se nem CT nem properties trouxerem (raro mas defensivo).
    rarity: str = ""

    @property
    def uid(self) -> str:
        """Identificador único para dedup: carta + seller + condição."""
        return f"{self.blueprint_id}:{self.seller_username}:{self.condition}"


@dataclass
class Opportunity:
    """Arbitragem detectada: listing CT com preço TCG conhecido e margem.
    Tudo em BRL (moeda interna do scanner). USD market é mantido como
    referência da fonte (TCGPlayer) pra auditoria.

    Campos `_real` / `live_*` / `markup_*` / `validation_status` são preenchidos
    pela v2.0 (validate_per_blueprint) — quando habilitada, faz uma chamada
    ao endpoint per-blueprint (preço com markup CT embutido) pros top N
    candidatos e recalcula margens reais.
    """
    listing: Listing
    tcg_market_usd: float            # preço market do TCGPlayer (fonte original)
    tcg_market_brl: float            # convertido USD→BRL
    ct_price_brl: float              # preço CT normalizado em BRL (per-expansion, RAW)
    margin_pct: float                # (tcg_brl - ct_brl) / tcg_brl
    margin_brl: float                # tcg_brl - ct_brl em R$ absolutos
    estimated_shipping_brl: float    # frete estimado em BRL (default 0 — modelo consolidação)
    net_margin_pct: float            # margem após dedução de frete (= margem bruta quando frete=0)
    scanned_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    # v2.0 — validação per-blueprint (preço REAL com markup CT)
    validation_status: str = "NOT_VALIDATED"
    live_price_brl: Optional[float] = None
    real_margin_pct: Optional[float] = None
    real_net_margin_pct: Optional[float] = None
    real_lucro_brl: Optional[float] = None
    markup_pct: Optional[float] = None
    markup_tier: Optional[str] = None
    # v2.7.1 (2026-05-18): URL TCGPlayer da carta exata matched no pokemontcg.io.
    # Vem de `card.tcgplayer.url` na response da API. Pra operador validar a
    # variante correta (ex: Lusamine 1st Place vs normal) antes de comprar.
    # Pode ser None quando: provider != pokemontcg, card sem entry TCGPlayer,
    # ou cobertura ruim da fonte.
    tcg_url: Optional[str] = None
    # v2.8 Layer 4 (2026-05-18): variante TCGPlayer usada no cálculo de margem.
    # Foil-aware disambiguation: listing CT foil → prioriza `holofoil`;
    # listing CT non-foil → prioriza `reverseHolofoil` (default landing TCG
    # pra 90% dos Holo Rare antigos). Sem essa coluna o operador não sabia
    # se o preço scanner vinha de Holofoil ($224 Pichu ecard1/22) ou Reverse
    # Holofoil ($50 mesma carta). Valores: "holofoil", "reverseHolofoil",
    # "normal", "unlimitedHolofoil", "" (provider != pokemontcg ou sem variant).
    price_variant_used: Optional[str] = None
    # v2.13 (2026-06-09): frente "chase cards" — tier de raridade (TOP/MID/
    # MODEST/BULK/""). Derivado da rarity oficial pokemontcg.io quando
    # disponível, senão da rarity do blueprint CT. Usado pelo --chase-only e
    # pela coluna Chase Tier no XLSX.
    chase_tier: str = ""
    # v2.13: frente "valorização" — score heurístico 0-100 (rarity + idade do
    # set + preço-âncora) + nota explicativa. NÃO é previsão (sem série
    # histórica na fonte); é triagem. Ver compute_valorization().
    valorization_score: int = 0
    valorization_note: str = ""
    # v2.13: data de lançamento do set (pokemontcg.io set.releaseDate), pra
    # auditoria da idade usada no score. "" quando a fonte não traz.
    set_release_date: str = ""
    # v2.14 (correção de resultado — candidato 4): FLAG de BAIXA CONFIANÇA de
    # variante. True quando o listing CT é NÃO-foil (foil=False) mas o único
    # preço disponível na pokemontcg.io foi de uma variante METÁLICA premium
    # (holofoil/unlimitedHolofoil) — i.e. nem `normal` nem `reverseHolofoil`
    # existiam. Nesse caso o preço de referência pode ser da variante ERRADA
    # (a holo costuma valer 3-10× a não-foil) e a margem fica inflada.
    # NÃO altera preço nem decisão — é só um sinal pro operador validar via
    # Link TCG (caso load-bearing: Gengar não-foil casando preço holo $1600).
    # reverseHolofoil NÃO dispara a flag de propósito: é o fallback INTENCIONAL
    # pra Holo Rare vintage (v2.10 Layer 4.1), onde non-foil↔reverse é correto.
    variant_low_confidence: bool = False


# ══════════════════════════════════════════════════════════════════════
# CACHE — SQLite local. Por quê?
#   - Evita hammer nas APIs (reduz custo em planos pagos, reduz banimento)
#   - Retomada em caso de crash
#   - Permite --dry-run para testar filtros sem re-buscar tudo
#   - 1 arquivo, zero setup (vs Redis/Postgres)
# ══════════════════════════════════════════════════════════════════════
class Cache:
    def __init__(self, db_path: Path = CACHE_DB):
        self.db = sqlite3.connect(db_path)
        self.db.execute("PRAGMA journal_mode=WAL")  # melhor concorrência
        self._init_schema()

    def _init_schema(self):
        self.db.executescript("""
        CREATE TABLE IF NOT EXISTS blueprints (
            ct_blueprint_id INTEGER PRIMARY KEY,
            ct_expansion_code TEXT,
            name TEXT,
            collector_number TEXT,
            raw_json TEXT,
            fetched_at TEXT
        );
        CREATE TABLE IF NOT EXISTS price_cache (
            key TEXT PRIMARY KEY,    -- "pokemontcg:<tcg_id>" ou similar
            market_usd REAL,
            low_usd REAL,
            mid_usd REAL,
            raw_json TEXT,
            fetched_at TEXT
        );
        CREATE TABLE IF NOT EXISTS fx_cache (
            pair TEXT PRIMARY KEY,   -- "USD_BRL", "EUR_BRL"
            rate REAL,
            fetched_at TEXT
        );
        CREATE TABLE IF NOT EXISTS blueprint_to_pricing (
            -- mapping CT blueprint → pricing provider product ID
            ct_blueprint_id INTEGER,
            provider TEXT,           -- "pokemontcg", "justtcg", "tcgplayer"
            provider_product_id TEXT,
            confidence REAL,         -- 0-1, quão confiante o match é
            matched_at TEXT,
            PRIMARY KEY (ct_blueprint_id, provider)
        );
        CREATE INDEX IF NOT EXISTS idx_bp_exp ON blueprints(ct_expansion_code);
        """)
        self.db.commit()

    def get_fx(self, pair: str) -> Optional[float]:
        row = self.db.execute(
            "SELECT rate, fetched_at FROM fx_cache WHERE pair = ?", (pair,)
        ).fetchone()
        if not row:
            return None
        rate, fetched_at = row
        if datetime.fromisoformat(fetched_at) < datetime.now() - FX_TTL:
            return None
        return rate

    def set_fx(self, pair: str, rate: float):
        self.db.execute(
            "INSERT OR REPLACE INTO fx_cache VALUES (?, ?, ?)",
            (pair, rate, datetime.now().isoformat()),
        )
        self.db.commit()

    def get_price(self, key: str) -> Optional[dict]:
        row = self.db.execute(
            "SELECT market_usd, low_usd, mid_usd, raw_json, fetched_at FROM price_cache WHERE key = ?",
            (key,),
        ).fetchone()
        if not row:
            return None
        market, low, mid, raw_json, fetched_at = row
        if datetime.fromisoformat(fetched_at) < datetime.now() - PRICE_TTL:
            return None
        # v2.7.1: extrai `tcgplayer.url` do raw_json pra hidratar Opportunity
        # mesmo em cache hit (sem refazer _search()). raw_json é o card dict
        # original do pokemontcg.io. Falha de parse = url None (não crash).
        # v2.8 Layer 4: também extrai `_variant_used` (gravado pelo provider
        # no card dict antes de json.dumps). Cache hit antigo (pre-v2.8) não
        # tem essa key → variant_used vira None, scanner persiste vazio.
        tcg_url = None
        variant_used = None
        # v2.13 (valorização): extrai metadados do card pokemontcg.io já cacheado
        # (set.releaseDate + rarity oficial). ZERO fetch extra — reusa o raw_json
        # que set_price já persistiu. Usado pelo score de valorização (idade do
        # set) e pela coluna Chase Tier (rarity oficial > rarity do CT).
        ptcg_rarity = None
        set_release_date = None
        try:
            if raw_json:
                raw = json.loads(raw_json)
                tcg_url = (raw.get("tcgplayer") or {}).get("url")
                variant_used = raw.get("_variant_used")
                ptcg_rarity = raw.get("rarity")
                set_release_date = (raw.get("set") or {}).get("releaseDate")
        except (json.JSONDecodeError, AttributeError, TypeError):
            tcg_url = None
            variant_used = None
        return {"market_usd": market, "low_usd": low, "mid_usd": mid,
                "tcg_url": tcg_url, "variant_used": variant_used,
                "ptcg_rarity": ptcg_rarity, "set_release_date": set_release_date}

    def set_price(self, key: str, market: float, low: float, mid: float, raw: dict):
        self.db.execute(
            "INSERT OR REPLACE INTO price_cache VALUES (?, ?, ?, ?, ?, ?)",
            (key, market, low, mid, json.dumps(raw), datetime.now().isoformat()),
        )
        self.db.commit()

    def get_mapping(self, blueprint_id: int, provider: str) -> Optional[str]:
        row = self.db.execute(
            "SELECT provider_product_id FROM blueprint_to_pricing WHERE ct_blueprint_id=? AND provider=?",
            (blueprint_id, provider),
        ).fetchone()
        return row[0] if row else None

    def set_mapping(self, blueprint_id: int, provider: str, product_id: str, confidence: float):
        self.db.execute(
            "INSERT OR REPLACE INTO blueprint_to_pricing VALUES (?, ?, ?, ?, ?)",
            (blueprint_id, provider, product_id, confidence, datetime.now().isoformat()),
        )
        self.db.commit()

    def clear_prices(self):
        """Limpa caches voláteis (preços TCG e FX). Útil pra --no-cache em
        cenários de arbitragem time-sensitive onde 24h TTL é tempo demais."""
        self.db.execute("DELETE FROM price_cache")
        self.db.execute("DELETE FROM fx_cache")
        self.db.commit()


# ══════════════════════════════════════════════════════════════════════
# FX — Frankfurter API (BCE, grátis, sem auth). Moeda interna = BRL.
# ══════════════════════════════════════════════════════════════════════
def _fetch_fx(pair: str, frm: str, to: str, cache: Cache) -> float:
    cached = cache.get_fx(pair)
    if cached is not None:
        return cached
    r = requests.get(f"{FX_BASE}?from={frm}&to={to}", timeout=TIMEOUT)
    r.raise_for_status()
    rate = r.json()["rates"][to]
    cache.set_fx(pair, rate)
    log.info(f"FX {frm}→{to} atualizado: {rate:.4f}")
    return rate

def get_usd_to_brl(cache: Cache) -> float:
    return _fetch_fx("USD_BRL", "USD", "BRL", cache)

def get_eur_to_brl(cache: Cache) -> float:
    return _fetch_fx("EUR_BRL", "EUR", "BRL", cache)


# ══════════════════════════════════════════════════════════════════════
# CARDTRADER CLIENT — API oficial, JSON REST, JWT Bearer.
# Rate limit (confirmado na docs):
#   - 10 req/s por endpoint
#   - 200 req / 10s no total
# Estratégia: delay de 0.15s entre requests (~6.7/s, folga de 30%)
# ══════════════════════════════════════════════════════════════════════
class CardTraderClient:
    def __init__(self, jwt: str, delay: float = REQUEST_DELAY_CT):
        if not jwt:
            raise ValueError("CT_JWT não configurado no .env")
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {jwt}",
            "Accept": "application/json",
            "User-Agent": "MasterBox-TCG-Scanner/1.0 (contato via cardtrader.com)",
        })
        self.delay = delay
        self._last_call = 0.0

    def _get(self, path: str, deadline_ts: Optional[float] = None, **params) -> dict | list:
        """GET com retry. Aceita `deadline_ts` (monotonic): se passar, retries
        e sleeps respeitam a deadline (v2.8 Codex H2 fix — antes uma chamada
        429 com Retry-After 60s podia segurar a set bem além do per-set-timeout).
        """
        elapsed = time.time() - self._last_call
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)
        url = f"{CT_BASE}{path}"

        def _deadline_exceeded() -> bool:
            return deadline_ts is not None and time.monotonic() > deadline_ts

        def _sleep_capped(wait_s: float) -> bool:
            """Sleep cap pra não passar da deadline. Retorna True se a deadline
            estourou OU se o backoff pedido (`wait_s`) não coube no tempo
            restante (não dá pra honrar a espera → não adianta tentar de novo).

            Fix 2026-06-15 (correção/robustez): antes o retorno era só
            `time.monotonic() > deadline_ts` APÓS dormir `min(wait_s, remaining)`.
            Quando `wait_s > remaining`, dormíamos exatamente até a deadline e o
            `>` voltava False por uma fração de segundo (sleep retorna um tiquinho
            cedo / arredondamento), o loop fazia `continue`, re-tentava e
            mascarava o timeout (ex.: 429 Retry-After 60s com 3s restantes
            "passava" em vez de abortar). Agora, se o backoff foi TRUNCADO pela
            deadline, sinalizamos estouro de forma determinística."""
            if deadline_ts is None:
                time.sleep(wait_s)
                return False
            remaining = deadline_ts - time.monotonic()
            if remaining <= 0:
                return True
            truncated = wait_s > remaining
            time.sleep(min(wait_s, remaining))
            # Estourou se: passou da deadline OU o backoff teve que ser cortado
            # (não conseguimos esperar o tempo necessário pra a próxima tentativa).
            return truncated or time.monotonic() > deadline_ts

        if _deadline_exceeded():
            raise TimeoutError(f"CT _get({path}): deadline already exceeded before request")

        # Retry com backoff 2/4/8s pra erros transientes (connection reset, 5xx,
        # rate limit). Descoberto em 2026-04-20 quando scan de 7 sets crashou
        # no 4º por ConnectionResetError no /blueprints/export.
        last_err: Optional[Exception] = None
        for attempt in range(3):
            # v2.8: usa min(TIMEOUT, tempo restante até deadline) pra request timeout
            req_timeout = TIMEOUT
            if deadline_ts is not None:
                remaining = max(1.0, deadline_ts - time.monotonic())
                req_timeout = min(TIMEOUT, remaining)
            try:
                r = self.session.get(url, params=params, timeout=req_timeout)
                self._last_call = time.time()
                if r.status_code == 429:
                    retry_after = int(r.headers.get("Retry-After", "5"))
                    log.warning(f"Rate limit CT, esperando {retry_after}s...")
                    if _sleep_capped(retry_after):
                        raise TimeoutError(
                            f"CT _get({path}): deadline excedida em sleep pós-429"
                        )
                    continue
                if 500 <= r.status_code < 600:
                    backoff = 2 ** attempt
                    log.warning(f"CT {r.status_code}, retry em {backoff}s...")
                    if _sleep_capped(backoff):
                        raise TimeoutError(
                            f"CT _get({path}): deadline excedida em sleep pós-5xx"
                        )
                    continue
                r.raise_for_status()
                return r.json()
            except (requests.ConnectionError, requests.Timeout) as e:
                last_err = e
                backoff = 2 ** attempt
                log.warning(f"CT {type(e).__name__}, retry em {backoff}s... ({attempt+1}/3)")
                if _sleep_capped(backoff):
                    raise TimeoutError(
                        f"CT _get({path}): deadline excedida em sleep pós-{type(e).__name__}"
                    ) from e
        if last_err:
            raise last_err
        raise requests.ConnectionError("CT: esgotou 3 tentativas")

    def list_expansions(self, game_id: int = CT_POKEMON_GAME_ID) -> list[dict]:
        """Lista todas as expansões (sets) de Pokemon no CT."""
        data = self._get("/expansions")
        # API retorna TODAS as expansões de TODOS os jogos; filtra por game_id
        return [e for e in data if e.get("game_id") == game_id]

    def list_blueprints(self, expansion_id: int, deadline_ts: Optional[float] = None) -> list[dict]:
        """Lista blueprints (cartas-molde) de uma expansão."""
        # Nota: /blueprints/export retorna tudo de uma vez; /blueprints paginado.
        # Usamos /blueprints/export por simplicidade.
        return self._get("/blueprints/export", deadline_ts=deadline_ts, expansion_id=expansion_id)

    def list_listings_by_blueprint(self, blueprint_id: int, language: str = "en",
                                    deadline_ts: Optional[float] = None) -> list[dict]:
        """Lista TODAS as ofertas ativas para um blueprint."""
        return self._get("/marketplace/products",
                         deadline_ts=deadline_ts,
                         blueprint_id=blueprint_id,
                         language=language)

    def list_listings_by_expansion(self, expansion_id: int, language: str = "en",
                                    deadline_ts: Optional[float] = None) -> list[dict]:
        """Lista TODAS as ofertas de uma expansão inteira (bem mais eficiente)."""
        data = self._get("/marketplace/products",
                         deadline_ts=deadline_ts,
                         expansion_id=expansion_id,
                         language=language)
        # API retorna dict[blueprint_id, list[listing]] — flatten para list
        if isinstance(data, dict):
            return [l for listings in data.values() for l in listings]
        return data


# ══════════════════════════════════════════════════════════════════════
# PRICING PROVIDERS — abstração. Trocar fonte = trocar 1 classe.
# Por que interface abstrata (Strategy Pattern)?
#   - TCGPlayer API oficial está fechada em 2026 → precisamos de plano B
#   - JustTCG cobra, pokemontcg.io é grátis → quer escolher dinamicamente
#   - Testabilidade: mock em testes sem chamar API real
# ══════════════════════════════════════════════════════════════════════
class PricingProvider(ABC):
    name: str = "base"
    # v2.7.1 (2026-05-18): URL TCGPlayer da última carta consultada com sucesso.
    # Provider sobrescreve em cada market_price_usd() bem-sucedido (set via
    # `self.last_tcg_url = ...`). scan_expansion lê após cada chamada e
    # propaga pra Opportunity.tcg_url. Reset pra None em cada chamada
    # (qualquer cache hit + lookup miss reseta) pra evitar carry-over entre
    # cards diferentes.
    last_tcg_url: Optional[str] = None
    # v2.8 Layer 4 (2026-05-18): variante TCGPlayer usada pra resolver o preço
    # (`holofoil`, `reverseHolofoil`, `normal`, `unlimitedHolofoil`). Provider
    # foil-aware sobrescreve a cada chamada. scan_expansion propaga pra
    # Opportunity.price_variant_used. Reset pra None igual `last_tcg_url`.
    last_variant_used: Optional[str] = None

    @abstractmethod
    def market_price_usd(self, card_name: str, set_code: str,
                         collector_number: str,
                         foil: bool = False,
                         rarity: Optional[str] = None) -> Optional[float]:
        """Retorna preço market em USD ou None se não achar.

        `foil`: indica se o listing CT é reverse-holo (pokemon_reverse=True
        em Pokémon, mtg_foil em MTG). Provider usa pra priorizar a variante
        correta (reverseHolofoil vs holofoil vs normal) no agregador
        pokemontcg.io (fix H2 2026-05-11; v2.18 2026-06-20).
        `rarity`: raridade do listing CT ("Holo Rare", "Common", ...). v2.18
        usa pra distinguir holo rare padrão (→ holofoil) de carta base
        (→ normal) quando foil=False.
        """
        ...


def _rarity_is_holo(*rarities: Optional[str]) -> bool:
    """True se QUALQUER rótulo de raridade indicar carta holográfica.

    Casa tanto a raridade do CardTrader ("Holo Rare") quanto a do
    pokemontcg.io ("Rare Holo", "Rare Holo EX", "Rare Holo Star",
    "Rare Holo LV.X", "Rare Shining", ...). Usado pela seleção de
    variante (v2.18) pra decidir se o preço de referência correto é
    `holofoil` (carta intrinsecamente holo) ou `normal` (carta base).

    NÃO casa "Reverse Holo" como raridade — no CardTrader o reverse é uma
    propriedade separada do listing (`pokemon_reverse`), não a raridade
    impressa; ele é tratado pelo sinal `foil`/reverse, não por aqui.
    """
    for r in rarities:
        if not r:
            continue
        rl = r.lower()
        # Guard defensivo: "reverse holo" como raridade (não existe no vocab CT
        # atual — reverse é o booleano pokemon_reverse) NÃO conta como holo aqui;
        # mantém o código fiel ao docstring se a fonte um dia mudar.
        if "reverse" in rl:
            continue
        if "holo" in rl or "shining" in rl:
            return True
    return False


class PokemonTcgIoProvider(PricingProvider):
    """
    pokemontcg.io v2 — GRÁTIS.
    Inclui campo `tcgplayer.prices.{variant}.market` (USD, TCGPlayer).
    Atualização diária. Limite: 20k req/dia com key grátis, 1k sem.

    Por que esse é o default? Porque a API oficial do TCGPlayer está
    fechada para novos devs, e esse agregador entrega o MESMO dado
    (preço market TCGPlayer) com zero custo.
    """
    name = "pokemontcg"

    def __init__(self, api_key: Optional[str], cache: Cache,
                 delay: float = REQUEST_DELAY_PRICING):
        self.session = requests.Session()
        if api_key:
            self.session.headers["X-Api-Key"] = api_key
        self.cache = cache
        self.delay = delay
        self._last_call = 0.0

    def _rate_limit(self):
        elapsed = time.time() - self._last_call
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)
        self._last_call = time.time()

    def _ptcg_get(self, params: dict):
        """GET na pokemontcg.io com retry bounded em erros TRANSIENTES
        (429 rate-limit, 5xx). Distingue:
          - 200            → retorna Response (caller parseia data/totalCount)
          - 429 / 5xx      → retry com backoff curto; se persistir, levanta
                             requests.Timeout (→ scan loop conta como
                             pricing_failure, NÃO como "miss" de cobertura)
          - 4xx ≠ 429      → retorna a Response (caller trata como "sem match";
                             é resposta legítima do servidor, não falha)

        Fix 2026-06-15 (correção/robustez — candidato 5): antes QUALQUER
        status != 200 (incluindo 429 e 5xx) virava `continue` → `_search`
        devolvia None → o loop de pricing contava como "miss" de cobertura.
        Consequências: (a) rate-limit/queda transitória da fonte sumia deals
        em silêncio; (b) o contador de "misses consecutivos" era envenenado e
        podia abortar sets bons com `no_coverage`; (c) inflava a taxa aparente
        de falha. Agora 429/5xx são tratados como falha transiente VISÍVEL
        (logada + contabilizada), não como ausência de carta.
        """
        last_status = None
        for attempt in range(3):
            self._rate_limit()
            r = self.session.get(f"{POKEMONTCG_BASE}/cards",
                                 params=params, timeout=TIMEOUT)
            if r.status_code == 200:
                return r
            last_status = r.status_code
            if r.status_code == 429 or 500 <= r.status_code < 600:
                # Transiente: respeita Retry-After (429) ou backoff exponencial.
                try:
                    retry_after = int(r.headers.get("Retry-After", "0"))
                except (TypeError, ValueError):
                    retry_after = 0
                backoff = retry_after if retry_after > 0 else 2 ** attempt
                # Cap defensivo: nunca dorme mais que TIMEOUT por tentativa.
                backoff = min(backoff, TIMEOUT)
                log.warning(
                    f"pokemontcg.io {r.status_code} (transiente), retry em "
                    f"{backoff}s... ({attempt + 1}/3)"
                )
                time.sleep(backoff)
                continue
            # 4xx ≠ 429 → resposta legítima do servidor; caller decide.
            return r
        # Esgotou retries num status transiente → falha VISÍVEL, não "miss".
        raise requests.Timeout(
            f"pokemontcg.io: status {last_status} persistente após 3 tentativas"
        )

    # v2.7 Layer 1.5 (bug-hunt 2026-05-18): alias map CT → pokemontcg.io.
    # CT usa códigos curtos derivados de TCGPlayer/comunidade; pokemontcg.io
    # usa codes oficiais Wizards/Nintendo da era. Sem alias, Layer 1 (strict
    # set match) rejeita TODA Jungle/Fossil/Neo/Gym/Wizards-era porque
    # ct `ju` ≠ ptcg `base2`. Estende set query no _search: além do código
    # CT, tentamos os aliases listados. Match em qualquer um aceita.
    #
    # Lista derivada do README + probes 2026-05-18. Não exhaustivo, foca
    # sets que apareceram no weekly v2.6 + vintage Wizards essenciais.
    SET_ALIAS_TO_PTCG = {
        # Wizards original era
        "ju": ["base2"],          # Jungle
        "fo": ["base3"],          # Fossil
        "b2": ["base4"],          # Base Set 2
        "tr": ["base5"],          # Team Rocket
        "g1": ["gym1"],           # Gym Heroes
        "g2": ["gym2"],           # Gym Challenge
        "n1": ["neo1"],           # Neo Genesis
        "n2": ["neo2"],           # Neo Discovery
        "n3": ["neo3"],           # Neo Revelation
        "n4": ["neo4"],           # Neo Destiny
        "lc": ["base6"],          # Legendary Collection
        # E-Card era
        "ex": ["ecard1"],         # Expedition Base Set
        "aq": ["ecard2"],         # Aquapolis
        "skg": ["ecard3"],        # Skyridge
        # EX series (Ruby & Sapphire era)
        "dr": ["ex3"],            # EX Dragon
        # Sun & Moon Promos
        "pupr": ["sm5"],          # Ultra Prism Promos → SM Ultra Prism
        # XY Promos
        "xybsp": ["xyp"],         # XY Black Star Promos
        # Sword & Shield era (v2.9 bug-hunt 2026-05-19):
        # CT uses 3-letter ptcgo-style codes; ptcg.io uses positional swsh*.
        # All EN main sets explicitly mapped to avoid Layer 1 strict-set
        # rejection. Verified via CT API (game_id=5) + pokemontcg.io /sets
        # 2026-05-19. Trainer Gallery splits (swsh9tg, swsh10tg, swsh11tg,
        # swsh12tg, swsh12pt5gg) NOT aliased here — TG cards have known
        # inflation bug (regex ^TG\d+ in postprocess routes to MANUAL).
        "ssh":   ["swsh1"],       # Sword & Shield base
        "rcl":   ["swsh2"],       # Rebel Clash
        "daa":   ["swsh3"],       # Darkness Ablaze
        "cpa":   ["swsh35"],      # Champion's Path
        "viv":   ["swsh4"],       # Vivid Voltage
        "shf":   ["swsh45"],      # Shining Fates
        "bst":   ["swsh5"],       # Battle Styles
        "cre":   ["swsh6"],       # Chilling Reign
        "evs":   ["swsh7"],       # Evolving Skies
        "fst":   ["swsh8"],       # Fusion Strike
        "brs":   ["swsh9"],       # Brilliant Stars
        "astr":  ["swsh10"],      # Astral Radiance (CT: astr, NOT asr)
        "pkmgo": ["pgo"],         # Pokémon GO (CT: pkmgo, NOT pgo)
        "lorg":  ["swsh11"],      # Lost Origin (CT: lorg; lot=Lost Thunder)
        "sit":   ["swsh12"],      # Silver Tempest
        "crz":   ["swsh12pt5"],   # Crown Zenith
        # Scarlet & Violet era (v2.9 expansion 2026-05-19):
        # Existing aliases (pre/twm/scr/ssp) retained. Added missing SV core
        # so Layer 1 doesn't drop matches for Paldea Evolved, Obsidian Flames,
        # 151, Paradox Rift, Paldean Fates, Temporal Forces, Shrouded Fable.
        # NOTE: CT codes sv6/sv7/sv8/sv9 point to JP-original sets (Mask of
        # Change, Stellar Miracle, Super Electric Breaker, Battle Partners) —
        # NOT aliased to EN ptcg.io ids on purpose. ptcg.io lacks JP coverage
        # and aliasing would produce wrong cards.
        "svi":   ["sv1"],         # Scarlet & Violet base
        "pal":   ["sv2"],         # Paldea Evolved
        "obf":   ["sv3"],         # Obsidian Flames
        "mew":   ["sv3pt5"],      # 151 (CT: mew)
        "par":   ["sv4"],         # Paradox Rift
        "paf":   ["sv4pt5"],      # Paldean Fates
        "tef":   ["sv5"],         # Temporal Forces
        "twm":   ["sv6"],         # Twilight Masquerade
        "sfa":   ["sv6pt5"],      # Shrouded Fable
        "scr":   ["sv7"],         # Stellar Crown
        "ssp":   ["sv8"],         # Surging Sparks
        "pre":   ["sv8pt5"],      # Prismatic Evolutions
        "jtg":   ["sv9"],         # Journey Together
        "dri":   ["sv10"],        # Destined Rivals
        "blk":   ["zsv10pt5"],    # Black Bolt
        "wht":   ["rsv10pt5"],    # White Flare
        # Mega Evolution era (v2.9 — first ME aliases added 2026-05-19):
        "meg":   ["me1"],         # Mega Evolution base
        "pfl":   ["me2"],         # Phantasmal Flames
        "asc":   ["me2pt5"],      # Ascended Heroes (sfa was timing out → asc same family)
        "por":   ["me3"],         # Perfect Order
        # === v2.10 BATCH (added 2026-05-19 from logs/alias_gaps_2026-05-19.md) ===
        # 110 high-confidence 1-to-1 aliases extracted from 558 rejected
        # (CT_set → api_set) pairs in weekly v2.9 scan log. Criterion: log shows
        # api_set mode with empty runner-ups OR runner-ups with hits ≤ 10% of
        # mode. Ambiguous codes (bs, c25, clb, crz, svpromo, etc.) deferred to
        # v2.11 case-by-case. xytkas/xytkos NOT included — log mapped them to
        # tk1a/tk1b but those are EX Trainer Kits, not XY (pokemontcg.io lacks
        # XY Trainer Kit coverage). EX series core sets dominate this batch.
        # Wizards / vintage era
        "wiz":   ["basep"],       # WotC Promos (Wizards Black Star)
        "nbsp":  ["np"],          # Nintendo Black Star Promos
        # E-Card era (extends existing aq/ex/skg)
        # (nothing new — ecard1/2/3 already covered by ex/aq/skg)
        # EX series (Ruby & Sapphire onwards)
        "rs":    ["ex1"],         # Ruby & Sapphire
        "ss":    ["ex2"],         # Sandstorm
        "exma":  ["ex4"],         # Team Magma vs Team Aqua
        "hl":    ["ex5"],         # Hidden Legends
        "rg":    ["ex6"],         # FireRed & LeafGreen
        "trr":   ["ex7"],         # Team Rocket Returns
        "dx":    ["ex8"],         # Deoxys
        "em":    ["ex9"],         # Emerald
        "uf":    ["ex10"],        # Unseen Forces
        "ds":    ["ex11"],        # Delta Species
        "lm":    ["ex12"],        # Legend Maker
        "hp":    ["ex13"],        # Holon Phantoms
        "cg":    ["ex14"],        # Crystal Guardians
        "df":    ["ex15"],        # Dragon Frontiers
        "pk":    ["ex16"],        # Power Keepers
        # Diamond & Pearl era
        "mt":    ["dp2"],         # Mysterious Treasures
        "sw":    ["dp3"],         # Secret Wonders
        "ge":    ["dp4"],         # Great Encounters
        "wcd2010": ["dp6"],       # World Championship Deck 2010 — Legends Awakened
        "sft":   ["dp7"],         # Stormfront
        # Platinum
        "pl":    ["pl1"],         # Platinum base
        "rr":    ["pl2"],         # Rising Rivals
        "prr":   ["pl2"],         # Pre-release Rising Rivals
        "sv":    ["pl3"],         # Supreme Victors (CT 'sv' — NOT Scarlet & Violet!)
        # HeartGold & SoulSilver
        "hgs":   ["hgss1"],       # HeartGold & SoulSilver base
        "ul":    ["hgss2"],       # Unleashed
        "und":   ["hgss3"],       # Undaunted
        "tri":   ["hgss4"],       # Triumphant
        "clo":   ["col1"],        # Call of Legends
        # Black & White era
        "blw":   ["bw1"],         # Black & White base
        "epo":   ["bw2"],         # Emerging Powers
        "nvi":   ["bw3"],         # Noble Victories
        "nxd":   ["bw4"],         # Next Destinies
        "dex":   ["bw5"],         # Dark Explorers
        "drx":   ["bw6"],         # Dragons Exalted
        "bcr":   ["bw7"],         # Boundaries Crossed
        "pbcr":  ["bw7"],         # Promo BCR
        "pls":   ["bw8"],         # Plasma Storm
        "ppls":  ["bw8"],         # Promo PLS
        "plf":   ["bw9"],         # Plasma Freeze
        "pplf":  ["bw9"],         # Promo PLF
        "plb":   ["bw10"],        # Plasma Blast
        "pplb":  ["bw10"],        # Promo PLB
        "ltr":   ["bw11"],        # Legendary Treasures
        "pltr":  ["bw11"],        # Promo LTR
        "drv":   ["dv1"],         # Dragon Vault
        # XY era
        "kss":   ["xy0"],         # Kalos Starter Set
        "xy-en": ["xy1"],         # XY base (EN)
        "flf":   ["xy2"],         # Flashfire
        "ffi":   ["xy3"],         # Furious Fists
        "phf":   ["xy4"],         # Phantom Forces
        "prc":   ["xy5"],         # Primal Clash
        "ros":   ["xy6"],         # Roaring Skies
        "aor":   ["xy7"],         # Ancient Origins
        "bkt":   ["xy8"],         # BREAKthrough
        "bkp":   ["xy9"],         # BREAKpoint
        "pbkp":  ["xy9"],         # Promo BKP
        "fco":   ["xy10"],        # Fates Collide
        "sts":   ["xy11"],        # Steam Siege
        "psts":  ["xy11"],        # Promo STS
        "evo":   ["xy12"],        # Evolutions
        "gen":   ["g1"],          # Generations
        "pgen":  ["g1"],          # Promo Generations
        "dcr":   ["dc1"],         # Double Crisis
        "det":   ["det1"],        # Detective Pikachu (movie set)
        # Sun & Moon era
        "sum":   ["sm1"],         # Sun & Moon base
        "gri":   ["sm2"],         # Guardians Rising
        "bus":   ["sm3"],         # Burning Shadows
        "pbus":  ["sm3"],         # Promo BUS
        "cinv":  ["sm4"],         # Crimson Invasion
        "upr":   ["sm5"],         # Ultra Prism
        "fli":   ["sm6"],         # Forbidden Light
        "ces":   ["sm7"],         # Celestial Storm
        "pces":  ["sm7"],         # Promo CES
        "lot":   ["sm8"],         # Lost Thunder
        "teu":   ["sm9"],         # Team Up
        "unb":   ["sm10"],        # Unbroken Bonds
        "unm":   ["sm11"],        # Unified Minds
        "cec":   ["sm12"],        # Cosmic Eclipse
        "pcec":  ["sm12"],        # Promo CEC
        "slg":   ["sm35"],        # Shining Legends
        "drm":   ["sm75"],        # Dragon Majesty
        "hif":   ["sm115"],       # Hidden Fates
        "pdp":   ["dp1"],         # Promo DP
        # McDonald's promos (trivial 1-to-1)
        "mc11":  ["mcd11"],
        "mc12":  ["mcd12"],
        "mc14":  ["mcd14"],
        "mc15":  ["mcd15"],
        "mc17":  ["mcd17"],
        "mc18":  ["mcd18"],
        "mc19":  ["mcd19"],
        "mc21":  ["mcd21"],
        # World Championship Decks (only mode-strong ones; ambiguous ones skipped)
        "wcd2005": ["ex7"],       # WCD 2005 — Team Rocket Returns
        "wcd2008": ["ex14"],      # WCD 2008 — Crystal Guardians
        "wcd2023": ["swsh9"],     # WCD 2023 — Brilliant Stars
        # Battle Arena / SV-era promo lines (mode-strong)
        "ba-20":   ["dp3"],       # Battle Arena 2020
        "ba-2024": ["svp"],       # Battle Arena 2024
        # Mega Evolution promo lines (m-blk, m-wht — mode-strong)
        "m-blk":   ["zsv10pt5"],  # Mega Black promo line
        "m-wht":   ["rsv10pt5"],  # Mega White promo line
        "p-blk":   ["zsv10pt5"],  # Promo Mega Black
        "p-wht":   ["rsv10pt5"],  # Promo Mega White
    }

    def _search(self, card_name: str, set_code: str, number: str) -> Optional[dict]:
        """Busca carta e retorna o primeiro match.
        Estratégia em cascata: primeiro tenta com set.id (mais específico, menor
        risco de falso-match entre reprints). Se retornar 0, tenta sem set.id
        (set codes divergem entre CT e pokemontcg.io — ex: CT `pre` = ptcg
        `sv8pt5`). Na segunda tentativa fica nome+número, que é razoavelmente
        único pra singles.

        v2.7 (bug-hunt 2026-05-18 Layer 1): strict set match no fallback.
        Antes — quando `set.id:X` retornava 0, o fallback aceitava `totalCount==1`
        sem checar set retornado. Isso fez Blastoise CT `clb` matchear
        `base5-3` (Dark Blastoise Team Rocket, 1stEdition $383) e Mewtwo CT
        `xybsp` matchear `rs:101/109` (Mewtwo ex Ruby & Sapphire $269) em vez
        de `xyp:XY101` ($25). Custou ~96% de falsos positivos no weekly v2.6.
        Fix: no fallback, exigir `card.set.id == set_code` (case-insensitive).
        Mismatch → log INFO + skip pra próximo candidato. Trade-off: alguns
        sets perdem cobertura (ex pupr→sm5 = Lusamine Promos Ultra Prism);
        compensado por UNSUPPORTED_SETS no postprocess (Layer 3).
        """
        num_clean = clean_collector_number(number)
        safe_name = card_name.replace('"', '\\"')
        base_q = f'name:"{safe_name}"'
        if num_clean:
            base_q += f" number:{num_clean}"

        # v2.7 Layer 1.5: monta lista de set_ids aceitos (canonical + aliases).
        # `expected_sets` é o conjunto de set.ids que satisfazem "set match".
        # Inclui o CT code original + qualquer alias mapeado em
        # SET_ALIAS_TO_PTCG. Sem alias → só o code CT (comportamento Layer 1
        # puro). Com alias → varia: pra `ju` aceita {`ju`, `base2`}.
        ct_code = (set_code or "").lower()
        expected_sets: set[str] = {ct_code} if ct_code else set()
        for alias in self.SET_ALIAS_TO_PTCG.get(ct_code, []):
            expected_sets.add(alias.lower())

        # Cada tupla: (query, strict_set_check).
        # Tentamos uma query por set candidato (CT code + cada alias).
        # strict_set_check só é necessário no fallback sem set.id.
        queries: list[tuple[str, bool]] = []
        for sid in sorted(expected_sets):  # determinism
            queries.append((f'{base_q} set.id:{sid}', False))
        queries.append((base_q, True))

        for q, strict_set_check in queries:
            # _ptcg_get faz retry em 429/5xx e LEVANTA requests.Timeout se
            # persistir (→ pricing_failure no scan loop, não "miss"). Erros 4xx
            # legítimos (≠429) voltam como Response e viram "sem match" abaixo.
            r = self._ptcg_get({"q": q, "pageSize": 5})
            if r.status_code != 200:
                log.debug(f"pokemontcg.io erro {r.status_code} para {q}")
                continue
            resp = r.json()
            results = resp.get("data", [])
            total = resp.get("totalCount", len(results))
            if not results:
                continue
            if strict_set_check:
                # v2.7 Layer 1: rejeita match se set retornado ∉ expected_sets.
                # Não basta totalCount==1 — precisa bater algum set esperado.
                for cand in results:
                    cand_set = (cand.get("set") or {}).get("id", "").lower()
                    if cand_set in expected_sets:
                        return cand
                # Nenhum candidato bate o set_code do CT ou alias
                first_set = (results[0].get("set") or {}).get("id", "")
                log.info(
                    f"set mismatch rejected: CT_set={set_code} (esperado {sorted(expected_sets)}) "
                    f"≠ api_set={first_set} ({total} candidatos) — '{base_q}'"
                )
                continue
            return results[0]
        return None

    def market_price_usd(self, card_name: str, set_code: str,
                         collector_number: str,
                         foil: bool = False,
                         rarity: Optional[str] = None) -> Optional[float]:
        # v2.7.1: reset last_tcg_url no início de cada chamada pra evitar
        # carry-over do card anterior caso este falhe sem set explícito.
        # v2.8 Layer 4: reset last_variant_used também.
        self.last_tcg_url = None
        self.last_variant_used = None
        # v2.13 (valorização): metadados pra score + chase tier. Reset a cada
        # chamada pra evitar carry-over do card anterior.
        self.last_ptcg_rarity = None
        self.last_set_release_date = None
        # Cache inclui foil pra evitar colisão entre versão normal e RH
        # do mesmo card (2026-05-11 H2 fix).
        cache_key = f"pokemontcg:{set_code}:{collector_number}:{card_name}:foil={foil}"
        cached = self.cache.get_price(cache_key)
        if cached:
            self.last_tcg_url = cached.get("tcg_url")
            self.last_variant_used = cached.get("variant_used")
            self.last_ptcg_rarity = cached.get("ptcg_rarity")
            self.last_set_release_date = cached.get("set_release_date")
            return cached["market_usd"]

        card = self._search(card_name, set_code, collector_number)
        if not card:
            return None
        # v2.7.1: captura tcgplayer.url da response (pode ser ausente em alguns
        # cards — set defensivo). Setado ANTES de validações restantes pra
        # garantir que sempre que retornarmos um preço, a url esteja disponível.
        self.last_tcg_url = (card.get("tcgplayer") or {}).get("url")
        # v2.13 (valorização): rarity oficial pokemontcg.io + data de lançamento
        # do set. Usados pelo score de valorização e pela coluna Chase Tier.
        self.last_ptcg_rarity = card.get("rarity")
        self.last_set_release_date = (card.get("set") or {}).get("releaseDate")

        # M5 fix (2026-05-12): instrumentação supranumerários.
        # Detecta quando o match retornado é SIR/SAR/HR/Gold/Rainbow (collector
        # number > printedTotal do set). Não muda comportamento — só sinaliza
        # pra operador investigar possíveis falso positivo de pricing inflado.
        # Padrão observado no MYP scanner irmão: SIR aparecendo como "Comum"
        # gera margem >100% falsa. CT pode ter caso análogo.
        try:
            card_num_int = int("".join(c for c in str(card.get("number", "")) if c.isdigit()) or "0")
            printed_total = int((card.get("set") or {}).get("printedTotal") or 0)
        except (TypeError, ValueError):
            card_num_int, printed_total = 0, 0
        if printed_total and card_num_int > printed_total:
            log.debug(
                f"supranumerary match: {card.get('name')} #{card_num_int} > "
                f"set.printedTotal={printed_total} "
                f"(set={card.get('set', {}).get('id')}, rarity={card.get('rarity')}). "
                f"Possível SIR/SAR/HR — validar match contra rarity do listing CT."
            )

        tcg = card.get("tcgplayer", {}).get("prices", {})
        if not tcg:
            return None

        # v2.18 (2026-06-20): seleção de variante CIENTE de raridade + reverse.
        #
        # PROBLEMA-RAIZ descoberto: pra cartas Pokémon, o sinal `foil` SEMPRE
        # chegava False — o parse do listing lia `props["mtg_foil"]/["foil"]`
        # (campos MTG, inexistentes em Pokémon) e IGNORAVA `pokemon_reverse`.
        # Logo a Holo Rare PADRÃO (pokemon_reverse=False) caía no ramo foil=False
        # e, sem `normal` na TCGPlayer (holo rare não tem printagem normal),
        # escorregava pra `reverseHolofoil` — variante reverse, mais rara e CARA.
        # Inflação sistemática em TODA holo rare vintage:
        #   Gengar Legendary Collection: holofoil $146.89 vs reverseHolofoil
        #     $1599.99 (10×!) — exatamente o "Gengar não-foil $1600" load-bearing.
        #   Claydol EX Hidden Legends: holofoil $14.47 vs reverse $44.95 (+210%).
        #   Shiftry EX Hidden Legends: holofoil $19.92 vs reverse $42.95 (+116%).
        # O fix v2.10 (normal-first) só corrigiu NÃO-holo (que têm `normal`);
        # holo rares continuaram inflando.
        #
        # CONSERTO: (1) o parse passa a ler `pokemon_reverse` → reverse holo
        #   chega como foil=True. (2) Aqui a prioridade vira ciente de raridade:
        #   - foil=True (reverse holo)      → reverseHolofoil primeiro.
        #   - foil=False + raridade holo    → holofoil/unlimited (printagem holo
        #       base); reverseHolofoil POR ÚLTIMO (nunca é o que o seller vende).
        #   - foil=False + não-holo         → normal; reverse por último.
        #   - foil=None (legado/direto)     → v2.7 Layer 2.
        # Empiricamente verificado no data model do CardTrader (2026-06-20):
        # `pokemon_rarity="Holo Rare"` + `pokemon_reverse=false` = holo padrão
        # (→ holofoil); `pokemon_reverse=true` = reverse (→ reverseHolofoil).
        #
        # EXCLUIR sempre: `1stEditionHolofoil`, `1stEditionNormal` (variantes
        # raras, inflam 3-10x; operador não vende 1st Ed).
        EXCLUDED = {"1stEditionHolofoil", "1stEditionNormal"}
        is_holo = _rarity_is_holo(rarity, card.get("rarity"))
        if foil is True:
            # Listing reverse-holo (pokemon_reverse=True) → reverseHolofoil.
            PRIORITY = ["reverseHolofoil", "holofoil", "unlimitedHolofoil", "normal"]
        elif foil is False:
            if is_holo:
                # Holo Rare padrão (não-reverse) → printagem holo base.
                PRIORITY = ["holofoil", "unlimitedHolofoil", "normal", "reverseHolofoil"]
            else:
                # Carta base não-holo → normal; reverse por último.
                PRIORITY = ["normal", "holofoil", "unlimitedHolofoil", "reverseHolofoil"]
        else:
            # Foil-flag ausente (cache miss antigo ou provider chamado direto).
            # Preserve v2.7 Layer 2 priority.
            PRIORITY = ["holofoil", "normal", "reverseHolofoil", "unlimitedHolofoil"]

        chosen = None
        chosen_variant = None
        for variant in PRIORITY:
            if variant in EXCLUDED:
                continue
            if variant in tcg and tcg[variant].get("market"):
                chosen = tcg[variant]
                chosen_variant = variant
                break
        if not chosen:
            # Nenhuma variante canônica → return None. Pre-fix tinha fallback
            # "primeira variante disponível" que pegava 1stEdition em vintage.
            available = [v for v in tcg.keys() if v not in EXCLUDED]
            log.debug(
                f"no canonical variant for {card_name} ({set_code}/{collector_number}): "
                f"available={list(tcg.keys())} canonical_after_exclusion={available} foil={foil}"
            )
            return None

        log.debug(
            f"variant chosen: {chosen_variant} for {card_name} "
            f"({set_code}/{collector_number}) foil={foil} — market=${chosen.get('market')}"
        )
        # v2.8 Layer 4: expõe variante usada pra Opportunity.price_variant_used.
        self.last_variant_used = chosen_variant
        market = chosen.get("market") or 0.0
        low = chosen.get("low") or 0.0
        mid = chosen.get("mid") or 0.0
        # v2.8 Layer 4: persiste variante no cache (raw payload tem o card dict
        # completo, mas extrair na leitura significa parsear JSON sempre.
        # Anotar `_variant_used` direto no card antes de json.dumps é O(1)).
        try:
            card["_variant_used"] = chosen_variant
        except (TypeError, AttributeError):
            pass
        self.cache.set_price(cache_key, market, low, mid, card)
        return market


class JustTcgProvider(PricingProvider):
    """
    JustTCG — PAGO (~$19-49/mês). Atualiza várias vezes ao dia.
    Stub: habilita quando/se você assinar. Troca no CLI: --provider justtcg
    """
    name = "justtcg"

    def __init__(self, api_key: Optional[str], cache: Cache,
                 delay: float = REQUEST_DELAY_PRICING):
        if not api_key:
            raise ValueError("JUSTTCG_API_KEY não configurado")
        self.session = requests.Session()
        self.session.headers["X-API-Key"] = api_key
        self.cache = cache
        self.delay = delay

    def market_price_usd(self, card_name: str, set_code: str,
                         collector_number: str,
                         foil: bool = False,
                         rarity: Optional[str] = None) -> Optional[float]:
        # TODO: implementar quando assinar JustTCG.
        # Endpoint: GET /cards?name=X&set=Y → prices.tcgplayer.marketPrice
        raise NotImplementedError("JustTCG provider: implementar ao assinar")


class TcgPlayerOfficialProvider(PricingProvider):
    """
    TCGPlayer API OFICIAL — FECHADA para novos devs em 2026.
    Stub: se você conseguir credenciais no futuro, só implementar aqui.
    """
    name = "tcgplayer"

    def market_price_usd(self, card_name: str, set_code: str,
                         collector_number: str,
                         foil: bool = False,
                         rarity: Optional[str] = None) -> Optional[float]:
        raise NotImplementedError(
            "TCGPlayer API oficial está fechada para novos devs em 2026. "
            "Quando obter credenciais (OAuth2 com public/private key), "
            "implementar aqui: POST /token → GET /pricing/product/{id}"
        )


PROVIDERS = {
    "pokemontcg": PokemonTcgIoProvider,
    "justtcg": JustTcgProvider,
    "tcgplayer": TcgPlayerOfficialProvider,
}


# ══════════════════════════════════════════════════════════════════════
# SCANNER — orquestração do pipeline de arbitragem.
# Fluxo:
#   1. Buscar expansões Pokemon no CT
#   2. Filtrar só as que interessam (via config ou --sets)
#   3. Para cada expansão, puxar TODAS listings EN
#   4. Aplicar filtros (NM, não-graded, preço mín)
#   5. Para cada listing, buscar preço TCG via pricing provider
#   6. Calcular margem bruta e líquida (default frete=0 modelo consolidação Hub)
#   7. Dedup: mantém só melhor oferta por (carta + condição)
# ══════════════════════════════════════════════════════════════════════
# --- v2.4 skip-list helpers (module-level) ---
#
# v2.7 (bug-hunt Codex H1, 2026-05-17): mutações agora são thread+process-safe.
# Antes: load → mutate → write_text() (read-modify-write sem lock). Dois scanners
# paralelos podiam corromper estado (último writer vencia). Pior: load_skip_list
# silenciosamente devolvia {} em qualquer erro de leitura/parse, então quem
# escrevesse depois apagava todo histórico real. Agora:
#   1. portalocker.Lock acquire exclusive antes de qualquer mutação
#   2. read+parse acontece DENTRO do lock (evita TOCTOU)
#   3. write é atomic: escreve em tempfile + os.replace() pra evitar truncate
#      parcial caso processo morra no meio do write
#   4. load_skip_list (read-only) usa shared lock e PROPAGA erro de parse
#      (não engole) — quem chama decide se aborta ou reseta.
# Sentinela `_SKIP_LIST_LOCK_TIMEOUT` evita deadlock se outro processo crashar
# segurando o lock.

_SKIP_LIST_LOCK_TIMEOUT = 30  # seg — operação é trivial, 30s é generoso


# ── Run-guard contra instâncias concorrentes (v2.14 robustez) ──────────────
# Incidente documentado (memória ct_scan_long_run_detached): dois scanners no
# MESMO state-dir disputam o cache.db SQLite + skip-list. A contenção de lock
# derruba o pricing pra ~9s/listing → o per-set-timeout estoura por falsa lentidão
# → a skip-list é ENVENENADA com sets bons. O guard segura um lock EXCLUSIVE
# não-bloqueante num lockfile do state-dir DURANTE TODO o scan. Se outro
# processo já o segura, recusamos iniciar (a menos de --allow-concurrent).
def acquire_run_guard(state_dir: Path):
    """Tenta adquirir o run-guard exclusivo do state-dir. Retorna o handle
    portalocker.Lock JÁ ADQUIRIDO (caller deve mantê-lo vivo até o fim do scan
    e dar .release() no final), ou None se já houver outra instância.

    Não levanta — o caller decide abortar (default) ou seguir (--allow-concurrent).
    """
    guard_path = state_dir / "scanner_run.lock"
    lock = portalocker.Lock(
        str(guard_path),
        mode="a+",
        timeout=0,
        flags=portalocker.LOCK_EX | portalocker.LOCK_NB,
    )
    try:
        lock.acquire()
    except portalocker.exceptions.LockException:
        return None
    # Grava PID + timestamp pra diagnóstico de quem segura o lock.
    try:
        fh = lock.fh
        fh.seek(0)
        fh.truncate()
        fh.write(f"pid={os.getpid()} since={datetime.now().astimezone().isoformat(timespec='seconds')}\n")
        fh.flush()
    except Exception:
        # Lock é o que importa; o conteúdo informativo é best-effort.
        pass
    return lock

# PR-F (2026-05-28): TTL pra entries TRANSIENTES da skip-list. Sets pulados por
# timeout (per_set_timeout_/ct_*_timeout_) ou mass-pricing-failure costumam ser
# stalls passageiros (rede ruim, 429 burst, cobertura pokemontcg.io temporária).
# Após 7 dias, devem ser re-tentados automaticamente. Já erros estruturais
# (unexpected_error_*) ficam PERMANENTES — provavelmente bug de schema/dados.
_SKIP_LIST_TTL_DAYS = 7
# Prefixos de reason considerados TRANSIENTES (elegíveis a TTL/expiração).
_SKIP_LIST_TRANSIENT_PREFIXES = (
    "per_set_timeout_",
    "mass_pricing_failure_",
    "ct_blueprints_timeout_",
    "ct_listings_timeout_",
)


def _skip_reason_str(entry) -> str:
    """Normaliza uma entry de reason (str legada OU dict {reason, added_at})
    para a string de motivo. Retrocompat com skip-lists pré-PR-F."""
    if isinstance(entry, dict):
        return str(entry.get("reason", ""))
    return str(entry) if entry is not None else ""


def _skip_entry_is_expired(entry, now: Optional[datetime] = None) -> bool:
    """PR-F: True se a entry é TRANSIENTE (prefixo conhecido) E tem added_at
    mais velho que _SKIP_LIST_TTL_DAYS.

    Regras de retrocompat:
      - entry legada (str pura, sem added_at) → NUNCA expira (permanente).
      - dict sem added_at (ou added_at inválido) → NUNCA expira (conservador).
      - reason não-transiente (ex. unexpected_error_*) → NUNCA expira.
    """
    reason = _skip_reason_str(entry)
    if not reason.startswith(_SKIP_LIST_TRANSIENT_PREFIXES):
        return False
    if not isinstance(entry, dict):
        # Legada (str) → sem timestamp → trata como permanente (não dropa).
        return False
    added_at = entry.get("added_at")
    if not added_at:
        return False
    try:
        # Formato gravado: "%Y-%m-%dT%H:%M:%SZ" (UTC). Parse robusto via fromisoformat.
        ts = datetime.fromisoformat(str(added_at).replace("Z", "+00:00"))
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
    except (ValueError, TypeError):
        # added_at corrompido → conservador: não dropar.
        return False
    now = now or datetime.now(timezone.utc)
    return (now - ts) > timedelta(days=_SKIP_LIST_TTL_DAYS)


def _skip_list_lock_path() -> Path:
    """Lock file separado (.lock sidecar) — portalocker em Windows não permite
    bloquear arquivo ainda inexistente; lock dedicado é mais robusto."""
    return SKIP_LIST_FILE.with_suffix(SKIP_LIST_FILE.suffix + ".lock")


def _empty_skip_payload() -> dict:
    return {"skipped": [], "reasons": {}, "updated_at": None}


def _read_skip_file_locked() -> dict:
    """Lê + parseia skip-list assumindo lock já adquirido pelo caller.
    Levanta exception em erro de parse (NÃO engole)."""
    if not SKIP_LIST_FILE.exists():
        return _empty_skip_payload()
    # utf-8-sig: tolera BOM UTF-8 (EF BB BF) que PS 5.1 `Set-Content -Encoding utf8`
    # e vários editores prepend ao reescrever a skip-list. `json.loads` puro rejeita
    # BOM com "Unexpected UTF-8 BOM" → crash pré-set-1 (já quebrou um daily 2026-05-29).
    # Comportamento idêntico ao utf-8 quando não há BOM.
    raw = SKIP_LIST_FILE.read_text(encoding="utf-8-sig")
    if not raw.strip():
        # Arquivo vazio (caso edge: write parcial / disk full). Tratado como vazio,
        # mas loga WARNING — operador deve saber.
        log.warning(f"Skip-list vazia ({SKIP_LIST_FILE}). Tratando como reset.")
        return _empty_skip_payload()
    return json.loads(raw)  # propaga JSONDecodeError


def _atomic_write_skip_file(payload: dict) -> None:
    """Escreve skip-list de forma atomica via tempfile + os.replace.
    Lock deve já estar adquirido pelo caller.

    Em Windows o `os.replace` pode levantar PermissionError[WinError 5] quando
    o destino esta sendo lido por outro processo (file sharing semantics,
    Google Drive watcher, antivirus). Retry curto com backoff cobre isso —
    pq estamos sob lock exclusive do `.lock` sidecar, o conflito eh transient
    (a janela termina assim que o reader concorrente fecha o handle)."""
    tmp_path = SKIP_LIST_FILE.with_suffix(SKIP_LIST_FILE.suffix + ".tmp")
    # Use PID + thread no nome do tmp pra evitar colisao entre workers paralelos
    # que estejam escapando do lock (defense-in-depth — não deveria acontecer)
    unique_tmp = tmp_path.with_suffix(f".tmp.{os.getpid()}")
    unique_tmp.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    # Retry os.replace pra contornar PermissionError transient no Windows
    last_err: Optional[Exception] = None
    for attempt in range(10):
        try:
            os.replace(unique_tmp, SKIP_LIST_FILE)
            return
        except PermissionError as e:
            last_err = e
            time.sleep(0.1 * (attempt + 1))  # backoff 0.1, 0.2, 0.3, ... 1.0s
    # Esgotou retries — tenta cleanup do tmp antes de propagar
    try:
        if unique_tmp.exists():
            unique_tmp.unlink()
    except Exception:
        pass
    raise last_err  # type: ignore[misc]


def load_skip_list() -> dict:
    """Lê scanner_skip_list.json (com shared lock pra coerência).
    Em erro de read/parse, levanta exception — caller decide.
    Retorna dict com 'skipped' (list[str]), 'reasons' (dict), 'updated_at'."""
    lock_path = _skip_list_lock_path()
    try:
        # SHARED lock — múltiplos readers OK, conflita com writer exclusive.
        # fail_when_locked=True habilita timeout polling (vs blocking infinito).
        with portalocker.Lock(
            lock_path,
            mode="a+",
            flags=portalocker.LOCK_SH | portalocker.LOCK_NB,
            timeout=_SKIP_LIST_LOCK_TIMEOUT,
            fail_when_locked=False,
            check_interval=0.1,
        ):
            return _read_skip_file_locked()
    except portalocker.exceptions.LockException as e:
        # Não consegue lock em 30s → outro processo travou. Log + propaga.
        log.error(f"Skip-list lock timeout ({lock_path}): {e}")
        raise
    except json.JSONDecodeError as e:
        log.error(f"Skip-list JSON inválido ({SKIP_LIST_FILE}): {e}")
        raise


def add_to_skip_list(exp_code: str, reason: str) -> None:
    """Adiciona um exp_code à skip-list com motivo + timestamp. Idempotente.
    Thread+process-safe (exclusive lock + atomic write)."""
    lock_path = _skip_list_lock_path()
    try:
        with portalocker.Lock(
            lock_path,
            mode="a+",
            flags=portalocker.LOCK_EX | portalocker.LOCK_NB,  # EXCLUSIVE não-bloqueante p/ polling
            timeout=_SKIP_LIST_LOCK_TIMEOUT,
            fail_when_locked=False,
            check_interval=0.1,
        ):
            # Read CURRENT state under lock (não confiar em estado pre-lock)
            try:
                data = _read_skip_file_locked()
            except json.JSONDecodeError as e:
                # Corruption detectada. Fail loud — operador investiga.
                # NÃO sobrescreve cegamente, isso apagaria histórico.
                log.error(
                    f"Skip-list corrupted ({SKIP_LIST_FILE}): {e}. "
                    f"Abortando add({exp_code}). Renomeie/delete o arquivo se ok."
                )
                raise
            skipped = set(data.get("skipped", []))
            skipped.add(exp_code)
            reasons = data.get("reasons") or {}
            # PR-F (2026-05-28): reason agora é {reason, added_at} (era str pura)
            # pra suportar TTL na leitura. Entries legadas (str) seguem aceitas
            # no read (tratadas como permanentes).
            reasons[exp_code] = {
                "reason": reason,
                "added_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            }
            payload = {
                "skipped": sorted(skipped),
                "reasons": reasons,
                "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            }
            _atomic_write_skip_file(payload)
            log.warning(f"  Adicionado à skip-list: {exp_code} (motivo: {reason})")
    except portalocker.exceptions.LockException as e:
        log.error(f"Skip-list lock timeout em add({exp_code}): {e}")
        raise


def clear_skip_list() -> None:
    """Apaga skip-list (no proximo scan todos os sets sao retentados).
    Thread+process-safe."""
    lock_path = _skip_list_lock_path()
    try:
        with portalocker.Lock(
            lock_path,
            mode="a+",
            flags=portalocker.LOCK_EX | portalocker.LOCK_NB,
            timeout=_SKIP_LIST_LOCK_TIMEOUT,
            fail_when_locked=False,
            check_interval=0.1,
        ):
            if SKIP_LIST_FILE.exists():
                SKIP_LIST_FILE.unlink()
                log.info(f"Skip-list apagada: {SKIP_LIST_FILE.name}")
    except portalocker.exceptions.LockException as e:
        log.error(f"Skip-list lock timeout em clear(): {e}")
        raise


# ══════════════════════════════════════════════════════════════════════
# CHECKPOINT WRITER — JSONL append-only crash-safe (v2.6, 2026-05-17)
#
# Por quê?
#   Scanner v2.5 mantinha `opps: list[Opportunity]` em memória até
#   `wb.save()` no fim do `scan()`. Crash mid-run (incidente 2026-05-17
#   17:56→20:21, weekly local) perdia TUDO — 686 sets de dados.
#
# Modelo:
#   - Append-only JSONL ao lado do XLSX final (.checkpoint.jsonl sidecar)
#   - Flush imediato + fsync após cada SET — não bufferizar
#   - Linha de header com args/total_sets, set_complete por set, e
#     opportunity por deal encontrado (preserva ordem temporal real)
#   - Last line parcial é descartável em parse (JSONL semantics)
#
# Recovery via scripts/recover_from_checkpoint.py: parseia JSONL,
# reconstrói lista de Opportunity equivalente, reusa export_xlsx().
# ══════════════════════════════════════════════════════════════════════
class CheckpointWriter:
    """Append-only JSONL writer crash-safe pra estado parcial do scan.

    Uso:
        cw = CheckpointWriter(path, every_n=10)   # every_n=0 desabilita
        cw.write_header(args_dict, total_sets)
        for ...:
            for opp in scan_expansion(exp):
                cw.write_opportunity(opp)
            cw.write_set_complete(set_code, set_name, stats_dict, elapsed_s)
        cw.close()

    Atributo `every_n` é UM NO-OP semântico no v2.6 — todos os writes
    fazem flush+fsync imediato (toda granularidade é per-set, não por
    batch). Mantido pra compatibilidade futura caso adicionemos
    bufferização opcional.
    """

    def __init__(self, path: Path, every_n: int = 10,
                 heartbeat_path: Optional[Path] = None,
                 flush_every_listings: int = 50):
        self.path = path
        self.every_n = every_n
        self.enabled = every_n > 0
        # PR-F (2026-05-28): heartbeat + flush per-listing.
        # heartbeat é independente do checkpoint JSONL — útil pra detectar stall
        # mesmo com --checkpoint-every 0. flush_every_listings controla a
        # cadência das linhas set_progress no JSONL (telemetria, não deal).
        self.heartbeat_path = heartbeat_path
        self.flush_every_listings = flush_every_listings
        self._fh = None
        self._opps_written = 0
        self._sets_written = 0
        if self.enabled:
            # Garante diretório existente
            self.path.parent.mkdir(parents=True, exist_ok=True)
            # Abre em modo append — preserva qualquer conteúdo de run anterior
            # incompleto. Recovery script pode unir múltiplas runs se necessário.
            self._fh = open(self.path, "a", encoding="utf-8")
            log.info(
                f"Checkpoint JSONL ativo: {self.path} (flush per-set, "
                f"every_n={every_n}, flush_every_listings={flush_every_listings})"
            )
        if self.heartbeat_path is not None:
            try:
                self.heartbeat_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                log.warning(f"Heartbeat dir mkdir falhou ({self.heartbeat_path}): {e}")

    def _write(self, obj: dict) -> None:
        """Append + flush + fsync. Síncrono. Falha silenciosa em log."""
        if not self.enabled or self._fh is None:
            return
        try:
            # default=str: torna Path (e outros não-serializáveis, ex. do args
            # dict no header) seguros. PR-F: --state-dir injeta um Path em args.
            self._fh.write(json.dumps(obj, ensure_ascii=False, default=str) + "\n")
            self._fh.flush()
            os.fsync(self._fh.fileno())
        except Exception as e:
            # NUNCA mata o scan por falha de checkpoint. Loga + continua.
            log.warning(f"Checkpoint write falhou ({obj.get('_type', '?')}): {e}")

    def touch_heartbeat(self, note: str = "") -> None:
        """PR-F: reescreve heartbeat.txt (modo 'w') com timestamp + nota, com
        flush+fsync. Independente do checkpoint JSONL — funciona mesmo com
        --checkpoint-every 0. Falha silenciosa em log (igual _write): heartbeat
        NUNCA mata o scan."""
        if self.heartbeat_path is None:
            return
        try:
            ts = datetime.now().astimezone().isoformat(timespec="seconds")
            with open(self.heartbeat_path, "w", encoding="utf-8") as hb:
                hb.write(f"{ts}\t{note}\n")
                hb.flush()
                os.fsync(hb.fileno())
        except Exception as e:
            log.warning(f"Heartbeat write falhou ({note!r}): {e}")

    def write_progress(self, set_code: str, i: int, total: int) -> None:
        """PR-F: linha de telemetria set_progress (NÃO é deal). Sinaliza ao
        recovery/monitor que o pricing loop está vivo dentro de um set. O
        parser de recovery DEVE ignorar _type=set_progress (não vira
        Opportunity). Append + flush + fsync via _write."""
        if not self.enabled:
            return
        self._write({
            "_type": "set_progress",
            "stamp": datetime.now().astimezone().isoformat(timespec="seconds"),
            "set_code": set_code,
            "i": i,
            "total": total,
        })

    def write_header(self, args_dict: dict, total_sets: int,
                     usd_brl: float = 0.0, eur_brl: float = 0.0) -> None:
        if not self.enabled:
            return
        # v2.14 (correção de resultado — candidato 3): grava as taxas de câmbio
        # do scan no header. O recovery (scripts/recover_from_checkpoint.py) lê
        # daqui pra reconstruir a célula Stats `usd_brl_rate` do XLSX. Sem isso,
        # o XLSX recuperado tinha usd_brl_rate=0.0 → a tabela de ENTREGA do
        # postprocess (coluna "CT US$", convertida via FX) ficava vazia. A
        # CLASSIFICAÇÃO (COMPRA/REVISAR) NÃO depende do FX (usa colunas em BRL),
        # mas a entrega ao operador sim.
        self._write({
            "_type": "scan_header",
            "stamp": datetime.now().astimezone().isoformat(timespec="seconds"),
            "args": args_dict,
            "total_sets": total_sets,
            "usd_brl": round(usd_brl, 4) if usd_brl else 0.0,
            "eur_brl": round(eur_brl, 4) if eur_brl else 0.0,
        })

    def write_opportunity(self, opp: "Opportunity") -> None:
        if not self.enabled:
            return
        # asdict() converte Opportunity+Listing nested em dict serializável.
        # Adiciona _type + denormaliza campos críticos pro topo (set_code,
        # blueprint_id, name) — facilita parse no recovery script sem precisar
        # entrar no nested listing.
        d = asdict(opp)
        d["_type"] = "opportunity"
        d["set_code"] = opp.listing.set_code
        d["blueprint_id"] = opp.listing.blueprint_id
        d["name"] = opp.listing.card_name
        self._write(d)
        self._opps_written += 1

    def write_set_complete(self, set_code: str, set_name: str,
                            per_set_stats: dict, elapsed_s: float) -> None:
        if not self.enabled:
            return
        self._write({
            "_type": "set_complete",
            "set_code": set_code,
            "set_name": set_name,
            "blueprints": per_set_stats.get("blueprints", 0),
            "filtered": per_set_stats.get("filtered", 0),
            "priced": per_set_stats.get("priced", 0),
            "opps_found": per_set_stats.get("opps_found", 0),
            "elapsed_s": round(elapsed_s, 2),
        })
        self._sets_written += 1

    def write_scan_complete(self, total_opps: int, total_elapsed_s: float) -> None:
        if not self.enabled:
            return
        self._write({
            "_type": "scan_complete",
            "stamp": datetime.now().astimezone().isoformat(timespec="seconds"),
            "total_opps": total_opps,
            "total_elapsed_s": round(total_elapsed_s, 2),
            "sets_written": self._sets_written,
            "opps_written": self._opps_written,
        })

    def close(self) -> None:
        if self._fh is not None:
            try:
                self._fh.close()
            except Exception:
                pass
            self._fh = None


class Scanner:
    def __init__(self, ct: CardTraderClient, pricing: PricingProvider, cache: Cache,
                 threshold: float = MARGIN_THRESHOLD,
                 min_price_usd: float = MIN_PRICE_USD,
                 exclude_graded: bool = EXCLUDE_GRADED,
                 shipping_brl_override: float = 0.0,
                 hub_fee_rate: float = 0.0,
                 per_set_timeout_s: float = DEFAULT_PER_SET_TIMEOUT_MIN * 60,
                 ignore_skip_list: bool = False,
                 chase_only: bool = False,
                 max_consecutive_misses: int = 0):
        self.ct = ct
        self.pricing = pricing
        self.cache = cache
        self.threshold = threshold
        self.min_price_usd = min_price_usd
        self.exclude_graded = exclude_graded
        # 2026-05-12: modelo operacional de consolidacao no Hub depot. Frete
        # default = 0 (cartas acumulam no deposito CT ~100 unidades e sao
        # enviadas em consolacao unica → frete dilui pra ~R$0.30/card).
        # Override via CLI --shipping-brl X pra simular envio direto.
        self.shipping_brl_override = shipping_brl_override
        # v2.12 (2026-06-06): hub fee aplicado sobre live_brl no recalc REAL.
        # DEFAULT 0.0 → margem reportada é BRUTA (custo = preço do site, sem
        # taxa). Operador calcula Hub fee/frete/cartão/IOF por fora. Override
        # via --hub-fee 0.06 pra reembutir os 6% históricos, se quiser.
        self.hub_fee_rate = hub_fee_rate
        # v2.4 (2026-05-15): per-set wall-clock timeout (segundos). Default 8min.
        # Quando excedido, scan_expansion aborta + adiciona o set à skip-list.
        self.per_set_timeout_s = per_set_timeout_s
        self.ignore_skip_list = ignore_skip_list
        # v2.13 (frente chase): --chase-only filtra pra raridades TOP/MID
        # (Special/Hyper/Secret/Illustration Rare, alt arts, ultra/rainbow/gold).
        self.chase_only = chase_only
        # v2.13 (frente eficiência): cap de "misses" consecutivos no pricing.
        # Quando N listings seguidos de um set não casam preço na pokemontcg.io
        # (buraco de cobertura — sets 2026 / mega evolução), aborta o set + skip
        # list em vez de moer centenas de listings no rate limit. 0 = desativado
        # (comportamento histórico preservado). Pendência v2.4 conhecida.
        self.max_consecutive_misses = max_consecutive_misses
        # PR-F (2026-05-28): defaults seguros pra heartbeat/checkpoint quando
        # scan_expansion é chamado direto (tests/diagnose) sem passar por scan().
        # scan() reaponta esses dois quando há um CheckpointWriter ativo.
        self.heartbeat = lambda *_: None
        self._checkpoint: Optional["CheckpointWriter"] = None
        self.usd_brl = get_usd_to_brl(cache)
        self.eur_brl = get_eur_to_brl(cache)
        self.stats = {
            "expansions_scanned": 0,
            "listings_fetched": 0,
            "listings_after_filters": 0,
            "tcg_price_found": 0,
            "opportunities_found": 0,
            "skipped_exotic_currency": 0,  # M4 fix: contador de listings em moedas exóticas (GBP/JPY/etc)
            "expansions_skipped_by_list": 0,  # v2.4: sets pulados via skip-list
            "expansions_timed_out": 0,        # v2.4: sets que estouraram per_set_timeout
            "pricing_failures": 0,            # v2.9 (Codex H5): erros desconhecidos no pricing
            "expansions_mass_pricing_abort": 0,  # v2.9: sets abortados por >50% pricing fails
            "skipped_non_chase": 0,           # v2.13: listings cortados por --chase-only
            "expansions_no_coverage_abort": 0,  # v2.13: sets abortados por N misses consecutivos
        }
        # v2.9 (Codex H5): threshold pra abortar set quando pricing está
        # massivamente quebrado (schema drift / SSL / endpoint down). 50% das
        # listings + minimum sample size de 20 evita falsos abort em sets
        # com 1-2 cards.
        self.mass_pricing_failure_threshold = 0.50
        self.mass_pricing_failure_min_sample = 20

    def _parse_listing(self, raw: dict, bp_index: dict) -> Optional[Listing]:
        """Converte o dict da API para dataclass Listing.
        `bp_index`: {blueprint_id: blueprint_dict} pré-carregado.
        """
        bp = bp_index.get(raw.get("blueprint_id"))
        if not bp:
            return None

        # Preço: CT retorna centavos + currency. Em 2026-04-29 descobrimos
        # que contas com país=Brasil recebem listings já convertidos pra BRL
        # pela API CT (independente do país do seller). Antes tudo vinha em EUR.
        # Estratégia: BRL é a moeda interna do scanner, com fallback EUR/USD
        # via Frankfurter pra sellers que ainda enviam moeda original.
        price = raw.get("price", {})
        cents = price.get("cents", 0)
        currency = (price.get("currency") or "BRL").upper()
        if currency == "BRL":
            price_brl_val = cents / 100.0
        elif currency == "EUR":
            price_brl_val = (cents / 100.0) * self.eur_brl
        elif currency == "USD":
            price_brl_val = (cents / 100.0) * self.usd_brl
        else:
            # GBP, JPY etc — raro, ignorado até termos demanda.
            # M4 fix (2026-05-11): log + stat counter pra não perder cobertura
            # silenciosa caso CT habilite GBP/JPY nativo no futuro.
            log.debug(
                f"moeda exótica ignorada: {currency} no listing "
                f"{raw.get('id')} (bp={raw.get('blueprint_id')})"
            )
            self.stats["skipped_exotic_currency"] += 1
            return None

        props = raw.get("properties_hash") or raw.get("properties") or {}
        user = raw.get("user", {})

        # Set code/name: blueprint só traz `expansion_id` numérico (sem dict
        # nested), então a fonte confiável é o próprio listing, que vem com
        # `expansion: {code, name_en, id}`. Descoberto em 2026-04-20 após
        # Mawile falso-positivo com set_code vazio.
        listing_exp = raw.get("expansion") or {}
        resolved_set_code = listing_exp.get("code") or ""
        resolved_set_name = listing_exp.get("name_en") or listing_exp.get("name") or ""

        return Listing(
            product_id=raw["id"],
            blueprint_id=raw["blueprint_id"],
            card_name=bp.get("name", ""),
            set_code=resolved_set_code,
            set_name=resolved_set_name,
            # Fonte preferida: properties_hash do listing, que traz "007" limpo.
            # Blueprint costuma ter collector_number vazio e cair em `version`
            # formatado como "Rarity | N/total" → aplicamos clean só nesse
            # fallback pra extrair o dígito via regex sem mexer nos paths
            # limpos (2026-05-11 fix).
            collector_number=(
                props.get("collector_number")
                or bp.get("collector_number")
                or clean_collector_number(bp.get("version", ""))
                or ""
            ),
            condition=props.get("condition", "Unknown"),
            # Pokemon usa sempre properties_hash.pokemon_language; mtg_language
            # é sempre None pra cartas Pokemon. Sem valor → string vazia (não "en")
            # para não deixar passar carta sem metadados de idioma.
            language=(props.get("pokemon_language") or "").lower(),
            price_cents=cents,
            price_currency=currency,
            price_brl=price_brl_val,
            quantity=raw.get("quantity", 0),
            # v2.18: Pokémon usa `pokemon_reverse` (não `foil`/`mtg_foil`, que
            # só existem em MTG). Antes esse campo era ignorado → TODA carta
            # Pokémon chegava foil=False, e holo rares padrão escorregavam pra
            # reverseHolofoil (inflado). Agora reverse-holo → foil=True.
            foil=(props.get("mtg_foil", False) or props.get("foil", False)
                  or props.get("pokemon_reverse", False)),
            graded=raw.get("graded", False),
            seller_username=user.get("username", ""),
            seller_can_sell_via_hub=user.get("can_sell_via_hub", False),
            seller_user_type=user.get("user_type", "private"),
            cardtrader_url=f"https://www.cardtrader.com/cards/{raw['blueprint_id']}",
            # 2026-05-16: Schema CT blueprint tem rarity em fixed_properties.pokemon_rarity,
            # NÃO no campo `rarity` raiz (verificado via inspect do bp[0] de sfa).
            # Patterns esperados: "Common"/"Uncommon"/"Rare"/"Ultra Rare"/"Secret Rare"/etc.
            # Vazio = sealed product (booster, ETB) — não-card, será filtrado downstream.
            rarity=(
                (bp.get("fixed_properties") or {}).get("pokemon_rarity")
                or bp.get("rarity")
                or props.get("pokemon_rarity")
                or props.get("rarity")
                or ""
            ).strip(),
        )

    def _passes_filters(self, l: Listing) -> bool:
        if l.condition != CONDITION_FILTER:
            return False
        if LANGUAGE_FILTER and l.language != LANGUAGE_FILTER:
            return False
        if self.exclude_graded and l.graded:
            return False
        # v2.1 — anti-Trainer Gallery: skipa cartas TG## (subset SWSH com
        # pricing corrompido no pokemontcg.io). Ver _TRAINER_GALLERY_RE.
        if l.collector_number and _TRAINER_GALLERY_RE.match(l.collector_number):
            return False
        # Filtro min_price_usd: converte BRL→USD via FX pra manter o threshold
        # em USD (referência TCGPlayer), sem importar a moeda do listing.
        price_usd_equiv = l.price_brl / self.usd_brl
        if price_usd_equiv < self.min_price_usd:
            return False
        return True

    def _estimate_shipping_brl(self, listing: Listing, units: int = 1) -> float:
        """Frete em BRL por listing. Default 0 (modelo de consolidacao no Hub depot).

        Modelo operacional 2026-05-12 (confirmado por Matheus):
        - Compras CT vao pro deposito Hub na Europa
        - Acumulam ~100 cartas
        - Enviadas em consolidacao unica pro Brasil
        - Frete daquele envio (~R$30-50) dilui entre as 100 cartas
        - Resultado per-card: R$0.30-0.50 → desprezivel

        Logo o scanner zera frete por default. Hub fee 6% e o unico custo
        adicional alem do preco CT, e ja vem embutido em `live_price_brl`
        via per-blueprint validation (markup_tier "Hub +6%").

        Override (`--shipping-brl X`): pra simular cenarios sem
        consolidacao (compra urgente, sellers non-Hub que enviam direto).
        Valor passa direto, dividido por `units` se compra multipla do
        mesmo listing.

        Implementacao antiga (estimativa EUR-based por tier) preservada
        como `_legacy_shipping_eur_estimate_unused` para referencia futura.
        """
        return self.shipping_brl_override / max(units, 1)

    def _legacy_shipping_eur_estimate_unused(self, listing: Listing) -> float:
        """[INATIVO desde 2026-05-12] Estimativa de frete baseada em EUR + FX.

        Calculo pre-consolidacao: tier de seller (Hub/Pro/Private) → base em
        EUR (5/10/12) → multiplicado por FX EUR→BRL → frete per-listing
        ~R$28-69.

        Substituido por shipping=0 default quando o modelo operacional de
        consolidacao foi confirmado. Mantido como referencia: se o fluxo de
        consolidacao mudar (ex: virar envio direto), reativar essa logica.
        """
        if listing.seller_user_type == "zero_fee" or listing.seller_can_sell_via_hub:
            base_eur = SHIPPING_EUR_HUB
        elif listing.seller_user_type == "professional":
            base_eur = SHIPPING_EUR_PROFESSIONAL
        else:
            base_eur = SHIPPING_EUR_PRIVATE
        return base_eur * self.eur_brl

    def _check_set_timeout(self, set_start: float, exp_code: str, exp_name: str,
                            stage: str, timeout_s: Optional[float] = None) -> bool:
        """Retorna True se set deve abortar. Loga + marca skip-list quando True.
        v2.8 (Codex H2): chamado antes de cada call externa do set.
        v2.14: aceita timeout_s explícito (o EFETIVO do set, c/ override). Quando
        None, cai no self.per_set_timeout_s (retrocompat p/ chamadas diretas)."""
        effective = self.per_set_timeout_s if timeout_s is None else timeout_s
        if not effective:
            return False
        elapsed = time.monotonic() - set_start
        if elapsed > effective:
            log.error(
                f"  ⏱️  TIMEOUT set {exp_code} ({exp_name}) at stage='{stage}': "
                f"{elapsed/60:.1f}min > {effective/60:.1f}min limite. "
                f"Abortando."
            )
            self.stats["expansions_timed_out"] += 1
            try:
                add_to_skip_list(
                    exp_code,
                    f"per_set_timeout_{int(effective)}s_at_{stage}",
                )
            except Exception as e:
                log.warning(f"  Falha ao gravar skip-list ({exp_code}): {e}")
            return True
        return False

    def scan_expansion(self, expansion: dict) -> Iterator[Opportunity]:
        exp_id = expansion["id"]
        exp_code = expansion.get("code", "")
        exp_name = expansion.get("name", "")
        log.info(f"→ Scan: {exp_name} ({exp_code}) [id={exp_id}]")
        # v2.4: marca início pra wall-clock timeout no pricing loop
        set_start = time.monotonic()
        # v2.14: timeout EFETIVO do set — aplica override de SET_TIMEOUT_OVERRIDES
        # (sets vintage pesados ganham mais fôlego sem o operador passar flag).
        # Tudo abaixo usa este valor, não self.per_set_timeout_s direto.
        effective_timeout_s = effective_set_timeout_s(exp_code, self.per_set_timeout_s)
        if effective_timeout_s != self.per_set_timeout_s:
            log.info(
                f"  ⏱️  Timeout override p/ {exp_code}: "
                f"{effective_timeout_s/60:.0f}min "
                f"(default {self.per_set_timeout_s/60:.0f}min)"
            )
        # v2.8 (Codex H2): deadline absoluta (monotonic). Passada pra todas
        # chamadas CT do set, então retries/429-sleeps respeitam o cap.
        deadline_ts: Optional[float] = (
            set_start + effective_timeout_s if effective_timeout_s else None
        )

        # v2.8: check ANTES de blueprints (cobre o caso de set já estourado
        # por algum estado externo, embora improvável aqui — set_start é now)
        if self._check_set_timeout(set_start, exp_code, exp_name, "pre_blueprints",
                                    timeout_s=effective_timeout_s):
            return

        # Carrega blueprints da expansão (indexa para O(1) lookup).
        # v2.8: deadline propaga pra _get → retries respeitam cap.
        try:
            blueprints = self.ct.list_blueprints(exp_id, deadline_ts=deadline_ts)
        except TimeoutError as e:
            log.error(f"  ⏱️  CT blueprints timeout para {exp_code}: {e}")
            self.stats["expansions_timed_out"] += 1
            try:
                add_to_skip_list(
                    exp_code,
                    f"ct_blueprints_timeout_{int(effective_timeout_s or 0)}s",
                )
            except Exception as ee:
                log.warning(f"  Falha ao gravar skip-list ({exp_code}): {ee}")
            return
        bp_index = {bp["id"]: bp for bp in blueprints}
        log.info(f"  {len(blueprints)} blueprints carregados")

        # v2.8: check entre blueprints e listings — se blueprints foi rápido
        # mas o set já consumiu o orçamento (improvável p/ 1 call), abortar
        if self._check_set_timeout(set_start, exp_code, exp_name, "pre_listings",
                                    timeout_s=effective_timeout_s):
            return

        # Puxa todas listings EN da expansão de uma vez (muito + eficiente que
        # 1 chamada por blueprint — economiza de 400+ calls para 1).
        try:
            raw_listings = self.ct.list_listings_by_expansion(
                exp_id, language=LANGUAGE_FILTER, deadline_ts=deadline_ts
            )
        except TimeoutError as e:
            log.error(f"  ⏱️  CT listings timeout para {exp_code}: {e}")
            self.stats["expansions_timed_out"] += 1
            try:
                add_to_skip_list(
                    exp_code,
                    f"ct_listings_timeout_{int(effective_timeout_s or 0)}s",
                )
            except Exception as ee:
                log.warning(f"  Falha ao gravar skip-list ({exp_code}): {ee}")
            return
        self.stats["listings_fetched"] += len(raw_listings)
        log.info(f"  {len(raw_listings)} listings EN encontrados")

        # Dedup por (blueprint + seller + condição): mantém o menor preço.
        # Comparação é em BRL (não em cents) porque listings podem vir em
        # moedas diferentes — comparar cents direto seria errado (1 cent BRL
        # ≠ 1 cent EUR).
        best_by_uid: dict[str, Listing] = {}
        for raw in raw_listings:
            l = self._parse_listing(raw, bp_index)
            if not l:
                continue
            if not self._passes_filters(l):
                continue
            existing = best_by_uid.get(l.uid)
            if not existing or l.price_brl < existing.price_brl:
                best_by_uid[l.uid] = l

        self.stats["listings_after_filters"] += len(best_by_uid)
        log.info(f"  {len(best_by_uid)} listings após filtros (NM, EN, não-graded, ≥${self.min_price_usd})")

        # Para cada listing filtrado, busca preço TCG e calcula margem
        total_listings = len(best_by_uid)
        # v2.9 (Codex H5): contadores per-set pra detectar mass pricing failure
        set_pricing_attempts = 0
        set_pricing_failures = 0
        # v2.13 (frente eficiência): contador de "misses" consecutivos (preço
        # não encontrado na fonte) pra cortar sets sem cobertura cedo.
        set_consecutive_misses = 0
        for i, l in enumerate(best_by_uid.values(), 1):
            # v2.4: wall-clock timeout check antes de cada call de pricing
            # v2.14: usa o timeout EFETIVO do set (com override vintage).
            if effective_timeout_s and (time.monotonic() - set_start) > effective_timeout_s:
                elapsed = time.monotonic() - set_start
                log.error(
                    f"  ⏱️  TIMEOUT set {exp_code} ({exp_name}): {elapsed/60:.1f}min "
                    f"> {effective_timeout_s/60:.1f}min limite. Abortando após "
                    f"{i-1}/{total_listings} listings priced."
                )
                self.stats["expansions_timed_out"] += 1
                add_to_skip_list(exp_code, f"per_set_timeout_{int(effective_timeout_s)}s_at_{i-1}_of_{total_listings}")
                return
            if i % 50 == 0 or i == total_listings:
                log.info(f"  Pricing progress: {i}/{total_listings} listings consultados")
            # PR-F (2026-05-28): heartbeat + set_progress JSONL na cadência
            # configurável (--flush-every-listings, default 50). Distinto do
            # INFO log acima (mantido em i%50 pra não quebrar contrato de log).
            flush_n = getattr(self._checkpoint, "flush_every_listings", 0) if self._checkpoint else 0
            if flush_n and (i % flush_n == 0 or i == total_listings):
                self.heartbeat(f"set={exp_code} i={i}/{total_listings}")
                if self._checkpoint is not None:
                    self._checkpoint.write_progress(exp_code, i, total_listings)
            set_pricing_attempts += 1
            tcg_market = None
            try:
                tcg_market = self.pricing.market_price_usd(
                    l.card_name, l.set_code, l.collector_number,
                    foil=l.foil,  # H2 fix: foil/reverse-aware variant selection
                    rarity=l.rarity,  # v2.18: holo rare → holofoil; base → normal
                )
            except (requests.ConnectionError, requests.Timeout) as e:
                # v2.9: erros de rede transientes — já foram retried pelo provider.
                # Log debug (não vira WARNING flood), conta como pricing_failure
                # pra mass-failure detection.
                log.debug(f"  Pricing transient {l.card_name}: {type(e).__name__}: {e}")
                set_pricing_failures += 1
                self.stats["pricing_failures"] += 1
            except Exception as e:
                # v2.9 (Codex H5): erros DESCONHECIDOS = WARNING + counter.
                # Pre-fix: log.debug + swallow → schema drift / SSL / JSON parse
                # silenciosamente removia deals. Agora aparece no INFO log.
                log.warning(
                    f"  Pricing FAILURE {l.card_name} ({l.set_code}/{l.collector_number}) "
                    f"bp={getattr(l, 'blueprint_id', '?')}: {type(e).__name__}: {e}"
                )
                set_pricing_failures += 1
                self.stats["pricing_failures"] += 1

            # v2.9: detecção de mass pricing failure → aborta set + skip-list.
            # Trigger só após sample mínimo pra evitar falso abort em sets
            # com poucos cards.
            if (
                set_pricing_attempts >= self.mass_pricing_failure_min_sample
                and (set_pricing_failures / set_pricing_attempts) > self.mass_pricing_failure_threshold
            ):
                fail_pct = set_pricing_failures / set_pricing_attempts * 100
                log.error(
                    f"  💥 MASS PRICING FAILURE em {exp_code} ({exp_name}): "
                    f"{set_pricing_failures}/{set_pricing_attempts} ({fail_pct:.0f}%) "
                    f"falharam. Abortando set."
                )
                self.stats["expansions_mass_pricing_abort"] += 1
                try:
                    add_to_skip_list(
                        exp_code,
                        f"mass_pricing_failure_{set_pricing_failures}_of_{set_pricing_attempts}",
                    )
                except Exception as ee:
                    log.warning(f"  Falha ao gravar skip-list ({exp_code}): {ee}")
                return

            if not tcg_market or tcg_market <= 0:
                # v2.13 (frente eficiência): conta miss consecutivo. Em sets sem
                # cobertura na pokemontcg.io (ex: 2026 / mega evolução), centenas
                # de listings seguidos não casam preço — cada um custa uma chamada
                # à API (com rate limit) que sempre falha. Quando bate o cap,
                # aborta o set e o joga na skip-list (mesma mecânica do timeout/
                # mass-pricing-failure). Distingue-se de pricing_failure (erro de
                # rede/exceção): aqui o request foi OK, só não há match.
                set_consecutive_misses += 1
                if (self.max_consecutive_misses > 0
                        and set_consecutive_misses >= self.max_consecutive_misses):
                    log.warning(
                        f"  🕳️  SEM COBERTURA em {exp_code} ({exp_name}): "
                        f"{set_consecutive_misses} listings consecutivos sem preço "
                        f"na fonte (cap={self.max_consecutive_misses}). Abortando set "
                        f"após {i}/{total_listings}."
                    )
                    self.stats["expansions_no_coverage_abort"] += 1
                    try:
                        add_to_skip_list(
                            exp_code,
                            f"no_coverage_{set_consecutive_misses}_consecutive_misses_at_{i}_of_{total_listings}",
                        )
                    except Exception as ee:
                        log.warning(f"  Falha ao gravar skip-list ({exp_code}): {ee}")
                    return
                continue
            # Hit: zera o contador de misses consecutivos.
            set_consecutive_misses = 0
            self.stats["tcg_price_found"] += 1

            # Margem calculada em BRL (a verdade operacional pro Matheus).
            # TCGPlayer market vem em USD → converte via Frankfurter.
            # v2.12 (2026-06-06): custo = preço do site × (1 + hub_fee_rate).
            # Com hub_fee_rate default 0.0, custo = preço do site → margem BRUTA
            # `(tcg − preço)/tcg`. Operador soma Hub fee/frete/cartão/IOF por fora.
            # (Passe --hub-fee 0.06 pra reembutir os 6% históricos.)
            tcg_brl = tcg_market * self.usd_brl
            ct_brl = l.price_brl
            custo_brl = ct_brl * (1.0 + self.hub_fee_rate)
            margin = (tcg_brl - custo_brl) / tcg_brl
            if margin < self.threshold:
                continue

            shipping_brl = self._estimate_shipping_brl(l)
            net_margin = (tcg_brl - custo_brl - shipping_brl) / tcg_brl

            self.stats["opportunities_found"] += 1
            # v2.7.1: tcg_url vem do provider (last call). Pode ser None se
            # provider != pokemontcg ou se a card não tem entry TCGPlayer.
            tcg_url = getattr(self.pricing, "last_tcg_url", None)
            # v2.8 Layer 4: variant usada (holofoil/reverseHolofoil/normal/etc).
            variant_used = getattr(self.pricing, "last_variant_used", None)
            # v2.13: rarity oficial pokemontcg.io (preferida) → fallback rarity
            # do blueprint CT. Set release date pra score de valorização.
            ptcg_rarity = getattr(self.pricing, "last_ptcg_rarity", None)
            # v2.14 (candidato 4) + v2.18: flag de baixa confiança de variante.
            # Listing NÃO-foil que só achou preço numa variante metálica premium
            # (holofoil/unlimitedHolofoil — nem normal nem reverseHolofoil).
            # v2.18: NÃO dispara pra holo rare — aí holofoil é a variante CORRETA
            # (esperada), não um casamento errado. Só sinaliza quando uma carta
            # NÃO-holo casou preço holo (o caso Gengar não-foil → $1600).
            variant_low_conf = (
                l.foil is False
                and not _rarity_is_holo(l.rarity, ptcg_rarity)
                and variant_used in ("holofoil", "unlimitedHolofoil")
            )
            set_release = getattr(self.pricing, "last_set_release_date", None) or ""
            rarity_for_tier = ptcg_rarity or l.rarity
            chase_tier = classify_chase_tier(rarity_for_tier)
            val_score, val_note = compute_valorization(
                rarity_for_tier, set_release, tcg_market
            )
            # Filtro --chase-only: só TOP/MID passam (cartas de maior interesse).
            if self.chase_only and chase_tier not in CHASE_ONLY_TIERS:
                self.stats["skipped_non_chase"] += 1
                continue
            yield Opportunity(
                listing=l,
                tcg_market_usd=tcg_market,
                tcg_market_brl=tcg_brl,
                ct_price_brl=ct_brl,
                margin_pct=margin,
                margin_brl=tcg_brl - custo_brl,
                estimated_shipping_brl=shipping_brl,
                net_margin_pct=net_margin,
                tcg_url=tcg_url,
                price_variant_used=variant_used,
                chase_tier=chase_tier,
                valorization_score=val_score,
                valorization_note=val_note,
                set_release_date=set_release,
                variant_low_confidence=variant_low_conf,
            )

        self.stats["expansions_scanned"] += 1

    def scan(self, expansions: list[dict],
             checkpoint: Optional["CheckpointWriter"] = None) -> list[Opportunity]:
        opps: list[Opportunity] = []
        # PR-F (2026-05-28): expõe heartbeat + checkpoint pro scan_expansion
        # (que não recebe checkpoint como arg). Se não há checkpoint, heartbeat
        # vira no-op (lambda) e write_progress é guardado por self._checkpoint=None.
        self._checkpoint = checkpoint
        self.heartbeat = checkpoint.touch_heartbeat if checkpoint else (lambda *_: None)
        # v2.4: carrega skip-list (sets que estouraram timeout em runs anteriores)
        skip_data = load_skip_list() if not self.ignore_skip_list else {"skipped": [], "reasons": {}}
        # PR-F (2026-05-28): TTL. Entries TRANSIENTES (timeout / mass-pricing-fail)
        # com added_at > 7d são re-tentadas (dropadas do skip_set deste run; o
        # arquivo NÃO é reescrito aqui — se o set falhar de novo, add_to_skip_list
        # refresca added_at). Entries permanentes (unexpected_error_*) e legadas
        # (str pura, sem timestamp) seguem pulando.
        reasons_map = skip_data.get("reasons") or {}
        skip_set = set(skip_data.get("skipped", []))
        expired = {
            code for code in skip_set
            if _skip_entry_is_expired(reasons_map.get(code))
        }
        if expired:
            log.info(
                f"Skip-list TTL ({_SKIP_LIST_TTL_DAYS}d): re-tentando {len(expired)} "
                f"set(s) transiente(s) expirado(s): {', '.join(sorted(expired))}"
            )
            skip_set -= expired
        if skip_set:
            log.info(
                f"Skip-list ativa ({SKIP_LIST_FILE.name}): {len(skip_set)} sets serão pulados "
                f"({', '.join(sorted(skip_set))}). Use --ignore-skip-list pra forçar."
            )
        # v2.5.1 (2026-05-16 night): heartbeat por set + total elapsed.
        # Operador pediu detectabilidade de stall: agora cada set loga
        # "ALIVE [HH:MM:SS] set N/TOTAL exp_code elapsed=Xmin" antes de scan_expansion.
        # Também captura Exception genérica (não só HTTPError) pra não matar
        # o loop inteiro se um set lançar algo inesperado.
        run_start = time.monotonic()
        total_sets = len(expansions)
        for idx, exp in enumerate(expansions, 1):
            exp_code = exp.get("code", "")
            elapsed_min = (time.monotonic() - run_start) / 60.0
            log.info(
                f"ALIVE [{datetime.now().strftime('%H:%M:%S')}] set {idx}/{total_sets} "
                f"({exp_code}) total_elapsed={elapsed_min:.1f}min"
            )
            # PR-F: heartbeat por set (detecta stall externamente via mtime)
            self.heartbeat(f"set {idx}/{total_sets} ({exp_code}) elapsed={elapsed_min:.1f}min")
            if exp_code in skip_set:
                self.stats["expansions_skipped_by_list"] += 1
                # PR-F: reason pode ser dict {reason, added_at} ou str legada.
                reason = _skip_reason_str(reasons_map.get(exp_code)) or "?"
                log.warning(f"⏭️  Pulando {exp.get('name')} ({exp_code}) — skip-list (motivo: {reason})")
                continue
            # v2.6 (2026-05-17): snapshot per-set stats antes do scan_expansion
            # pra calcular DIFFs (self.stats é cumulativo). Permite registrar
            # blueprints/filtered/priced/opps_found per set no checkpoint JSONL.
            stats_pre = dict(self.stats) if checkpoint and checkpoint.enabled else None
            set_start = time.monotonic()
            opps_pre_count = len(opps)
            try:
                # v2.6: itera o generator e escreve cada Opportunity no checkpoint
                # ASSIM QUE for yielded. Antes era `opps.extend(scan_expansion(exp))`
                # — esperar o generator esgotar dava melhor desempenho mas perdia
                # tudo em crash mid-set. Agora: live append-only.
                for opp in self.scan_expansion(exp):
                    opps.append(opp)
                    if checkpoint:
                        checkpoint.write_opportunity(opp)
            except requests.HTTPError as e:
                log.error(f"Falha em {exp.get('name')}: {e}")
                # Não emite set_complete em falha — recovery vai mostrar lacuna
                continue
            except Exception as e:
                # v2.5.1: nunca deixa exceção genérica matar o full scan
                log.error(
                    f"Erro inesperado em {exp.get('name')} ({exp_code}): "
                    f"{type(e).__name__}: {e}. Continuando próximos sets."
                )
                add_to_skip_list(exp_code, f"unexpected_error_{type(e).__name__}")
                continue
            # v2.6: set completou sem exception → emite set_complete com DIFFs
            if checkpoint and checkpoint.enabled and stats_pre is not None:
                per_set_stats = {
                    "blueprints": 0,  # sem stat agregado de blueprints; deixar 0
                    "filtered": self.stats.get("listings_after_filters", 0) - stats_pre.get("listings_after_filters", 0),
                    "priced": self.stats.get("tcg_price_found", 0) - stats_pre.get("tcg_price_found", 0),
                    "opps_found": len(opps) - opps_pre_count,
                }
                checkpoint.write_set_complete(
                    set_code=exp_code,
                    set_name=exp.get("name", ""),
                    per_set_stats=per_set_stats,
                    elapsed_s=time.monotonic() - set_start,
                )
        # Ordena por margem bruta desc (maior oportunidade primeiro)
        opps.sort(key=lambda o: o.margin_pct, reverse=True)
        return opps

    def validate_per_blueprint(self, opps: list[Opportunity], top_n: int = 30) -> None:
        """Fase 3 / v2.0 — Valida top N candidatos via endpoint per-blueprint.

        A API CT retorna preços diferentes nos endpoints per-expansion (RAW do
        seller, sem markup) e per-blueprint (preço FINAL com markup CT embutido,
        +6% tier Hub ou +20% tier non-VAT). O scanner usa per-expansion por
        eficiência (1 chamada por set vs 1 por carta), mas isso infla margens
        em 5-20% vs o preço que o comprador realmente paga.

        Esta validação faz chamadas per-blueprint pros top N por margem bruta,
        encontra o seller correspondente, e enriquece a Opportunity com:
          - live_price_brl (preço REAL com markup)
          - markup_pct, markup_tier (Hub +6% / non-VAT +20% / anômalo)
          - real_margin_pct, real_net_margin_pct, real_lucro_brl
          - validation_status (VALIDATED_REAL, VALIDATED_MARKUP, STALE,
            PRICE_CHANGED, API_ERROR)

        Mutate in place. Não filtra — caller decide threshold via min_net_margin.
        """
        if not opps:
            return
        # Top N por margem bruta (a mesma ordem do scan default)
        sorted_opps = sorted(opps, key=lambda o: o.margin_pct, reverse=True)
        candidates = sorted_opps[:top_n]
        unique_bp_ids = {o.listing.blueprint_id for o in candidates}
        log.info(f"v2.0: validando top {len(candidates)} candidatos ({len(unique_bp_ids)} blueprints únicos) per-blueprint...")

        # Cache resultado per-blueprint (uma chamada por bp_id, mesmo se múltiplos sellers do mesmo card)
        bp_listings: dict[int, Optional[list[dict]]] = {}
        for bp_id in unique_bp_ids:
            try:
                listings = self.ct.list_listings_by_blueprint(bp_id, language=LANGUAGE_FILTER)
                if isinstance(listings, dict):
                    listings = [l for sub in listings.values() for l in sub]
                bp_listings[bp_id] = listings
            except Exception as e:
                log.warning(f"v2.0: per-blueprint falhou bp={bp_id}: {e}")
                bp_listings[bp_id] = None

        # Enriquece cada candidato
        for o in candidates:
            bp_id = o.listing.blueprint_id
            seller = o.listing.seller_username
            listings = bp_listings.get(bp_id)
            if listings is None:
                o.validation_status = "API_ERROR"
                continue
            # v2.19 (2026-06-21): casar o listing per-blueprint pela MESMA
            # condição (NM) E mesmo reverse/variante do listing que o scan achou
            # — não só pelo vendedor. Antes pegava o PRIMEIRO listing do vendedor
            # (qualquer condição/variante): um vendedor com uma cópia Poor barata
            # OU a versão não-reverse "contaminava" a validação. Caso real
            # Shiftry hl: scan achou a NM reverse (~R$140), mas a validação casava
            # a Poor não-reverse do mesmo vendedor (R$46,46) → comparava preço de
            # carta Poor contra referência reverse → falso "79%". Agora exige
            # condition == o.listing.condition (Near Mint) e reverse igual; entre
            # os que batem, pega o MAIS BARATO (melhor oferta daquele vendedor
            # naquela exata variante/condição). Se nenhum bate → STALE honesto
            # (a cópia exata que o scan viu sumiu), em vez de casar a errada.
            def _variant_matches(listing_dict) -> bool:
                p = (listing_dict.get("properties_hash")
                     or listing_dict.get("properties") or {})
                cond = p.get("condition", "")
                is_rev = bool(p.get("pokemon_reverse", False)
                              or p.get("mtg_foil", False)
                              or p.get("foil", False))
                return cond == o.listing.condition and is_rev == bool(o.listing.foil)

            seller_matches = [
                l for l in listings
                if (l.get("user") or {}).get("username") == seller
                and _variant_matches(l)
            ]
            match = (min(seller_matches, key=lambda l: l.get("price_cents", float("inf")))
                     if seller_matches else None)
            if not match:
                # STALE: vendedor não tem mais a cópia na MESMA condição+variante
                # que o scan achou (ou só tinha outra condição/variante).
                o.validation_status = "STALE"
                continue
            # Live BRL (per-blueprint sempre retorna na moeda da conta = BRL pra Matheus,
            # mas tratamos as 3 moedas pra robustez)
            live_cents = match.get("price_cents", 0)
            live_currency = (match.get("price_currency") or "BRL").upper()
            if live_currency == "BRL":
                live_brl = live_cents / 100.0
            elif live_currency == "EUR":
                live_brl = (live_cents / 100.0) * self.eur_brl
            elif live_currency == "USD":
                live_brl = (live_cents / 100.0) * self.usd_brl
            else:
                o.validation_status = "API_ERROR"
                continue

            o.live_price_brl = live_brl
            markup = (live_brl - o.ct_price_brl) / o.ct_price_brl if o.ct_price_brl > 0 else 0.0
            o.markup_pct = markup
            # Classificação tier (com base no padrão observado 2026-04-29 e refinado 2026-05-11)
            if markup < 0.02:
                o.markup_tier = "Real (sem markup)"
                o.validation_status = "VALIDATED_REAL"
            elif markup < 0.12:
                o.markup_tier = "Hub (+6%)"
                o.validation_status = "VALIDATED_MARKUP"
            elif markup < 0.30:
                o.markup_tier = "non-VAT (+20%)"
                o.validation_status = "VALIDATED_MARKUP"
            elif markup < 0.45:
                # 2026-05-11: tiers 30-45% (TheDragonsVault, A2Z TCG, Fun Gs)
                # parecem legítimos — provavelmente non-Hub VAT-exempt ou
                # categoria comissionada alta. Validados, não erro de preço.
                o.markup_tier = f"Alto markup ({markup*100:+.0f}%)"
                o.validation_status = "VALIDATED_MARKUP"
            else:
                # markup ≥45% — provavelmente preço mudou entre scan e validate
                o.markup_tier = f"Anômalo ({markup*100:+.0f}%)"
                o.validation_status = "PRICE_CHANGED"

            # Recalc margens com preço LIVE (per-blueprint, preço de checkout).
            # v2.12 (2026-06-06): custo real = live_brl × (1 + hub_fee_rate). Com
            # hub_fee_rate default 0.0, custo = live_brl → margens REAIS são BRUTAS
            # `(tcg − live)/tcg`. Operador soma Hub fee/frete/cartão/IOF por fora.
            # (--hub-fee 0.06 reembute os 6% históricos.) Alinha com postprocess.
            hub_fee_brl = live_brl * self.hub_fee_rate
            custo_real = live_brl + hub_fee_brl
            o.real_margin_pct = (o.tcg_market_brl - custo_real) / o.tcg_market_brl
            o.real_net_margin_pct = (o.tcg_market_brl - custo_real - o.estimated_shipping_brl) / o.tcg_market_brl
            o.real_lucro_brl = o.tcg_market_brl - custo_real - o.estimated_shipping_brl

        # Stats sumarizadas
        validated_real = sum(1 for o in candidates if o.validation_status == "VALIDATED_REAL")
        validated_markup = sum(1 for o in candidates if o.validation_status == "VALIDATED_MARKUP")
        stale = sum(1 for o in candidates if o.validation_status == "STALE")
        log.info(f"v2.0: {validated_real} REAL, {validated_markup} com markup, {stale} STALE, {len(candidates) - validated_real - validated_markup - stale} outros")
        self.stats["v2_validated"] = validated_real + validated_markup
        self.stats["v2_stale"] = stale


# ══════════════════════════════════════════════════════════════════════
# EXCEL EXPORT — openpyxl para compatibilidade com MYP scanner.
# Layout:
#   - Aba "Oportunidades": uma linha por deal, ordenado por margem desc
#   - Aba "Stats": métricas da execução (funnel de filtragem)
# Formatação condicional: margem >= 40% fica verde forte, <30% vermelho.
# ══════════════════════════════════════════════════════════════════════
def export_xlsx(opportunities: list[Opportunity], stats: dict,
                out_path: Path, usd_brl: float, eur_brl: float,
                threshold: float, hub_fee_rate: float = 0.0) -> Path:
    wb = Workbook()

    # Aba 1: Oportunidades
    ws = wb.active
    ws.title = "Oportunidades"

    headers = [
        "Card Name", "Set", "Nº", "Rarity", "Condição", "Idioma",
        "Scan R$ (raw)", "Moeda Original CT", "LIVE R$ (real)", "Markup %", "Markup Tier",
        "Validation Status",
        "TCG Market (BRL)", "TCG Market (USD)",
        "Margem % (scan)", "Margem % REAL", "Net Margin % (scan)", "Net Margin % REAL",
        "Lucro R$ REAL", "Frete Est. R$",
        "Qtd", "Foil",
        # v2.8 Layer 4 (2026-05-18): variante TCGPlayer usada no cálculo.
        # Operador valida: foil-listing deve casar com `holofoil`; non-foil
        # com `reverseHolofoil` na maioria dos Holo Rare antigos. Mismatch
        # entre Foil="No" e Variant="holofoil" indica preço inflado.
        "Variant",
        "Seller", "Tipo Seller", "Hub",
        # v2.7.1 (2026-05-18): Link TCG — URL TCGPlayer da carta exata
        # matched no pokemontcg.io. Operador valida variante (ex Lusamine 1st Place
        # vs normal) antes de comprar. Vazio quando provider != pokemontcg ou
        # card sem entry TCGPlayer (Promos antigos comum).
        "Link CardTrader", "Link TCG", "Scanned At",
        # v2.13 — frentes "chase" e "valorização". APPEND no fim (não realocar
        # colunas existentes: os testes de coerência checam índices fixos 6-19,
        # 26, 27). Chase Tier = raridade top (TOP/MID/MODEST/BULK). Desconto %
        # = margem bruta repetida com rótulo do ranking de chase. Valorização =
        # score heurístico 0-100 (rarity + idade do set + preço) + notas.
        "Chase Tier", "Desconto %", "Valorização (0-100)", "Valorização — Notas",
        # v2.14 (candidato 4): flag de baixa confiança de variante. APPEND no
        # FIM (não realocar — testes de coerência checam índices fixos).
        "Variante Baixa Confiança",
    ]
    ws.append(headers)

    for opp in opportunities:
        l = opp.listing
        ws.append([
            l.card_name,
            f"{l.set_name} ({l.set_code})",
            l.collector_number,
            l.rarity,
            l.condition,
            l.language.upper(),
            round(opp.ct_price_brl, 2),
            l.price_currency,
            round(opp.live_price_brl, 2) if opp.live_price_brl is not None else None,
            round(opp.markup_pct, 4) if opp.markup_pct is not None else None,
            opp.markup_tier or "",
            opp.validation_status,
            round(opp.tcg_market_brl, 2),
            round(opp.tcg_market_usd, 2),
            round(opp.margin_pct, 4),
            round(opp.real_margin_pct, 4) if opp.real_margin_pct is not None else None,
            round(opp.net_margin_pct, 4),
            round(opp.real_net_margin_pct, 4) if opp.real_net_margin_pct is not None else None,
            round(opp.real_lucro_brl, 2) if opp.real_lucro_brl is not None else None,
            round(opp.estimated_shipping_brl, 2),
            l.quantity,
            "Yes" if l.foil else "No",
            opp.price_variant_used or "",
            l.seller_username,
            l.seller_user_type,
            "Yes" if l.seller_can_sell_via_hub else "No",
            l.cardtrader_url,
            opp.tcg_url or "",
            opp.scanned_at,
            # v2.13 — chase + valorização (colunas append, índices 29-32 0-based)
            opp.chase_tier or "",
            round(opp.margin_pct, 4),  # Desconto % (= margem bruta, rótulo chase)
            opp.valorization_score,
            opp.valorization_note or "",
            # v2.14 (candidato 4): "Sim" só quando non-foil casou variante holo.
            "Sim" if opp.variant_low_confidence else "",
        ])

    # ─── Formatação ───
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(*(Side(border_style="thin", color="D1D5DB"),) * 4)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Larguras de coluna ajustadas à leitura
    # v2.8 Layer 4: nova coluna Variant inserida entre Foil e Seller
    widths = [
        28, 28, 8, 12, 8,         # A-E: card/set/num/cond/idioma
        12, 10, 12, 9, 18,        # F-J: scan/moeda/live/markup%/markup_tier
        18,                       # K: validation_status
        14, 14,                   # L-M: TCG BRL/USD
        12, 12, 12, 16,           # N-Q: margem scan/real/net scan/net REAL
        12, 10,                   # R-S: lucro REAL / frete
        6, 6,                     # T-U: qtd/foil
        18,                       # V: variant (holofoil/reverseHolofoil/normal)
        20, 14, 6,                # W-Y: seller/tipo/hub
        40, 40, 18,               # Z-AB: link CT / link TCG / scanned_at
        # v2.13 — AC-AF: chase tier / desconto% / valorização / notas
        12, 11, 16, 60,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Formato número/percentual/moeda
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
        # Colunas monetárias (índices 0-based)
        # v2.8 Layer 4: índices corrigidos após inserção da coluna "Idioma"
        # (idx 5). O bloco anterior estava todo -1 (escrito pré-Idioma).
        row[6].number_format = '"R$"#,##0.00'   # G: Scan R$ (raw)
        row[8].number_format = '"R$"#,##0.00'   # I: LIVE R$ (real)
        row[9].number_format = "0.00%"          # J: Markup %
        row[12].number_format = '"R$"#,##0.00'  # M: TCG Market BRL
        row[13].number_format = "$#,##0.00"     # N: TCG Market USD
        row[14].number_format = "0.00%"         # O: Margem % (scan)
        row[15].number_format = "0.00%"         # P: Margem % REAL
        row[16].number_format = "0.00%"         # Q: Net Margin % (scan)
        row[17].number_format = "0.00%"         # R: Net Margin % REAL
        row[18].number_format = '"R$"#,##0.00'  # S: Lucro R$ REAL
        row[19].number_format = '"R$"#,##0.00'  # T: Frete Est. R$
        # v2.13: Desconto % (idx30 0-based) — formato percentual. Demais novas
        # colunas (Chase Tier idx29, Valorização idx31 int, Notas idx32 texto)
        # ficam no formato geral.
        if len(row) > 30:
            row[30].number_format = "0.00%"
        # v2.8 Layer 4: nova coluna Variant inserida entre Foil e Seller.
        # Headers atuais (29 cols): ... "Foil"(22), "Variant"(23), "Seller"(24),
        # "Tipo Seller"(25), "Hub"(26), "Link CardTrader"(27), "Link TCG"(28),
        # "Scanned At"(29). 0-based: 26=Link CT, 27=Link TCG.
        # Apenas quando célula tem URL não-vazia.
        link_font = Font(color="0563C1", underline="single")
        for col_0 in (26, 27):  # 0-based: Link CardTrader, Link TCG
            if col_0 < len(row):
                cell = row[col_0]
                v = cell.value
                if isinstance(v, str) and v.startswith("http"):
                    cell.hyperlink = v
                    cell.font = link_font

    # Freeze header + filtro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Formatação condicional na coluna Net Margin REAL (R, idx17): verde forte
    # = oportunidade real. Antes apontava p/ Q (Net Margin % scan, otimista
    # pré-validação) — bug de off-by-one igual ao bloco de number_format.
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"R2:R{ws.max_row}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FEE2E2",
                mid_type="num", mid_value=0.15, mid_color="FEF3C7",
                end_type="num", end_value=0.30, end_color="86EFAC",
            ),
        )

    # Aba 2: Stats (auditoria do funnel)
    ws2 = wb.create_sheet("Stats")
    ws2.append(["Métrica", "Valor"])
    for k, v in stats.items():
        ws2.append([k, v])
    ws2.append(["usd_brl_rate", round(usd_brl, 4)])
    ws2.append(["eur_brl_rate", round(eur_brl, 4)])
    ws2.append(["threshold_margin", f"{threshold:.0%}"])
    ws2.append(["hub_fee_rate", f"{hub_fee_rate:.0%}"])
    ws2.append(["margin_basis", "BRUTA (sem taxa)" if hub_fee_rate <= 0 else f"líquida (+{hub_fee_rate:.0%} hub fee)"])
    ws2.append(["scanned_at", datetime.now().isoformat(timespec="seconds")])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(out_path)
    log.info(f"✓ Planilha salva: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser(
        description="CardTrader Arbitrage Scanner — Pokémon TCG Singles",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--sets", nargs="*", help="Códigos de expansão CT (ex: sv1 sv3pt5). Default: todas do config.yaml")
    p.add_argument("--all-sets", action="store_true",
                   help=("Escaneia TODAS as expansões Pokemon do CardTrader (~832), "
                         "ignorando tanto --sets quanto a lista 'sets' do config.yaml. "
                         "Use para o scan COMPLETO (weekly). Sem esta flag e sem --sets, "
                         "o escopo cai na lista do config.yaml."))
    p.add_argument("--skip-backcatalog", action="store_true",
                   help=("Escaneia só as coleções modernas/curadas (PRIORITY_SET_CODES), "
                         "pulando o back-catalog (mercado eficiente, ~0 deal acionável). "
                         "Mais útil com --all-sets: corta ~832 → ~30 sets. Combinado com "
                         "--sets, intersecta com a lista do usuário."))
    p.add_argument("--threshold", type=float, default=MARGIN_THRESHOLD,
                   help=f"Margem mínima bruta (default: {MARGIN_THRESHOLD})")
    p.add_argument("--min-price-usd", type=float, default=MIN_PRICE_USD,
                   help=f"Preço mínimo por carta em USD (default: {MIN_PRICE_USD})")
    p.add_argument("--include-graded", action="store_true",
                   help="Incluir cartas graded (PSA/BGS/CGC). Default: excluir")
    p.add_argument("--provider", choices=list(PROVIDERS.keys()), default="pokemontcg",
                   help="Fonte de preços TCG (default: pokemontcg)")
    p.add_argument("--output", "-o", type=str, default=None,
                   help="Arquivo de saída .xlsx (default: cardtrader_scan_<timestamp>.xlsx)")
    p.add_argument("--dry-run", action="store_true",
                   help="Usa só cache, não chama APIs (debug)")
    p.add_argument("--no-cache", action="store_true",
                   help="Limpa caches voláteis (preços TCG + FX) antes de rodar — força refresh em janelas de arbitragem urgente")
    p.add_argument("--max-expansions", type=int, default=None,
                   help="Limita o número de expansões (teste rápido)")
    p.add_argument("--validate-top", type=int, default=0,
                   help="v2.0: valida os top N candidatos via per-blueprint (preço REAL com markup CT). 0 = desativado. Recomendado: 30")
    p.add_argument("--min-net-margin", type=float, default=0.0,
                   help="v2.0: filtra apenas oportunidades com margem liq REAL >= X apos validacao. Ex: 0.20 = so 20%% liq+")
    p.add_argument("--shipping-brl", type=float, default=0.0,
                   help="Frete fixo per-listing em BRL (default 0 — modelo de "
                        "consolidacao no Hub depot, frete dilui per-card). "
                        "Override pra simular envio direto: ex --shipping-brl 28.84")
    p.add_argument("--hub-fee", type=float, default=0.0,
                   help=("v2.12: DEFAULT 0.0 — margem reportada e BRUTA "
                         "(custo = preco do site, SEM taxa). Operador calcula "
                         "Hub fee/frete/cartao/IOF por fora, manualmente. "
                         "Para reembutir a taxa antiga de 6%%, passe --hub-fee 0.06 "
                         "(custo real = live_brl x (1 + hub_fee))."))
    # v2.4: per-set timeout + skip-list controls
    p.add_argument("--per-set-timeout", type=float, default=DEFAULT_PER_SET_TIMEOUT_MIN,
                   help=(f"v2.4: wall-clock timeout per set (minutos). Default "
                         f"{DEFAULT_PER_SET_TIMEOUT_MIN}. Quando excedido durante pricing loop, "
                         f"set é abortado e adicionado à skip-list. 0 = desativado."))
    p.add_argument("--ignore-skip-list", action="store_true",
                   help="v2.4: ignora scanner_skip_list.json (retenta sets que travaram em runs anteriores).")
    # v2.13 (frente chase cards): filtra pra raridades top.
    p.add_argument("--chase-only", action="store_true",
                   help=("v2.13: só inclui chase cards — raridades TOP/MID "
                         "(Special/Hyper/Secret/Illustration Rare, alt/full art, "
                         "ultra/rainbow/gold rare). Corta common/uncommon/holo "
                         "comum. Mantém piso $10 e validação per-blueprint. "
                         "Output já vem ranqueado por desconto (margem desc)."))
    # v2.13 (frente eficiência): cap de misses consecutivos sem cobertura.
    p.add_argument("--max-consecutive-misses", type=int, default=0,
                   help=("v2.13: aborta um set após N listings consecutivos sem "
                         "preço na fonte (buraco de cobertura — sets 2026/mega "
                         "evolução) e o joga na skip-list. Evita moer centenas de "
                         "chamadas que sempre falham. 0 = desativado (default, "
                         "comportamento histórico). Sugestão p/ --all-sets: 40."))
    p.add_argument("--clear-skip-list", action="store_true",
                   help="v2.4: apaga scanner_skip_list.json antes de rodar (reset total).")
    # v2.6 (2026-05-17): partial JSONL checkpoint crash-recovery
    p.add_argument("--checkpoint-every", type=int, default=10,
                   help=("v2.6: emite checkpoint JSONL append-only a cada N sets "
                         "(default 10). 0 = desativado. Path: "
                         "<state-dir>/<output-name>.checkpoint.jsonl sidecar. "
                         "Crash mid-run não "
                         "perde mais dados — `scripts/recover_from_checkpoint.py "
                         "--checkpoint <path>` regenera XLSX equivalente."))
    # PR-F (2026-05-28): state dir + flush per-listing.
    p.add_argument("--state-dir", type=Path, default=None,
                   help=("v2.11 (PR-F): diretório de estado mutável (cache.db, "
                         "scanner_skip_list.json, heartbeat.txt, *.checkpoint.jsonl). "
                         "Default: %%LOCALAPPDATA%%/CardTraderScanner. Tira estado "
                         "volátil do Google Drive (evita conflitos de sync + lock). "
                         "O XLSX final continua indo pra --output."))
    p.add_argument("--flush-every-listings", type=int, default=50,
                   help=("v2.11 (PR-F): emite linha de telemetria set_progress + "
                         "heartbeat a cada N listings precificados (default 50). "
                         "Distinto de --checkpoint-every (que é por SET). Permite "
                         "detectar stall mid-set. 0 = desativado."))
    p.add_argument("--allow-concurrent", action="store_true",
                   help=("v2.14 (robustez): por padrão o scanner RECUSA iniciar se "
                         "já houver outra instância usando o MESMO state-dir "
                         "(lockfile scanner_run.lock). Dois scanners no mesmo "
                         "cache.db/skip-list disputam lock SQLite (throughput cai a "
                         "~9s/listing) e podem ENVENENAR a skip-list por timeout "
                         "falso. Passe esta flag pra rodar mesmo assim (ex.: "
                         "state-dirs diferentes que o guard não distingue, ou "
                         "diagnóstico). Use com cautela."))
    return p.parse_args()


# Sets de MAIOR PRIORIDADE (era Scarlet & Violet + curados do daily). No modo
# --all-sets, estes são escaneados PRIMEIRO — o catálogo CT vem em ordem
# old→new, e sem reordenar os sets modernos (foco da operação) ficariam no fim
# da fila, sob risco de serem cortados pelo timeout. Códigos que não existirem
# no catálogo simplesmente não casam (inócuo). Ordem aqui = ordem de scan.
PRIORITY_SET_CODES = [
    # Curados (daily) — códigos CardTrader
    "sfa", "scr", "par", "paf", "tef", "twm", "ssp", "dri", "blk", "jtg", "asc",
    # Scarlet & Violet (config.yaml) — sv1..sv10 + meio-sets
    "sv1", "sv2", "sv3", "sv3pt5", "sv4", "sv4pt5", "sv5", "sv6", "sv6pt5",
    "sv7", "sv8", "sv8pt5", "sv9", "sv10", "zsv10pt5", "me2pt5",
]


def filter_modern_sets(
    expansions: list[dict], priority_codes: list[str] = PRIORITY_SET_CODES
) -> list[dict]:
    """Mantém só as expansões cujo `code` está em `priority_codes` (modernas/
    curadas), descartando o back-catalog. É o suporte de `--skip-backcatalog`.

    Lição operacional (auditoria 2026-06-08): back-catalog = mercado eficiente
    = ~0 deal acionável; o gap de arbitragem mora em lançamentos novos. Função
    pura (sem rede) → testável offline. Preserva a ordem de entrada;
    case-insensitive; tolera `code` ausente/None.
    """
    prio = {c.lower() for c in priority_codes}
    return [e for e in expansions if (e.get("code") or "").lower() in prio]


def load_config() -> dict:
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    return {}


def resolve_state_dir(state_dir_arg: Optional[Path]) -> Path:
    """PR-F (2026-05-28): resolve o diretório de estado MUTÁVEL.

    Precedência:
      1. --state-dir explícito (se passado).
      2. %LOCALAPPDATA%/CardTraderScanner (default em Windows).
      3. SCRIPT_DIR (fallback legado quando LOCALAPPDATA ausente — nunca
         quebra a inicialização em CI/não-Windows).

    NÃO cria o diretório (caller faz mkdir) — fica testável puro.
    """
    if state_dir_arg is not None:
        return Path(state_dir_arg)
    localappdata = os.environ.get("LOCALAPPDATA")
    if localappdata:
        return Path(localappdata) / "CardTraderScanner"
    return SCRIPT_DIR


def main():
    args = parse_args()

    # Sanity check: --threshold e --min-net-margin são frações (0.30 = 30%).
    # Trap recorrente: usuário passa "25" achando que é percent → 2500%, scan
    # zera silenciosamente. Auto-converte com warning (2026-05-11 fix).
    if args.threshold > 1.0:
        log.warning(
            f"--threshold {args.threshold} > 1.0 parece percentual, "
            f"convertendo para fração: {args.threshold/100}"
        )
        args.threshold = args.threshold / 100.0
    if args.min_net_margin > 1.0:
        log.warning(
            f"--min-net-margin {args.min_net_margin} > 1.0 parece percentual, "
            f"convertendo para fração: {args.min_net_margin/100}"
        )
        args.min_net_margin = args.min_net_margin / 100.0
    # v2.3: mesma proteção para --hub-fee. Operador pode passar `6` ou `0.06`.
    if args.hub_fee > 1.0:
        log.warning(
            f"--hub-fee {args.hub_fee} > 1.0 parece percentual, "
            f"convertendo para fração: {args.hub_fee/100}"
        )
        args.hub_fee = args.hub_fee / 100.0

    load_dotenv(ENV_FILE)
    cfg = load_config()

    # PR-F (2026-05-28): resolve state dir. Tudo que é estado MUTÁVEL
    # (cache.db, skip-list, heartbeat, checkpoint JSONL) vive aqui — fora do
    # Google Drive, que causa lock-conflict + sync-conflict (.tmp/desktop.ini).
    # O XLSX final (out_path) continua indo pra --output (default SCRIPT_DIR).
    # Reaponta os globais CACHE_DB + SKIP_LIST_FILE pra que as 6 funções de
    # skip-list (que referenciam o nome do módulo em call-time) sigam o novo
    # path. ARMADILHA: Cache.__init__ tem db_path=CACHE_DB como DEFAULT-ARG
    # (bindado no import) — mudar o global não basta; passamos explícito abaixo.
    global CACHE_DB, SKIP_LIST_FILE
    state_dir = resolve_state_dir(args.state_dir)
    state_dir.mkdir(parents=True, exist_ok=True)
    CACHE_DB = state_dir / "cache.db"
    SKIP_LIST_FILE = state_dir / "scanner_skip_list.json"
    log.info(f"State dir: {state_dir}")

    # v2.14 (robustez): run-guard contra instâncias concorrentes no mesmo
    # state-dir. Dois scanners no mesmo cache.db/skip-list disputam lock e
    # envenenam a skip-list por timeout falso (ver acquire_run_guard).
    run_guard = acquire_run_guard(state_dir)
    if run_guard is None:
        if args.allow_concurrent:
            log.warning(
                "⚠️  Outra instância parece estar usando este state-dir "
                f"({state_dir}). --allow-concurrent passado → seguindo mesmo "
                "assim. Risco: contenção de lock SQLite + skip-list envenenada."
            )
        else:
            log.error(
                f"💥 Já há um scanner rodando neste state-dir ({state_dir}). "
                "Dois scanners no mesmo cache.db/skip-list disputam lock "
                "(pricing despenca pra ~9s/listing) e podem ENVENENAR a "
                "skip-list por timeout falso. Aborte a outra instância, OU use "
                "--state-dir diferente, OU passe --allow-concurrent (com cautela)."
            )
            sys.exit(2)

    ct_jwt = os.getenv("CT_JWT", "").strip()
    if not ct_jwt:
        log.error("CT_JWT não definido. Configure no arquivo .env")
        log.error("Como obter: CardTrader → Settings → API Access → Create New Token")
        sys.exit(1)

    # PR-F: db_path EXPLÍCITO (não confiar no default-arg bindado no import).
    cache = Cache(db_path=state_dir / "cache.db")
    if args.no_cache:
        log.info("--no-cache: limpando price_cache e fx_cache (forçando refresh)...")
        cache.clear_prices()
    ct = CardTraderClient(ct_jwt)

    # Pricing provider
    provider_cls = PROVIDERS[args.provider]
    if args.provider == "pokemontcg":
        pricing = provider_cls(os.getenv("POKEMONTCG_API_KEY"), cache)
    elif args.provider == "justtcg":
        pricing = provider_cls(os.getenv("JUSTTCG_API_KEY"), cache)
    else:
        pricing = provider_cls()
    log.info(f"Pricing provider: {pricing.name}")

    # Seleção de expansões
    log.info("Listando expansões Pokemon no CardTrader...")
    all_expansions = ct.list_expansions(CT_POKEMON_GAME_ID)
    log.info(f"Total: {len(all_expansions)} expansões")

    # --all-sets força o escopo COMPLETO: ignora --sets E a lista do config.yaml,
    # caindo no branch `else` abaixo (expansions = all_expansions ~832).
    sets_cfg = None if args.all_sets else (args.sets or cfg.get("sets"))
    if sets_cfg:
        wanted = {s.lower() for s in sets_cfg}
        expansions = [e for e in all_expansions if e.get("code", "").lower() in wanted]
        missing = wanted - {e.get("code", "").lower() for e in expansions}
        if missing:
            log.warning(f"Sets não encontrados: {missing}")
    else:
        expansions = all_expansions

    # v2.17: --skip-backcatalog restringe às coleções modernas/curadas
    # (PRIORITY_SET_CODES), pulando o back-catalog (mercado eficiente, ~0 deal).
    if args.skip_backcatalog:
        before = len(expansions)
        expansions = filter_modern_sets(expansions)
        log.info(
            f"--skip-backcatalog: {len(expansions)}/{before} expansões modernas "
            f"(puladas {before - len(expansions)} de back-catalog)"
        )

    # Reordena no escopo COMPLETO: sets de maior valor (SV/curados) PRIMEIRO.
    # O catálogo CT vem old→new; sem isso, os sets modernos (foco) ficariam no
    # fim e poderiam ser cortados pelo timeout. sort é estável → o resto mantém
    # a ordem do catálogo. Só afeta --all-sets (em --sets a ordem é do usuário).
    if args.all_sets:
        prio = {c.lower(): i for i, c in enumerate(PRIORITY_SET_CODES)}
        expansions.sort(key=lambda e: prio.get((e.get("code") or "").lower(), 10_000))
        matched = [e.get("code") for e in expansions if (e.get("code") or "").lower() in prio]
        log.info(
            f"Prioridade --all-sets: {len(matched)} sets curados/SV primeiro"
            + (f" ({', '.join(matched[:25])})" if matched else "")
        )

    if args.max_expansions:
        expansions = expansions[: args.max_expansions]
    log.info(f"Scan em {len(expansions)} expansões")

    # v2.4: gerencia skip-list antes de criar o Scanner
    if args.clear_skip_list:
        clear_skip_list()
    per_set_timeout_s = args.per_set_timeout * 60 if args.per_set_timeout and args.per_set_timeout > 0 else 0

    # Scanner
    scanner = Scanner(
        ct=ct,
        pricing=pricing,
        cache=cache,
        threshold=args.threshold,
        min_price_usd=args.min_price_usd,
        exclude_graded=not args.include_graded,
        shipping_brl_override=args.shipping_brl,
        hub_fee_rate=args.hub_fee,
        per_set_timeout_s=per_set_timeout_s,
        ignore_skip_list=args.ignore_skip_list,
        chase_only=args.chase_only,
        max_consecutive_misses=args.max_consecutive_misses,
    )

    # v2.6: resolve output_path ANTES de scan() pra calcular checkpoint sidecar.
    # XLSX final continua no --output (default SCRIPT_DIR / Drive). Só o JSONL
    # de checkpoint + heartbeat migram pro state_dir (PR-F).
    out_path = Path(args.output) if args.output else (
        SCRIPT_DIR / f"cardtrader_scan_{datetime.now():%Y%m%d_%H%M}.xlsx"
    )
    # PR-F: checkpoint sidecar vive no state_dir (fora do Drive), nomeado a
    # partir do XLSX final pra rastreabilidade (<output-name>.checkpoint.jsonl).
    checkpoint_path = state_dir / (out_path.name + ".checkpoint.jsonl")
    heartbeat_path = state_dir / "heartbeat.txt"
    checkpoint = CheckpointWriter(
        checkpoint_path,
        every_n=args.checkpoint_every,
        heartbeat_path=heartbeat_path,
        flush_every_listings=args.flush_every_listings,
    )
    if checkpoint.enabled:
        # Header com args dict pra reprodutibilidade no recovery.
        # v2.14: inclui FX do scan (usd_brl/eur_brl) pra o recovery reconstruir
        # a célula Stats usd_brl_rate (candidato 3).
        checkpoint.write_header(
            vars(args), total_sets=len(expansions),
            usd_brl=scanner.usd_brl, eur_brl=scanner.eur_brl,
        )

    t0 = time.time()
    try:
        opps = scanner.scan(expansions, checkpoint=checkpoint)
    finally:
        # Flush + close mesmo em exception. fsync já roda per-write.
        pass
    dt = time.time() - t0
    if checkpoint.enabled:
        checkpoint.write_scan_complete(total_opps=len(opps), total_elapsed_s=dt)
        checkpoint.close()
    log.info(f"Scan completo em {dt:.1f}s — {len(opps)} oportunidades ≥ {args.threshold:.0%}")
    if args.hub_fee > 0:
        log.info(f"Hub fee aplicado no recalc REAL: {args.hub_fee:.0%} (custo = site_price × {1 + args.hub_fee:.2f})")
    else:
        log.info("Margem BRUTA (--hub-fee 0.0): custo = preço do site, SEM taxa. "
                 "Operador calcula Hub fee/frete/cartão/IOF por fora (v2.12).")

    # v2.0 — validação per-blueprint dos top N
    if args.validate_top > 0 and opps:
        t1 = time.time()
        scanner.validate_per_blueprint(opps, top_n=args.validate_top)
        log.info(f"Validação per-blueprint completa em {time.time()-t1:.1f}s")

        # M1 fix (2026-05-11): sempre filtra status inválidos (STALE, API_ERROR,
        # PRICE_CHANGED) quando há validação, independente do min_net_margin.
        # Antes só rodava quando min_net_margin > 0, deixando lixo no XLSX.
        before_status = len(opps)
        opps = [
            o for o in opps
            if o.validation_status in ("VALIDATED_REAL", "VALIDATED_MARKUP")
        ]
        dropped = before_status - len(opps)
        if dropped > 0:
            log.info(f"Filtro de status (STALE/API_ERROR/PRICE_CHANGED): {before_status} → {len(opps)} oportunidades")

        # Filtro adicional por margem líq REAL (só se min_net_margin > 0)
        if args.min_net_margin > 0:
            before = len(opps)
            opps = [
                o for o in opps
                if (o.real_net_margin_pct or 0) >= args.min_net_margin
            ]
            log.info(f"Filtro --min-net-margin {args.min_net_margin:.0%}: {before} → {len(opps)} oportunidades")

        # M2 fix (2026-05-11): sempre re-ordena por margem líq REAL desc
        # quando houve validação, não só quando min_net_margin > 0.
        opps.sort(key=lambda o: o.real_net_margin_pct or -999, reverse=True)

    # Export — out_path já resolvido acima (v2.6 fix: precisa antes do scan
    # pra calcular .checkpoint.jsonl sidecar)
    export_xlsx(opps, scanner.stats, out_path,
                usd_brl=scanner.usd_brl, eur_brl=scanner.eur_brl,
                threshold=args.threshold, hub_fee_rate=args.hub_fee)

    # Top 5 por margem líquida REAL se validado, senão líquida estimada
    def net_key(o: Opportunity) -> float:
        return o.real_net_margin_pct if o.real_net_margin_pct is not None else o.net_margin_pct
    opps_by_net = sorted(opps, key=net_key, reverse=True)

    # Resumo no terminal
    print("\n" + "═" * 60)
    print(f"  RESUMO — {datetime.now():%Y-%m-%d %H:%M}")
    print("═" * 60)
    print(f"  Expansões escaneadas   : {scanner.stats['expansions_scanned']}")
    print(f"  Listings baixados       : {scanner.stats['listings_fetched']}")
    print(f"  Após filtros           : {scanner.stats['listings_after_filters']}")
    print(f"  Com preço TCG          : {scanner.stats['tcg_price_found']}")
    print(f"  Oportunidades ≥ {args.threshold:.0%}   : {scanner.stats['opportunities_found']}")
    print(f"  Câmbio USD→BRL          : {scanner.usd_brl:.4f}")
    print(f"  Câmbio EUR→BRL          : {scanner.eur_brl:.4f}")
    # v2.9 (Codex H5): expor pricing failures pra operador detectar drift cedo
    pricing_failures = scanner.stats.get("pricing_failures", 0)
    if pricing_failures > 0:
        attempts = scanner.stats.get("listings_after_filters", 0)
        pct = (pricing_failures / attempts * 100) if attempts > 0 else 0
        marker = "  ⚠️  " if pricing_failures >= 5 else "  "
        print(f"{marker}Pricing failures        : {pricing_failures} ({pct:.1f}% das tentativas)")
        if pricing_failures >= 5:
            print(f"      → revisar log WARNING acima (schema drift / SSL / endpoint down?)")
    if scanner.stats.get("expansions_mass_pricing_abort", 0) > 0:
        print(f"  💥 Sets abortados (mass pricing fail): {scanner.stats['expansions_mass_pricing_abort']}")
    print(f"  Planilha                : {out_path}")
    print("═" * 60)

    if opps_by_net:
        validated_label = " (validadas v2.0)" if args.validate_top > 0 else ""
        print(f"\n  TOP 5 OPORTUNIDADES{validated_label}:\n")
        for i, o in enumerate(opps_by_net[:5], 1):
            l = o.listing
            # Frete: por default = 0 (modelo consolidação Hub depot). Mostra
            # parêntese só quando há override via --shipping-brl > 0.
            freight_suffix = (
                f" c/ frete (R${o.estimated_shipping_brl:.2f})"
                if o.estimated_shipping_brl > 0 else ""
            )
            print(f"  {i}. {l.card_name} ({l.set_code} {l.collector_number})")
            if o.live_price_brl is not None:
                print(f"     Scan R${o.ct_price_brl:.2f} → LIVE R${o.live_price_brl:.2f}  [{o.markup_tier}]")
                print(f"     TCG: R${o.tcg_market_brl:.2f} (${o.tcg_market_usd:.2f})  | Lucro REAL: R${o.real_lucro_brl:.2f}")
                print(f"     Margem REAL: {o.real_margin_pct:.1%} bruta | {o.real_net_margin_pct:.1%} líquida{freight_suffix}")
            else:
                print(f"     CT: R${o.ct_price_brl:.2f} ({l.price_currency})  →  TCG: R${o.tcg_market_brl:.2f} (${o.tcg_market_usd:.2f})")
                print(f"     Margem: {o.margin_pct:.1%} bruta | {o.net_margin_pct:.1%} líquida{freight_suffix}  [não validado]")
            print(f"     Seller: {l.seller_username} ({l.seller_user_type})")
            print(f"     Link: {l.cardtrader_url}\n")


if __name__ == "__main__":
    main()
