"""PR-J: células de preço (Preço CT, TCG) viram hyperlink → página de conferência.
Roda via pytest E standalone."""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
import pandas as pd
import cardtrader_postprocess as pp


def _ws_from_df(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append(list(r))
    return ws


def test_price_cells_get_hyperlinks():
    df = pd.DataFrame({
        "Carta": ["Plusle 193"],
        "Preço CT (R$)": [178.62],
        "TCG (R$)": [270.22],
        "Link CT": ["https://www.cardtrader.com/cards/265388"],
        "Link TCG": ["https://prices.pokemontcg.io/tcgplayer/sv4-193"],
    })
    ws = _ws_from_df(df)
    pp.apply_card_hyperlinks(ws, df)
    cols = list(df.columns)
    pc = ws.cell(row=2, column=cols.index("Preço CT (R$)") + 1)
    tc = ws.cell(row=2, column=cols.index("TCG (R$)") + 1)
    assert pc.hyperlink and "cardtrader.com" in pc.hyperlink.target
    assert tc.hyperlink and "pokemontcg.io" in tc.hyperlink.target
    # valor numérico preservado (só a célula virou clicável)
    assert pc.value == 178.62 and tc.value == 270.22


def test_no_link_cols_is_noop():
    df = pd.DataFrame({"Preço CT (R$)": [10.0], "TCG (R$)": [20.0]})
    ws = _ws_from_df(df)
    pp.apply_card_hyperlinks(ws, df)  # sem Link CT/TCG → não deve quebrar
    assert ws.cell(row=2, column=1).hyperlink is None


if __name__ == "__main__":
    import traceback
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    passed = 0
    for fn in fns:
        try:
            fn()
            passed += 1
            print(f"PASS {fn.__name__}")
        except Exception:
            print(f"FAIL {fn.__name__}")
            traceback.print_exc()
    print(f"{passed}/{len(fns)} passed")
    sys.exit(0 if passed == len(fns) else 1)
