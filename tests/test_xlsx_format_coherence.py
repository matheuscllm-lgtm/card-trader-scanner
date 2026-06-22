#!/usr/bin/env python3
"""Review-fix tests (2026-05-29) — coerência de formatação do XLSX + recovery.

Cobre 3 achados de revisão:

  FIX 1 — formatação condicional apontava p/ coluna Q (Net Margin % SCAN,
          otimista pré-validação) em vez de R (Net Margin % REAL, idx17).
  FIX 2 — bloco de `number_format` estava todo off-by-one (-1), escrito antes
          da coluna "Idioma" (idx5) ser inserida. LIVE R$ (idx8) deve ser
          money, Markup (idx9) "0.00%", Net REAL (idx17) "0.00%" (NÃO money).
  FIX 3 — recover_from_checkpoint perdia `price_variant_used` ao reconstruir
          a Opportunity.

Mapa de headers (0-based, 29 cols, verificado em cardtrader_scanner.py):
  6=Scan R$(raw)→G, 8=LIVE R$→I, 9=Markup%→J, 12=TCG BRL→M, 13=TCG USD→N,
  14=Margem%scan→O, 15=Margem%REAL→P, 16=Net%scan→Q, 17=Net%REAL→R,
  18=Lucro→S, 19=Frete→T.

Roda de dois jeitos (espelha tests existentes):
    pytest tests/test_xlsx_format_coherence.py -v
    python tests/test_xlsx_format_coherence.py     # fallback standalone
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

# Repo root no sys.path (tests/ → ..)
REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import cardtrader_scanner as sc  # noqa: E402
from cardtrader_scanner import Listing, Opportunity, export_xlsx  # noqa: E402

# scripts/ no sys.path p/ importar o recovery
SCRIPTS_DIR = REPO_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))


# ─────────────────────────── fixtures sintéticas ───────────────────────────
def _make_listing(foil: bool = False) -> Listing:
    return Listing(
        product_id=111,
        blueprint_id=222,
        card_name="Pichu",
        set_code="ecard1",
        set_name="Expedition",
        collector_number="22",
        condition="Near Mint",
        language="en",
        price_cents=5000,
        price_currency="EUR",
        price_brl=300.0,
        quantity=1,
        foil=foil,
        graded=False,
        seller_username="someseller",
        seller_can_sell_via_hub=True,
        seller_user_type="professional",
        cardtrader_url="https://cardtrader.com/cards/222",
        rarity="Holo Rare",
    )


def _make_opp(foil: bool = False) -> Opportunity:
    return Opportunity(
        listing=_make_listing(foil=foil),
        tcg_market_usd=120.0,
        tcg_market_brl=600.0,
        ct_price_brl=300.0,
        margin_pct=0.50,
        margin_brl=300.0,
        estimated_shipping_brl=0.0,
        net_margin_pct=0.50,
        validation_status="VALIDATED",
        live_price_brl=318.0,
        real_margin_pct=0.47,
        real_net_margin_pct=0.47,
        real_lucro_brl=282.0,
        markup_pct=0.06,
        markup_tier="6%",
        tcg_url="https://tcgplayer.com/product/123",
        price_variant_used="holofoil",
    )


def _build_xlsx(tmp_path: Path) -> Path:
    opps = [_make_opp(foil=True), _make_opp(foil=False)]
    out = tmp_path / "deals.xlsx"
    export_xlsx(
        opps,
        stats={"listings_scanned": 2},
        out_path=out,
        usd_brl=5.05,
        eur_brl=5.50,
        threshold=0.25,
    )
    return out


# ─────────────── Regressão: export_xlsx cria o diretório-alvo ────────────────
def test_export_xlsx_creates_missing_output_dir(tmp_path):
    """export_xlsx deve CRIAR o diretório-alvo se ele não existir.

    outputs/ é gitignored e NÃO vem num clone limpo (sessão Claude Code na
    nuvem) — antes do fix o wb.save() quebrava com FileNotFoundError DEPOIS de um
    rastreio inteiro (horas no --all-sets), perdendo todo o trabalho em memória.
    """
    missing = tmp_path / "outputs" / "nested"  # ainda NÃO existe
    assert not missing.exists(), "pré-condição: dir-alvo não deve existir"
    out = missing / "deals.xlsx"
    export_xlsx(
        [_make_opp(foil=True)],
        stats={"listings_scanned": 1},
        out_path=out,
        usd_brl=5.05,
        eur_brl=5.50,
        threshold=0.25,
    )  # não deve levantar FileNotFoundError
    assert out.exists(), "xlsx não foi criado (dir-alvo não foi gerado)"


# ─────────────────────────── FIX 2: number_format ──────────────────────────
def test_number_format_money_percent_coherence(tmp_path):
    """LIVE R$ (I/idx8) money, Markup (J/idx9) %, Net REAL (R/idx17) % (NÃO money)."""
    out = _build_xlsx(tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb["Oportunidades"]

    # Linha 2 = primeira data row. Colunas Excel são 1-based; idx0 +1 = col.
    cell_live = ws.cell(row=2, column=8 + 1)   # I: LIVE R$ (real)
    cell_markup = ws.cell(row=2, column=9 + 1)  # J: Markup %
    cell_net_real = ws.cell(row=2, column=17 + 1)  # R: Net Margin % REAL

    assert "R$" in cell_live.number_format, (
        f"col I (LIVE R$) devia ser money, got {cell_live.number_format!r}"
    )
    assert cell_markup.number_format == "0.00%", (
        f"col J (Markup) devia ser 0.00%, got {cell_markup.number_format!r}"
    )
    assert cell_net_real.number_format == "0.00%", (
        f"col R (Net REAL) devia ser 0.00%, got {cell_net_real.number_format!r}"
    )
    # Guarda anti-regressão: Net REAL NÃO pode ser money.
    assert "R$" not in cell_net_real.number_format
    wb.close()


def test_number_format_full_layout(tmp_path):
    """Valida todos os índices do bloco corrigido contra o layout real."""
    out = _build_xlsx(tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb["Oportunidades"]
    r = 2  # primeira data row

    expected = {
        6: ("money", "G Scan R$ raw"),
        8: ("money", "I LIVE R$"),
        9: ("0.00%", "J Markup %"),
        12: ("money", "M TCG BRL"),
        13: ("$#,##0.00", "N TCG USD"),
        14: ("0.00%", "O Margem scan"),
        15: ("0.00%", "P Margem REAL"),
        16: ("0.00%", "Q Net scan"),
        17: ("0.00%", "R Net REAL"),
        18: ("money", "S Lucro"),
        19: ("money", "T Frete"),
    }
    for idx0, (kind, label) in expected.items():
        fmt = ws.cell(row=r, column=idx0 + 1).number_format
        if kind == "money":
            assert "R$" in fmt, f"{label} (idx{idx0}) devia ser money R$, got {fmt!r}"
        elif kind == "$#,##0.00":
            assert fmt == "$#,##0.00", f"{label} (idx{idx0}) got {fmt!r}"
        else:
            assert fmt == kind, f"{label} (idx{idx0}) devia ser {kind}, got {fmt!r}"
    wb.close()


# ───────────────────── FIX 1: conditional formatting R não Q ────────────────
def test_conditional_formatting_covers_R_not_Q(tmp_path):
    """ColorScale deve cobrir R2:R (Net REAL, idx17), nunca Q2:Q (Net scan)."""
    out = _build_xlsx(tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb["Oportunidades"]

    # ws.conditional_formatting itera objetos ConditionalFormatting; o range
    # coberto está em `.sqref` (ex MultiCellRange "R2:R3").
    ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    joined = " ".join(ranges)

    assert any(rng.startswith("R2") for rng in ranges), (
        f"esperava conditional_formatting começando em R2, got {ranges!r}"
    )
    assert not any(rng.startswith("Q2") for rng in ranges), (
        f"conditional_formatting NÃO pode estar em Q (Net scan), got {ranges!r}"
    )
    assert "R2:R" in joined, f"esperava range R2:R{ws.max_row}, got {joined!r}"
    wb.close()


# ───────────────────── FIX 3: recovery preserva variant ─────────────────────
def _recovery_dict() -> dict:
    """Dict de checkpoint mínimo (formato .checkpoint.jsonl) p/ reconstruir."""
    return {
        "listing": {
            "product_id": 111,
            "blueprint_id": 222,
            "card_name": "Pichu",
            "set_code": "ecard1",
            "set_name": "Expedition",
            "collector_number": "22",
            "condition": "Near Mint",
            "language": "en",
            "price_cents": 5000,
            "price_currency": "EUR",
            "price_brl": 300.0,
            "quantity": 1,
            "foil": True,
            "graded": False,
            "seller_username": "someseller",
            "seller_can_sell_via_hub": True,
            "seller_user_type": "professional",
            "cardtrader_url": "https://cardtrader.com/cards/222",
            "rarity": "Holo Rare",
        },
        "tcg_market_usd": 120.0,
        "tcg_market_brl": 600.0,
        "ct_price_brl": 300.0,
        "margin_pct": 0.50,
        "margin_brl": 300.0,
        "estimated_shipping_brl": 0.0,
        "net_margin_pct": 0.50,
        "validation_status": "VALIDATED",
        "live_price_brl": 318.0,
        "real_margin_pct": 0.47,
        "real_net_margin_pct": 0.47,
        "real_lucro_brl": 282.0,
        "markup_pct": 0.06,
        "markup_tier": "6%",
        "tcg_url": "https://tcgplayer.com/product/123",
        "price_variant_used": "holofoil",
    }


def test_recovery_preserves_price_variant_used():
    import recover_from_checkpoint as rec

    opp = rec._reconstruct_opportunity(_recovery_dict())
    assert opp.price_variant_used == "holofoil", (
        f"recovery perdeu price_variant_used, got {opp.price_variant_used!r}"
    )


def test_recovery_variant_missing_defaults_none():
    """Checkpoint pré-v2.8 (sem o campo) → default None, sem crash."""
    import recover_from_checkpoint as rec

    d = _recovery_dict()
    del d["price_variant_used"]
    opp = rec._reconstruct_opportunity(d)
    assert opp.price_variant_used is None
    # tcg_url ainda preservado (não quebrou o get-com-default existente).
    assert opp.tcg_url == "https://tcgplayer.com/product/123"


# ─────────────────────────── standalone fallback ───────────────────────────
if __name__ == "__main__":
    import tempfile

    failures = 0
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        for fn in (
            test_export_xlsx_creates_missing_output_dir,
            test_number_format_money_percent_coherence,
            test_number_format_full_layout,
            test_conditional_formatting_covers_R_not_Q,
        ):
            try:
                fn(tp)
                print(f"PASS {fn.__name__}")
            except AssertionError as e:
                failures += 1
                print(f"FAIL {fn.__name__}: {e}")
    for fn in (
        test_recovery_preserves_price_variant_used,
        test_recovery_variant_missing_defaults_none,
    ):
        try:
            fn()
            print(f"PASS {fn.__name__}")
        except AssertionError as e:
            failures += 1
            print(f"FAIL {fn.__name__}: {e}")
    sys.exit(1 if failures else 0)
