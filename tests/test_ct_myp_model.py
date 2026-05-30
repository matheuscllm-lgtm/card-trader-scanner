"""PR-K: alinhamento CT postprocess com modelo MYP.

Cobre:
  - _combine_name_number formato "Plusle (193/197)" com total mapeado
  - Fallback "Plusle (193)" quando set_code não mapeado
  - CT_SET_TOTAL contém os 12 sets do daily moderno
  - 3 sheets novas (Top 50 Margin, Validate Manually, TCG Suspect) existem no XLSX final
  - Ordenação Top 50 por Net % desc
  - TG## vai pra Validate Manually
  - Set vintage (lc) vai pra TCG Suspect

Roda via pytest E standalone (python tests/test_ct_myp_model.py).
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
import pandas as pd
import cardtrader_postprocess as pp


# ─── FIX 1: nome formato (NNN/MMM) ────────────────────────────────────────────

def test_combine_with_total_par():
    df = pd.DataFrame({
        "card_name": ["Plusle"],
        "card_number": [193],
        "set_code": ["Paradox Rift (par)"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Plusle (193/182)"
    # car_number preservado pra coluna Nº
    assert out["card_number"].iloc[0] == 193


def test_combine_with_total_scr():
    df = pd.DataFrame({
        "card_name": ["Milcery"],
        "card_number": [152],
        "set_code": ["Stellar Crown (scr)"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Milcery (152/142)"


def test_combine_set_code_unmapped_fallback():
    """set_code não mapeado → 'Nome (NNN)' sem total."""
    df = pd.DataFrame({
        "card_name": ["Foo"],
        "card_number": [42],
        "set_code": ["Random Set (xyz)"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Foo (42)"


def test_combine_set_code_direct_code():
    """set_code direto (sem parens) também resolve via lower-case match."""
    df = pd.DataFrame({
        "card_name": ["Pikachu"],
        "card_number": [180],
        "set_code": ["asc"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Pikachu (180/217)"


def test_combine_no_set_column_fallback():
    """Sem coluna set_code → fallback 'Nome (NNN)' sem total."""
    df = pd.DataFrame({"card_name": ["Bar"], "card_number": [99]})
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Bar (99)"


def test_combine_no_number():
    df = pd.DataFrame({
        "card_name": ["Mew"],
        "card_number": [None],
        "set_code": ["Paradox Rift (par)"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Mew"


def test_combine_nan_number():
    df = pd.DataFrame({
        "card_name": ["X"],
        "card_number": [float("nan")],
        "set_code": ["par"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "X"


def test_combine_float_strips_dot_zero():
    df = pd.DataFrame({
        "card_name": ["Plusle"],
        "card_number": [193.0],
        "set_code": ["par"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Plusle (193/182)"


def test_combine_alphanumeric_tg():
    df = pd.DataFrame({
        "card_name": ["Trainer"],
        "card_number": ["TG12"],
        "set_code": ["scr"],
    })
    out = pp._combine_name_number(df)
    assert out["card_name"].iloc[0] == "Trainer (TG12/142)"


def test_ct_set_total_covers_daily_sets():
    """Os 12 sets do daily moderno (par-blk) precisam estar mapeados."""
    expected = {"par", "paf", "tef", "twm", "sfa", "scr", "ssp",
                "dri", "jtg", "pre", "asc", "blk"}
    missing = expected - set(pp.CT_SET_TOTAL.keys())
    assert not missing, f"Sets faltando em CT_SET_TOTAL: {missing}"
    # Todos os valores > 0
    for code, total in pp.CT_SET_TOTAL.items():
        assert total > 0, f"Total {code}={total} inválido"


# ─── FIX 2: 3 sheets MYP-style ───────────────────────────────────────────────

def _build_realistic_raw_df():
    """Raw DF imitando colunas do scanner output, várias categorias."""
    return pd.DataFrame({
        "card_name": ["Plusle", "Trainer", "Cheap", "Hot", "Vintage", "InflatedTcg"],
        "card_number": [193, "TG12", 10, 50, 5, 22],
        "set_code": [
            "Paradox Rift (par)",   # normal
            "Stellar Crown (scr)",  # TG##
            "Stellar Crown (scr)",  # baseline cheap
            "Stellar Crown (scr)",  # high margin
            "Legendary Collection (lc)",  # vintage
            "Stellar Crown (scr)",  # tcg 20× mediana
        ],
        "rarity": ["Illustration Rare", "Trainer Gallery", "Common",
                   "Special Illustration Rare", "Holo Rare", "Holo Rare"],
        "language": ["EN"] * 6,
        "live_brl": [100.0, 50.0, 5.0, 50.0, 20.0, 30.0],
        "reference_price_brl": [180.0, 200.0, 10.0, 500.0, 100.0, 5000.0],
        "validation_status": ["VALIDATED_REAL", "VALIDATED_REAL",
                              "VALIDATED_REAL", "STALE",
                              "VALIDATED_REAL", "VALIDATED_REAL"],
        "seller": ["s1", "s2", "s3", "s4", "s5", "s6"],
        "link_ct": ["https://www.cardtrader.com/cards/1"] * 6,
        "link_tcg": ["https://prices.pokemontcg.io/tcgplayer/sv4-193"] * 6,
        "quantity": [1] * 6,
    })


def _write_and_load(df):
    """Escreve via write_report num tmpfile e reabre como Workbook."""
    cfg = pp.DecisionConfig()
    tmp = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    try:
        pp.write_report(df, cfg, tmp)
        wb = openpyxl.load_workbook(tmp)
        return wb
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass


def test_three_new_sheets_exist():
    wb = _write_and_load(_build_realistic_raw_df())
    names = wb.sheetnames
    assert "Top 50 Margin" in names
    assert "Validate Manually" in names
    assert "TCG Suspect" in names
    # Ordem esperada: Deals, All Listings, Top 50 Margin, Validate Manually,
    # TCG Suspect, Summary
    assert names.index("Top 50 Margin") > names.index("All Listings")
    assert names.index("Summary") > names.index("TCG Suspect")


def test_top50_sorted_by_net_desc():
    wb = _write_and_load(_build_realistic_raw_df())
    ws = wb["Top 50 Margin"]
    headers = [c.value for c in ws[1]]
    assert "Net %" in headers
    net_idx = headers.index("Net %") + 1
    vals = [ws.cell(row=r, column=net_idx).value for r in range(2, ws.max_row + 1)]
    vals_num = [v for v in vals if isinstance(v, (int, float))]
    assert vals_num == sorted(vals_num, reverse=True), f"Top 50 não está desc: {vals_num}"


def test_validate_manually_includes_tg():
    wb = _write_and_load(_build_realistic_raw_df())
    ws = wb["Validate Manually"]
    headers = [c.value for c in ws[1]]
    assert "Nº" in headers
    nº_idx = headers.index("Nº") + 1
    nums = [str(ws.cell(row=r, column=nº_idx).value) for r in range(2, ws.max_row + 1)]
    assert any(n.upper().startswith("TG") for n in nums), \
        f"TG## não apareceu em Validate Manually: {nums}"


def test_tcg_suspect_includes_vintage():
    wb = _write_and_load(_build_realistic_raw_df())
    ws = wb["TCG Suspect"]
    headers = [c.value for c in ws[1]]
    assert "Set" in headers
    set_idx = headers.index("Set") + 1
    sets = [str(ws.cell(row=r, column=set_idx).value).lower()
            for r in range(2, ws.max_row + 1)]
    assert any("(lc)" in s or s == "lc" for s in sets), \
        f"Set vintage 'lc' não apareceu em TCG Suspect: {sets}"


def test_new_sheets_have_card_columns():
    """Sheets novas usam MESMAS colunas das sheets existentes."""
    wb = _write_and_load(_build_realistic_raw_df())
    base_headers = [c.value for c in wb["All Listings"][1]]
    for sheet_name in ("Top 50 Margin", "Validate Manually", "TCG Suspect"):
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        first_row = [c.value for c in ws[1]]
        # Se vazia, primeira célula é placeholder
        if first_row == ["Nenhum listing nesta categoria."] + [None] * (len(first_row) - 1):
            continue
        assert first_row == base_headers, \
            f"{sheet_name} headers diferem de All Listings:\n{first_row}\nvs\n{base_headers}"


if __name__ == "__main__":
    import traceback
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    passed = 0
    for fn in fns:
        try:
            fn()
            passed += 1
            print(f"PASS {fn.__name__}")
        except Exception:
            print(f"FAIL {fn.__name__}")
            traceback.print_exc()
    print(f"{passed}/{len(fns)} passed")
    sys.exit(0 if passed == len(fns) else 1)
