#!/usr/bin/env python3
"""Testes v2.13 — frentes chase cards + valorização médio/longo prazo.

Cobre:
  • classify_chase_tier: taxonomia de raridade → TOP/MID/MODEST/BULK/"".
  • compute_valorization: score 0-100 honesto (rarity + idade + preço), com
    nota explicativa e SEM inventar série histórica.
  • _valorization_age_component: janela de maturidade do set.
  • export_xlsx: novas colunas (Chase Tier / Desconto % / Valorização / Notas)
    aparecem no fim, sem deslocar as colunas existentes (índices 6-19/26/27
    intactos — contrato dos testes de coerência).

Roda de dois jeitos:
    pytest tests/test_chase_valorization.py -v
    python tests/test_chase_valorization.py     # fallback standalone
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import cardtrader_scanner as sc  # noqa: E402
from cardtrader_scanner import (  # noqa: E402
    Listing, Opportunity, classify_chase_tier, compute_valorization,
    _valorization_age_component, export_xlsx, CHASE_ONLY_TIERS,
)


# ───────────────────────────── chase tier ─────────────────────────────
def test_chase_tier_top():
    for r in ["Special Illustration Rare", "SIR", "Hyper Rare",
              "Secret Rare", "Illustration Rare"]:
        assert classify_chase_tier(r) == "TOP", r


def test_chase_tier_mid():
    for r in ["Full Art", "Alt Art", "Rainbow Rare", "Gold Rare",
              "Double Rare", "Ultra Rare"]:
        assert classify_chase_tier(r) == "MID", r


def test_chase_tier_modest():
    for r in ["Holo Rare", "Reverse Holo", "Promo", "Rare Holo"]:
        assert classify_chase_tier(r) == "MODEST", r


def test_chase_tier_bulk():
    for r in ["Common", "Comum", "Uncommon", "Incomum"]:
        assert classify_chase_tier(r) == "BULK", r


def test_chase_tier_unknown_rarity_is_modest():
    # Rarity presente mas não mapeada → MODEST (conservador).
    assert classify_chase_tier("Radiant Rare") == "MODEST"


def test_chase_tier_empty_is_blank():
    # Sem rarity → "" (não-chase pelo filtro).
    assert classify_chase_tier("") == ""
    assert classify_chase_tier(None) == ""


def test_chase_only_tiers_are_top_mid():
    assert CHASE_ONLY_TIERS == {"TOP", "MID"}


# ─────────────────────────── valorização ──────────────────────────────
def test_age_component_mature_set_scores_highest():
    now = datetime(2026, 6, 9)
    # Set de ~3 anos = janela madura → 35.
    pts, note = _valorization_age_component("2023/06/01", now=now)
    assert pts == 35
    assert "maduro" in note


def test_age_component_brand_new_scores_low():
    now = datetime(2026, 6, 9)
    pts, _ = _valorization_age_component("2026/05/01", now=now)
    assert pts == 12  # <6 meses


def test_age_component_missing_date_is_neutral():
    pts, note = _valorization_age_component(None)
    assert pts == 10
    assert "desconhecida" in note


def test_age_component_handles_dash_format():
    now = datetime(2026, 6, 9)
    pts, _ = _valorization_age_component("2023-06-01", now=now)
    assert pts == 35


def test_valorization_top_rarity_premium_price_high_score():
    now = datetime(2026, 6, 9)
    score, note = compute_valorization(
        "Special Illustration Rare", "2023/06/01", 150.0, now=now
    )
    # 45 (TOP) + 35 (maduro) + 20 (premium) = 100.
    assert score == 100
    assert "chase forte" in note
    # Honestidade: a nota deixa explícito que é heurística sem histórico.
    assert "sem série histórica" in note


def test_valorization_bulk_cheap_new_low_score():
    now = datetime(2026, 6, 9)
    score, _ = compute_valorization("Common", "2026/05/01", 11.0, now=now)
    # 0 (BULK) + 12 (novo) + 4 (barato) = 16.
    assert score == 16


def test_valorization_never_invents_historical_data():
    # Garante que a nota NUNCA afirma "alta histórica" / "% do topo" — dados
    # que a fonte não fornece. (Anti-fabricação de dados.)
    _, note = compute_valorization("Ultra Rare", "2024/01/01", 50.0)
    lowered = note.lower()
    for forbidden in ["alta histórica", "topo histórico", "% do topo", "ath"]:
        assert forbidden not in lowered


# ─────────────────────── export_xlsx novas colunas ────────────────────
def _make_opp() -> Opportunity:
    listing = Listing(
        product_id=1, blueprint_id=2, card_name="Charizard ex",
        set_code="sv3pt5", set_name="151", collector_number="199",
        condition="Near Mint", language="en", price_cents=10000,
        price_currency="EUR", price_brl=500.0, quantity=1, foil=True,
        graded=False, seller_username="s", seller_can_sell_via_hub=True,
        seller_user_type="professional",
        cardtrader_url="https://cardtrader.com/cards/2",
        rarity="Special Illustration Rare",
    )
    return Opportunity(
        listing=listing, tcg_market_usd=200.0, tcg_market_brl=1010.0,
        ct_price_brl=500.0, margin_pct=0.50, margin_brl=510.0,
        estimated_shipping_brl=0.0, net_margin_pct=0.50,
        chase_tier="TOP", valorization_score=100,
        valorization_note="raridade top (chase forte); set maduro; preço-âncora premium",
        set_release_date="2023/06/16",
    )


def test_xlsx_has_new_columns_at_end(tmp_path):
    out = tmp_path / "deals.xlsx"
    export_xlsx([_make_opp()], stats={"x": 1}, out_path=out,
                usd_brl=5.05, eur_brl=5.50, threshold=0.30)
    wb = openpyxl.load_workbook(out)
    ws = wb["Oportunidades"]
    headers = [c.value for c in ws[1]]
    # Novas colunas existem.
    for h in ["Chase Tier", "Desconto %", "Valorização (0-100)", "Valorização — Notas"]:
        assert h in headers, f"coluna {h} ausente: {headers}"
    # E estão NO FIM (depois de Scanned At) — não deslocaram as antigas.
    assert headers.index("Scanned At") < headers.index("Chase Tier")
    # Contrato de índices antigos preservado.
    assert headers[6] == "Scan R$ (raw)"
    assert headers[17] == "Net Margin % REAL"
    wb.close()


def test_xlsx_new_column_values(tmp_path):
    out = tmp_path / "deals.xlsx"
    export_xlsx([_make_opp()], stats={"x": 1}, out_path=out,
                usd_brl=5.05, eur_brl=5.50, threshold=0.30)
    wb = openpyxl.load_workbook(out)
    ws = wb["Oportunidades"]
    headers = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert row[headers.index("Chase Tier")] == "TOP"
    assert row[headers.index("Valorização (0-100)")] == 100
    assert abs(row[headers.index("Desconto %")] - 0.50) < 1e-9
    wb.close()


# ─────────────────────────── standalone fallback ───────────────────────
if __name__ == "__main__":
    import tempfile

    no_arg = [
        test_chase_tier_top, test_chase_tier_mid, test_chase_tier_modest,
        test_chase_tier_bulk, test_chase_tier_unknown_rarity_is_modest,
        test_chase_tier_empty_is_blank, test_chase_only_tiers_are_top_mid,
        test_age_component_mature_set_scores_highest,
        test_age_component_brand_new_scores_low,
        test_age_component_missing_date_is_neutral,
        test_age_component_handles_dash_format,
        test_valorization_top_rarity_premium_price_high_score,
        test_valorization_bulk_cheap_new_low_score,
        test_valorization_never_invents_historical_data,
    ]
    with_tmp = [test_xlsx_has_new_columns_at_end, test_xlsx_new_column_values]

    failures = 0
    for fn in no_arg:
        try:
            fn()
            print(f"PASS {fn.__name__}")
        except AssertionError as e:
            failures += 1
            print(f"FAIL {fn.__name__}: {e}")
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td)
        for fn in with_tmp:
            try:
                fn(tp)
                print(f"PASS {fn.__name__}")
            except AssertionError as e:
                failures += 1
                print(f"FAIL {fn.__name__}: {e}")
    print(f"\n{len(no_arg) + len(with_tmp) - failures}/{len(no_arg) + len(with_tmp)} passed")
    sys.exit(1 if failures else 0)
