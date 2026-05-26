"""PSA grading profitability analysis via PriceCharting scrape.

For each card in a scanner XLSX output, fetches Ungraded / PSA 9 / PSA 10
prices from pricecharting.com and reports cards where grading is profitable
(PSA 9 sale price > raw NM price + grading fee).

Usage:
    python psa_grading_analysis.py --input <scan.xlsx> [--input <scan2.xlsx> ...] \
        --output psa_analysis.xlsx [--grading-fee 50] [--rate-limit 1.0]
"""
import argparse
import json
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote_plus

import requests
from openpyxl import Workbook, load_workbook

SCRIPT_DIR = Path(__file__).parent
CACHE_PATH = SCRIPT_DIR / ".pricecharting_cache.json"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
SEARCH_URL = "https://www.pricecharting.com/search-products?q={q}&type=prices"

PRICE_ID_RE = re.compile(
    r'<td\s+id="([a-z_]+_price)"[^>]*>\s*<span class="price js-price">\s*\$?([\d,\.]+|N/A)',
    re.I,
)
GAME_URL_RE = re.compile(r'href="(https://www\.pricecharting\.com/game/pokemon-[a-z0-9\-]+/[a-z0-9\-]+)"')


PRICE_FIELD_MAP = {
    "used_price": "ungraded",
    "complete_price": "psa7",
    "new_price": "psa8",
    "graded_price": "psa9",
    "box_only_price": "psa95",
    "manual_only_price": "psa10",
}

# Grading probability rules-of-thumb for NM raw cards bought from sellers
# WITHOUT photo/scan inspection. Conservative estimates from the PSA grading
# community (vintage is harder due to centering/edges/print, modern more lenient).
# Source: aggregated from PSA pop reports + r/PSACard heuristics + Coleção experience.
ERA_GRADE_PROBS = {
    # era_key: (p_below8, p_psa8, p_psa9, p_psa10)
    "wotc":      (0.25, 0.30, 0.40, 0.05),   # 1999-2003 (Base..Skyridge)
    "ex":        (0.17, 0.25, 0.50, 0.08),   # 2003-2007 (EX Ruby/Sapphire..Power Keepers)
    "dp":        (0.13, 0.20, 0.55, 0.12),   # 2007-2010 (Diamond/Pearl..Arceus)
    "hgss_bw":   (0.09, 0.18, 0.58, 0.15),   # 2010-2013 (HeartGold..Legendary Treasures)
    "xy":        (0.07, 0.15, 0.60, 0.18),   # 2014-2016
    "sm":        (0.06, 0.12, 0.60, 0.22),   # 2017-2019 (Sun & Moon)
    "swsh":      (0.05, 0.10, 0.60, 0.25),   # 2020-2023 (Sword & Shield)
    "sv":        (0.04, 0.08, 0.60, 0.28),   # 2024+ (Scarlet & Violet)
    "unknown":   (0.10, 0.20, 0.55, 0.15),   # fallback
}


def classify_era(set_code: str, set_name: str) -> str:
    """Map a CT set code/name to a PSA-difficulty era bucket."""
    code = (set_code or "").lower()
    name = (set_name or "").lower()
    sv_codes = {"svi","pal","obf","mew","par","paf","tef","twm","sfa","scr","ssp","pre","jtg","dri","blk","wht","sv1v","sv1s","sv2a","svpv","svp","promosv","svproducts","svpk","p-pre","m-pre","p-sv2a","m-151c","151c","151pkmr","csv2","csv9","csv9-p","csv9-m","sv9a","sv9","sv10","sv11b","sv11w","p-sv11b","p-sv11w","m-sv11b","m-sv11w","m-blk","m-wht","p-blk","p-wht","svp-c","svp-id","svl","svk","svls","svln","svb","sveen","svp1","svm","svod","svom","sd-v","csvn","csm15","csmd","csm1d","csvh2","csvh4ac","csvh4ec","csvh4pc","csmjc","csma","csv1","cbb5c","cs15","cs25","cs35","cs45","cs55","cs65","csf","cs41","cs51","cs1","cs1a","cs1b","cs2a","cs2b","cs3b","cs4a","cs4b","cs5a","cs5b","cs6a","cs6b","cs61","csb","csd","csg","csgc","csuc","csve1","30th-ch"}
    swsh_codes = {"ssh","rcl","daa","viv","shf","bst","cre","evs","fst","brs","astr","lorg","sit","crz","cpa","s-p","ms&s","sma","smlf","sll","sld","upcc","wcs23","s1w","sh","sm6b","cp4","s8ap","s8a","s8ag","wcdecks","wcp"}
    sm_codes = {"sum","gri","cinv","upr","fli","ces","drm","lot","teu","unb","unm","hif","cec","bus","det","tk10","tk11","msm","slg","sm3"}
    xy_codes = {"xy-en","tk6","flf","ffi","tk7","phf","prc","tk8","ros","aor","bkt","bkp","gen","tk9","fco","sts","evo","xytkas","xytkos","xyths","xytkn","bisharp","wigglytuff","xytkp","suicune","svm","hxy","mc19"}
    hgss_bw_codes = {"blw","epo","nvi","drx","bcr","pls","plf","plb","dex","ltr","bw7","ppd","pbg","hgs","ul","und","tri","clo","l1","l1ss","pbus","pplf"}
    dp_codes = {"dp","mt","sw","ge","md","la","sft","rr","sv","aoa","pt-a-lp","dp-a-gf","pl"}
    ex_codes = {"hl","lm","uf","ds","cg","pk","rg","em","dr","trr","jun","myf","gls","gcd","ec1","smd","sm5","sm4","sm2","sm10","sm8","sm6","sm1","ex","aq","skg","lc","trr"}
    wotc_codes = {"bs","ju","fo","b2","tr","g1","g2","n1","n2","n3","shbs","wiz","pr1"}

    if code in sv_codes or "scarlet" in name or "violet" in name or "paldea" in name or "151" in name or "stellar" in name or "shrouded" in name or "destined" in name or "journey" in name or "twilight" in name or "temporal" in name or "paradox" in name or "surging" in name or "prismatic" in name or "black bolt" in name or "white flare" in name:
        return "sv"
    if code in swsh_codes or "sword" in name or "shield" in name or "vivid voltage" in name or "rebel clash" in name or "darkness ablaze" in name or "chilling reign" in name or "evolving skies" in name or "fusion strike" in name or "brilliant stars" in name or "astral radiance" in name or "lost origin" in name or "silver tempest" in name or "crown zenith" in name or "battle styles" in name or "shining fates" in name or "champion's path" in name:
        return "swsh"
    if code in sm_codes or "sun & moon" in name or "guardians rising" in name or "burning shadows" in name or "crimson invasion" in name or "ultra prism" in name or "forbidden light" in name or "celestial storm" in name or "lost thunder" in name or "team up" in name or "unbroken bonds" in name or "unified minds" in name or "hidden fates" in name or "cosmic eclipse" in name or "dragon majesty" in name or "shining legends" in name:
        return "sm"
    if code in xy_codes or "xy " in name or name.startswith("xy") or "flashfire" in name or "furious fists" in name or "phantom forces" in name or "primal clash" in name or "roaring skies" in name or "ancient origins" in name or "breakthrough" in name or "breakpoint" in name or "generations" in name or "fates collide" in name or "steam siege" in name or "evolutions" in name or "mcdonald" in name:
        return "xy"
    if code in hgss_bw_codes or "heartgold" in name or "soulsilver" in name or "unleashed" in name or "undaunted" in name or "triumphant" in name or "call of legends" in name or "black & white" in name or "emerging powers" in name or "noble victories" in name or "dragons exalted" in name or "boundaries crossed" in name or "plasma" in name or "dark explorers" in name or "legendary treasures" in name:
        return "hgss_bw"
    if code in dp_codes or "diamond" in name or "pearl" in name or "platinum" in name or "arceus" in name or "rising rivals" in name or "supreme victors" in name or "mysterious treasures" in name or "stormfront" in name or "majestic dawn" in name or "legends awakened" in name or "great encounters" in name or "secret wonders" in name:
        return "dp"
    if code in ex_codes or "expedition" in name or "aquapolis" in name or "skyridge" in name or "legendary collection" in name or name.startswith("ex "):
        # e-Card era (Expedition/Aquapolis/Skyridge) + Legendary Collection are border-line WOTC/EX but grade like EX-era
        return "ex"
    if code in wotc_codes or "base set" in name or "jungle" in name or "fossil" in name or "team rocket" in name or "gym " in name or name.startswith("neo "):
        return "wotc"
    return "unknown"


def expected_value(prices: dict, era_probs: tuple) -> dict:
    """Compute EV per copy graded, plus per-grade probability and per-grade profit.
    EV does NOT subtract grading fee here (caller does that)."""
    p_below8, p8, p9, p10 = era_probs
    psa8 = prices.get("psa8") or 0.0
    psa9 = prices.get("psa9") or 0.0
    psa10 = prices.get("psa10") or 0.0
    ungraded = prices.get("ungraded") or 0.0
    # Below PSA 8 -> sell at ungraded price (worst case)
    weighted = (p_below8 * ungraded) + (p8 * psa8) + (p9 * psa9) + (p10 * psa10)
    return {
        "p_below8": p_below8,
        "p_psa8": p8,
        "p_psa9": p9,
        "p_psa10": p10,
        "weighted_sell_usd": weighted,
    }


def _load_cache() -> dict:
    if CACHE_PATH.exists():
        try:
            return json.loads(CACHE_PATH.read_text())
        except json.JSONDecodeError:
            return {}
    return {}


def _save_cache(cache: dict) -> None:
    CACHE_PATH.write_text(json.dumps(cache, indent=2))


def _clean_set_name(set_name: str) -> str:
    """Strip trailing `(code)` and similar noise from scanner set names."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", set_name).strip()


def _slugify(s: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s


def _set_tokens(set_name: str) -> list[str]:
    """Tokens from set name that must appear in URL set-slug. EX/promo noise filtered."""
    clean = _clean_set_name(set_name)
    raw = _slugify(clean).split("-")
    stop = {"ex", "the", "of", "pokemon", "set", "tcg", "collection"}
    return [t for t in raw if t and t not in stop and len(t) > 2]


def _url_set_slug(url: str) -> str:
    """Extract the set slug between `/game/pokemon-` and `/card`."""
    m = re.search(r"/game/pokemon-([a-z0-9\-]+)/", url)
    return m.group(1) if m else ""


def _url_matches_card_set(url: str, card_name: str, set_name: str) -> bool:
    """Require URL slug to match BOTH the card name AND the set."""
    if not url:
        return False
    card_slug_part = url.rsplit("/", 1)[-1].lower()
    card_slug = _slugify(card_name)
    first = card_slug.split("-")[0]
    if not card_slug:
        return False
    card_ok = (card_slug in card_slug_part) or (
        first and first in card_slug_part
    )
    if not card_ok:
        return False
    set_slug = _url_set_slug(url)
    tokens = _set_tokens(set_name)
    if not tokens:
        return True  # nothing to enforce
    # At least one distinctive set token must appear in URL set slug
    return any(t in set_slug for t in tokens)


def search_card_url(card_name: str, set_name: str, number: str,
                    session: requests.Session) -> str | None:
    """Search PriceCharting; return first URL whose slug matches BOTH card+set."""
    clean_set = _clean_set_name(set_name)
    queries = []
    if number:
        queries.append(f"{card_name} {clean_set} {number}")
    queries.append(f"{card_name} {clean_set}")

    for q in queries:
        url_search = SEARCH_URL.format(q=quote_plus(q))
        r = session.get(url_search, timeout=20)
        if r.status_code != 200:
            continue
        for m in GAME_URL_RE.finditer(r.text):
            candidate = m.group(1)
            if _url_matches_card_set(candidate, card_name, set_name):
                return candidate
    return None


def fetch_prices(url: str, session: requests.Session) -> dict:
    """Fetch the card page and parse all price fields by ID."""
    r = session.get(url, timeout=20)
    if r.status_code != 200:
        return {}
    prices: dict[str, float | None] = {}
    for m in PRICE_ID_RE.finditer(r.text):
        field = PRICE_FIELD_MAP.get(m.group(1))
        if not field:
            continue
        raw = m.group(2).replace(",", "")
        try:
            prices[field] = float(raw)
        except ValueError:
            prices[field] = None
    return prices


def get_card_prices(card_name: str, set_name: str, number: str,
                    session: requests.Session,
                    cache: dict, rate_limit: float = 1.0) -> dict:
    """Lookup PSA prices for one card, with disk cache."""
    key = f"{card_name}|{set_name}|{number}".lower()
    if key in cache:
        return cache[key]

    result = {"card_name": card_name, "set_name": set_name, "number": number,
              "url": None, "prices": {}, "error": None}
    try:
        url = search_card_url(card_name, set_name, number, session)
        if not url:
            result["error"] = "no_search_match"
        else:
            result["url"] = url
            time.sleep(rate_limit)
            result["prices"] = fetch_prices(url, session)
            if not result["prices"]:
                result["error"] = "no_prices_parsed"
    except requests.RequestException as e:
        result["error"] = f"request_error: {type(e).__name__}"

    cache[key] = result
    _save_cache(cache)
    time.sleep(rate_limit)
    return result


def load_scan_rows(paths: list[Path]) -> list[dict]:
    """Read all scanner XLSX outputs and return a deduped list of cards."""
    seen: set[tuple] = set()
    cards: list[dict] = []
    for p in paths:
        if not p.exists():
            print(f"WARN: missing input {p}", file=sys.stderr)
            continue
        wb = load_workbook(p, read_only=True)
        if "Oportunidades" not in wb.sheetnames:
            print(f"WARN: no Oportunidades sheet in {p}", file=sys.stderr)
            continue
        ws = wb["Oportunidades"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(headers)}

        # Best-effort column resolution
        def col(*candidates):
            for c in candidates:
                if c in idx:
                    return idx[c]
            return None

        i_card = col("Card Name", "Carta", "Card")
        i_set = col("Set", "Expansion", "Expansão")
        i_num = col("Nº", "Number", "Num")
        # v2.x scanner columns: live BRL (markup-adjusted price actually paid) + TCG USD
        i_live_brl = col("LIVE R$ (real)", "Live R$", "LIVE R$")
        i_scan_brl = col("Scan R$ (raw)", "Scan R$")
        i_tcg_usd = col("TCG Market (USD)", "TCG USD")
        i_tcg_brl = col("TCG Market (BRL)", "TCG BRL")
        i_link = col("Link CardTrader", "Link CT", "Link")
        i_seller = col("Seller")
        i_foil = col("Foil")
        i_net = col("Net Margin % REAL", "Net Margin %", "Net %")
        i_rarity = col("Rarity", "Raridade")

        for r in rows[1:]:
            if i_card is None or i_set is None:
                break
            card_name = str(r[i_card] or "").strip()
            set_name = str(r[i_set] or "").strip()
            if not card_name or not set_name:
                continue
            num = str(r[i_num] or "").strip() if i_num is not None else ""
            key = (card_name.lower(), set_name.lower(), num)
            if key in seen:
                continue
            seen.add(key)

            # Compute USD/BRL FX rate from TCG columns (live in scan)
            fx = None
            if i_tcg_usd is not None and i_tcg_brl is not None:
                u = r[i_tcg_usd]
                b = r[i_tcg_brl]
                if u and b:
                    fx = float(b) / float(u)
            # Convert live BRL → USD (the price actually paid by the buyer after markup)
            live_brl = r[i_live_brl] if i_live_brl is not None else None
            scan_brl = r[i_scan_brl] if i_scan_brl is not None else None
            cost_brl = live_brl or scan_brl
            cost_usd = (float(cost_brl) / fx) if (cost_brl and fx) else None

            # Set name like "Stellar Crown (scr)" — split out code
            m = re.search(r"^(.*?)\s*\(([^)]+)\)\s*$", set_name)
            set_clean = m.group(1).strip() if m else set_name
            set_code = m.group(2).strip().lower() if m else ""

            cards.append({
                "card_name": card_name,
                "set_name": set_clean,
                "set_code": set_code,
                "number": num,
                "rarity": str(r[i_rarity]) if i_rarity is not None and r[i_rarity] else "",
                "price_usd": cost_usd,
                "live_brl": float(live_brl) if live_brl else None,
                "tcg_usd": float(r[i_tcg_usd]) if i_tcg_usd is not None and r[i_tcg_usd] else None,
                "fx_rate": fx,
                "link_ct": str(r[i_link] or "") if i_link is not None else "",
                "seller": str(r[i_seller] or "") if i_seller is not None else "",
                "foil": r[i_foil] if i_foil is not None else None,
                "net_margin_pct": r[i_net] if i_net is not None else None,
                "source_file": p.name,
            })
        wb.close()
    return cards


def build_analysis(cards: list[dict], grading_fee: float, rate_limit: float) -> list[dict]:
    """For each card, fetch PSA prices and compute grading profit."""
    cache = _load_cache()
    session = requests.Session()
    session.headers.update({"User-Agent": UA})

    rows: list[dict] = []
    n = len(cards)
    for i, c in enumerate(cards, 1):
        print(f"[{i}/{n}] {c['card_name']} ({c['set_name']})", flush=True)
        pc = get_card_prices(c["card_name"], c["set_name"], c.get("number", ""),
                             session, cache, rate_limit)
        prices = pc.get("prices", {}) or {}
        raw_nm = c.get("price_usd")
        psa8 = prices.get("psa8")
        psa9 = prices.get("psa9")
        psa10 = prices.get("psa10")
        ungraded_pc = prices.get("ungraded")

        # Era + grading probabilities
        era = classify_era(c.get("set_code", ""), c["set_name"])
        era_probs = ERA_GRADE_PROBS.get(era, ERA_GRADE_PROBS["unknown"])
        ev = expected_value(prices, era_probs)

        def per_grade(sell_usd):
            if sell_usd is None or raw_nm is None:
                return None, None
            profit = sell_usd - raw_nm - grading_fee
            margin = profit / (raw_nm + grading_fee) * 100 if (raw_nm + grading_fee) else None
            return round(profit, 2), margin

        psa8_profit, psa8_margin_pct = per_grade(psa8)
        psa9_profit, psa9_margin_pct = per_grade(psa9)
        psa10_profit, psa10_margin_pct = per_grade(psa10)

        # Expected value: weighted sell across grade distribution, minus raw cost + fee
        ev_profit = None
        ev_margin_pct = None
        if raw_nm and ev["weighted_sell_usd"] > 0:
            ev_profit = ev["weighted_sell_usd"] - raw_nm - grading_fee
            ev_margin_pct = ev_profit / (raw_nm + grading_fee) * 100 if (raw_nm + grading_fee) else None

        grade_worth_it_psa9 = (raw_nm and psa9 and psa9 > raw_nm + grading_fee)
        ev_worth_grading = ev_profit is not None and ev_profit > 0

        rows.append({
            "card_id": f"{c['card_name']} #{c.get('number') or '-'}",
            "card_name": c["card_name"],
            "number": c.get("number", ""),
            "set_name": c["set_name"],
            "set_code": c.get("set_code", ""),
            "rarity": c.get("rarity", ""),
            "era": era,
            "seller": c["seller"],
            "foil": c["foil"],
            "raw_nm_usd": raw_nm,
            "live_brl": c.get("live_brl"),
            "fx_rate": c.get("fx_rate"),
            "tcg_market_usd": c.get("tcg_usd"),
            "pc_ungraded_usd": ungraded_pc,
            "psa8_usd": psa8, "psa8_profit_usd": psa8_profit, "psa8_margin_pct": psa8_margin_pct,
            "psa9_usd": psa9, "psa9_profit_usd": psa9_profit, "psa9_margin_pct": psa9_margin_pct,
            "psa10_usd": psa10, "psa10_profit_usd": psa10_profit, "psa10_margin_pct": psa10_margin_pct,
            "p_psa8": era_probs[1],
            "p_psa9": era_probs[2],
            "p_psa10": era_probs[3],
            "p_below8": era_probs[0],
            "ev_sell_usd": round(ev["weighted_sell_usd"], 2) if ev["weighted_sell_usd"] else None,
            "ev_profit_usd": round(ev_profit, 2) if ev_profit is not None else None,
            "ev_margin_pct": ev_margin_pct,
            "ev_worth_grading": ev_worth_grading,
            "psa9_worth_grading": grade_worth_it_psa9,
            "link_ct": c["link_ct"],
            "link_pc": pc.get("url") or "",
            "error": pc.get("error"),
            "source_file": c["source_file"],
        })
    return rows


def write_xlsx(rows: list[dict], path: Path, grading_fee: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PSA Analysis"
    headers = [
        "Card+Nº", "Set", "Era", "Rarity", "Seller", "Foil",
        "Raw NM (USD)", "TCG Mkt (USD)", "PC Ungraded (USD)",
        "PSA 8 (USD)", "PSA 8 Lucro", "PSA 8 Margem %",
        "PSA 9 (USD)", "PSA 9 Lucro", "PSA 9 Margem %",
        "PSA 10 (USD)", "PSA 10 Lucro", "PSA 10 Margem %",
        "P(PSA 8)", "P(PSA 9)", "P(PSA 10)",
        "EV Sell (USD)", "EV Lucro (USD)", "EV Margem %", "Vale graded? (EV)",
        "Link CT", "Link PriceCharting", "Erro", "Origem",
    ]
    ws.append(headers)
    def pct(v):
        return f"{v:.1f}%" if v is not None else None
    for r in rows:
        ws.append([
            r["card_id"], r["set_name"], r.get("era"), r.get("rarity", ""), r["seller"], r["foil"],
            round(r["raw_nm_usd"], 2) if r["raw_nm_usd"] else None,
            r.get("tcg_market_usd"),
            r["pc_ungraded_usd"],
            r["psa8_usd"], r["psa8_profit_usd"], pct(r["psa8_margin_pct"]),
            r["psa9_usd"], r["psa9_profit_usd"], pct(r["psa9_margin_pct"]),
            r["psa10_usd"], r["psa10_profit_usd"], pct(r["psa10_margin_pct"]),
            pct(r["p_psa8"]*100), pct(r["p_psa9"]*100), pct(r["p_psa10"]*100),
            r["ev_sell_usd"], r["ev_profit_usd"], pct(r["ev_margin_pct"]),
            "SIM" if r["ev_worth_grading"] else ("NÃO" if r["ev_worth_grading"] is False else None),
            r["link_ct"], r["link_pc"], r["error"], r["source_file"],
        ])

    info_ws = wb.create_sheet("Info")
    info_ws.append(["Parâmetros"])
    info_ws.append(["Grading fee (USD)", grading_fee])
    info_ws.append(["Total cards", len(rows)])
    info_ws.append(["Decisão SIM = PSA 9 > Raw NM + Grading Fee"])
    info_ws.append(["Fonte preços graded", "pricecharting.com"])

    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", nargs="+", required=True, help="Scanner XLSX outputs")
    ap.add_argument("--output", required=True, help="Resultado XLSX")
    ap.add_argument("--grading-fee", type=float, default=50.0,
                    help="Custo de grading em USD (default 50)")
    ap.add_argument("--rate-limit", type=float, default=1.5,
                    help="Segundos entre requests PriceCharting (default 1.5)")
    args = ap.parse_args()

    paths = [Path(p) for p in args.input]
    out = Path(args.output)

    cards = load_scan_rows(paths)
    print(f"Loaded {len(cards)} unique cards from {len(paths)} input file(s)")

    if not cards:
        print("Nothing to analyze.")
        return

    rows = build_analysis(cards, args.grading_fee, args.rate_limit)
    write_xlsx(rows, out, args.grading_fee)
    print(f"Wrote {out}")

    # Console summary — rank by Expected Value (accounts for grading difficulty)
    candidates = [r for r in rows if r["ev_profit_usd"] is not None]
    candidates.sort(key=lambda x: x["ev_profit_usd"], reverse=True)
    worth = [r for r in candidates if r["ev_worth_grading"]]
    print(f"\n=== {len(worth)}/{len(rows)} cards com EV de grading POSITIVO ===\n")
    print(f"{'Card+Nº':40} {'Era':8} {'Raw':>8} {'PSA8':>8} {'PSA9':>8} {'PSA10':>8} {'EV':>8} {'EV%':>7}")
    for r in candidates[:30]:
        ev_marker = "[OK]" if r["ev_worth_grading"] else "[ ]"
        psa8 = f"${r['psa8_usd']:.0f}" if r['psa8_usd'] else "-"
        psa9 = f"${r['psa9_usd']:.0f}" if r['psa9_usd'] else "-"
        psa10 = f"${r['psa10_usd']:.0f}" if r['psa10_usd'] else "-"
        raw = f"${r['raw_nm_usd']:.0f}" if r['raw_nm_usd'] else "-"
        ev_prof = f"${r['ev_profit_usd']:.0f}" if r['ev_profit_usd'] is not None else "-"
        ev_pct = f"{r['ev_margin_pct']:.0f}%" if r['ev_margin_pct'] is not None else "-"
        print(f" {ev_marker} {r['card_id'][:38]:38} {r['era']:8} {raw:>8} {psa8:>8} {psa9:>8} {psa10:>8} {ev_prof:>8} {ev_pct:>7}")


if __name__ == "__main__":
    main()
