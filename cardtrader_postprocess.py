"""CardTrader Postprocess v2.0 — núcleo simplificado (operador 2026-05-16).

Mudancas vs v1.5:
- REMOVE bucket CORE/HYPE/DEAD classification
- REMOVE fundamental_analysis original (iconicidade/chase/meta heuristicas subjetivas)
- REMOVE long_term tier
- REMOVE coluna "Acao" (substituida por Decisao mecanica)
- ADICIONA Chase Tier (TOP/MID/MODEST/BULK) baseado em rarity oficial PokemonTCG
- ADICIONA Fundamental Score (0-100) derivado de metricas OBJETIVAS (chase + margin + lucro + validation)
- ADICIONA Decisao (COMPRA/REVISAR/NAO) via regra MECANICA (nao opiniao Claude)
- ADICIONA Porque (1 linha factual)
- REDUZ de 10 sheets pra 3 (Deals, All Listings, Summary). Expandido pra 6
  desde PR-K (+ Top 50 Margin, Validate Manually, TCG Suspect — modelo MYP)
- MANTEM: TG## auto-filter, hyperlinks, alias fixes
- v2.12 (2026-06-06): margem BRUTA — custo = preco_pagina, SEM taxa (default
  --hub-fee 0.0). SUPERSEDE o × 1.06. Operador soma fees por fora.

v2.1 (bug-hunt 2026-05-17):
- #2 Defensive recompute de net_margin/lucro_liq via fórmula 1.06 quando o
  input carrega `live_brl` + `reference_price_brl` válidos. Compara com o
  net_margin do input — drift > 0.5pp → WARN + usa recomputado. Coluna
  `margin_source` ("input" | "recomputed") preserva auditoria.
- #3 Força UTF-8 no stdout/encoding pra eliminar mojibake em headers
  (`Decis�o` → `Decisão`) quando invocado fora do wrapper PS sem
  PYTHONIOENCODING=utf-8. openpyxl pega encoding do interpreter locale.
- #4 Summary separa `Total listings escaneados` (input completo) de
  `Total deals exportados` (COMPRA + REVISAR). Antes o label estava
  errado — número era de deals, não listings totais.

v2.3 Layer 5 (bug-hunt 2026-05-18):
- ALPHA_SUFFIX_RE detecta `153a`, `022a`, `156b` no collector number.
  Tipicamente promo/League variant (1st/2nd/3rd/4th Place ou Prerelease)
  cega pra pokemontcg.io. classify_decision retorna REVISAR antes de
  qualquer outro check.
- Pareado com Scanner v2.8 Layer 4 (foil-aware variant disambiguation
  no provider) — cobre as 2 classes de falsos positivos de variante
  detectados na validação manual do weekly v2.6 (Pichu/Tyranitar e
  Lusamine).

Decisao mecanica (thresholds configuraveis via CLI):
  COMPRA:  net_margin >= 25% AND lucro_liq >= R$50 AND chase_tier in {TOP, MID}
           AND validation_status in {VALIDATED_REAL, VALIDATED_MARKUP}
           AND NOT trainer_gallery_potential_fp
  NAO:     chase_tier == BULK OR net_margin < 20% OR validation_status == STALE
           OR trainer_gallery_potential_fp
  REVISAR: else (zona cinza — margem 20-25% OU chase MODEST com margem alta)

Memorias relevantes respeitadas:
  - margem BRUTA (operador 2026-06-06): custo = preco do site, sem taxa;
    fees calculados FORA do scanner. SUPERSEDE ct_margin_formula (× 1.06).
  - feedback_no_purchase_decisions: Decisao e REGRA mecanica, nao opiniao Claude
  - cardtrader_trainer_gallery_bug: TG## auto-filter mantido
"""
from __future__ import annotations
import argparse, os, re, sys
from dataclasses import dataclass
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Caminho 1 DoubleHolo (2ª opinião): join por productId TCGplayer + nota DH.
# Opcional — só ativa com a flag --doubleholo. Sem ela, comportamento idêntico.
import doubleholo_join

# ─── v2.1 fix #3 (bug-hunt 2026-05-17): força UTF-8 no I/O ───────────────────
# openpyxl.Workbook.save() pega encoding do interpreter locale via algum
# caminho interno. Em Windows com locale pt-BR, sys.stdout.encoding default
# é cp1252 → headers UTF-8 são gravados como bytes UTF-8 mas LIDOS como
# latin-1, produzindo mojibake (`Decis�o`, `Pre�o CT`). PYTHONIOENCODING=utf-8
# resolve, mas só quando o invocador seta a var (wrappers PS setam; scripts
# ad-hoc esquecem). Garantia defensiva aqui:
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    # Python <3.7 ou stdout não-reconfigurável (raro). Fallback silencioso.
    pass

# ─── Margem BRUTA (decisão do operador 2026-06-06 — v2.12) ───────────────────
# SUPERSEDE a fórmula `× 1.06`. O postprocess NÃO embute mais nenhuma taxa por
# default: custo = preço do site, margem = (tcg − preço)/tcg. O operador soma
# Hub fee/frete/cartão/IOF por fora, manualmente. Para reembutir os 6% históricos
# passe `--hub-fee 0.06` (mantém paridade com o scanner, que também aceita).
HUB_FEE_RATE = 0.0

# ─── Net % threshold pra Validate Manually (PR-L 2026-05-30) ─────────────────
# Net % chega em fração (0.30 = 30%). EXTREME_NET_PCT = 2.0 significa 200% —
# margens dessa magnitude são quase sempre falso positivo (provider error,
# inflação reverseHolofoil vintage, etc.) e vão pra revisão manual.
EXTREME_NET_PCT = 2.0

# ─── Gallery subset filter (Trainer Gallery TG## + Galarian Gallery GG##) ─────
# Cartas dos subsets de galeria geram FP massivo: o pricing pokemontcg.io casa
# com uma variante secret-rare alt-art de MESMO NOME (preço 5-10x maior) →
# margem falsa (35/38 oportunidades de um scan SWSH 2026-04-29 eram FP). Produção
# guardava só TG## (^TG\d+), mas o subset irmão Galarian Gallery é numerado GG##
# (confirmado: pokemontcg.io numera "GG01"… com rarity "Trainer Gallery Rare
# Holo", a MESMA classe) e escapava do guard → virava COMPRA com margem falsa.
# O próprio scanner já trata swsh12pt5gg (Crown Zenith Galarian Gallery) como
# gallery split problemático (não-aliasado de propósito) — esta regex só COMPLETA
# essa intenção no postprocess. Achado via ASI-Evolve (experimento
# cardtrader_classify: baseline COMPRA-F1 0,73 → 1,0 ao generalizar o guard).
# CONSERVADOR: roteia GG## pra NAO/manual igual TG##, nunca auto-compra. O nome
# TRAINER_GALLERY_RE é mantido por compat (agora cobre TG+GG).
TRAINER_GALLERY_RE = re.compile(r'^(?:TG|GG)\d+', re.IGNORECASE)

# ─── CT set base totals (PR-K — alinhamento com modelo MYP, 2026-05-29) ──────
# Formato Carta MYP-style: "Plusle (193/197)". Total = printedTotal do
# pokemontcg.io (números BASE do set, NÃO incluindo SIR/HR/secret rares).
# Fonte: api.pokemontcg.io/v2/sets queried 2026-05-29.
#   sv4 par   = 182 base / 266 total (Paradox Rift)
#   sv4pt5 paf= 91 base / 245 total  (Paldean Fates)
#   sv5 tef   = 162 base / 218 total (Temporal Forces)
#   sv6 twm   = 167 base / 226 total (Twilight Masquerade)
#   sv6pt5 sfa= 64 base / 99 total   (Shrouded Fable)
#   sv7 scr   = 142 base / 175 total (Stellar Crown)
#   sv8 ssp   = 191 base / 252 total (Surging Sparks)
#   sv8pt5 pre= 131 base / 180 total (Prismatic Evolutions)
#   sv9 jtg   = 159 base / 190 total (Journey Together)
#   sv10 dri  = 182 base / 244 total (Destined Rivals)
#   zsv10pt5 blk = 86 base / 172 total (Black Bolt)
#   me2pt5 asc= 217 base / 295 total (Ascended Heroes)
# Para sets fora desta tabela → fallback "Nome (NNN)" sem total.
CT_SET_TOTAL: dict[str, int] = {
    "par": 182,
    "paf": 91,
    "tef": 162,
    "twm": 167,
    "sfa": 64,
    "scr": 142,
    "ssp": 191,
    "pre": 131,
    "jtg": 159,
    "dri": 182,
    "asc": 217,
    "me2pt5": 217,  # PR-L CT-LOW-a: alias do code pokemontcg.io (= asc)
    "blk": 86,
}

# ─── Vintage suspect filter (PR-K — Validate Manually heuristic) ─────────────
# Sets confirmados como inflados na memória ct_scanner_session_handoff_2026_05_19
# (LC + BA-20 inflados confirmados). Pokemontcg.io aplica reverseHolofoil
# fallback que infla 5-30× quando o card-base tem preço baixo.
VINTAGE_SUSPECT_SETS = {
    "lc",       # Legendary Collection
    "ba-20", "ba20",  # Battle Academy 2020
    "ba-22", "ba22",  # Battle Academy 2022
}

# ─── Alpha suffix filter (v2.3 Layer 5 — bug-hunt 2026-05-18) ────────────────
# Cards com collector number `\d+[a-zA-Z]+` (ex `153a`, `022a`, `156b`) são
# tipicamente variantes promo/league cegas pra pokemontcg.io:
#   - "1st Place Pokemon League" / "2nd Place" / "3rd Place" / "4th Place"
#   - Prerelease (alguns sets)
#   - Staff variants (Champion's Festival, World Championships)
# Pokemontcg.io agrega TODAS sob o número-base (ex `153`) com o preço da
# variante mais cara → margem inflada 5-30× quando seller CT está vendendo
# a promo barata. Caso histórico operador: Lusamine sm5/153a — scanner
# pegou main set Lusamine ($160) mas seller CT vende 1st Place ($13.77).
#
# Estratégia: REVISAR forçado (operador valida via Link TCG → variant
# selector visual). NÃO marcar NÃO automático pq alguns alpha suffixes
# legítimos existem em sets vintage (Topps Movie, Black Star Promos
# numerados como `TVR_1a`, etc — fora do escopo CT).
#
# NÃO confundir com TG##: TG## tem letras ANTES do número, então a regex
# `^\d+[a-zA-Z]+` não casa. TG## fica com sua flag separada.
ALPHA_SUFFIX_RE = re.compile(r'^\d+[a-zA-Z]+', re.IGNORECASE)

# ─── Unsupported sets filter (v2.2 Layer 3 bug-hunt 2026-05-18) ──────────────
# Sets onde pokemontcg.io tem cobertura ruim OU os codes do CT divergem de
# forma que mesmo com Layer 1 (strict set match) acabamos pegando o set
# errado. Quando um listing CT vem de um destes sets, qualquer Decisão
# automática vira REVISAR_MANUAL — operador valida preço TCG manual antes
# de comprar.
#
# Lista canônica (operador 2026-05-18 pós-debug weekly v2.6):
#   - Promo/legacy sets sem alias em pokemontcg.io: phs, pplf, pupr, xybsp
#   - World Championship decks (não-tradeable na ptcg.io): wcd2004/06/07
#   - Theme deck exclusives sem reverse map: deckexclusives, xytkn
#   - Pokemon TCG Classic reprint set: clb (matched Team Rocket originals)
#   - McDonald's promo sets: m24
UNSUPPORTED_SETS = {
    "clb", "wcd2004", "wcd2006", "wcd2007", "deckexclusives",
    "xytkn", "m24", "phs", "pplf", "xybsp",
}

def _extract_set_code_from_label(set_label: str | None) -> str:
    """Set label vem como 'Jungle (ju)' ou direto 'ju'. Extrai código entre
    parênteses se presente, senão devolve label normalizado."""
    if not set_label:
        return ""
    s = str(set_label).strip()
    m = re.search(r"\(([^)]+)\)\s*$", s)
    if m:
        return m.group(1).strip().lower()
    return s.lower()

# ─── Chase Tier classification ────────────────────────────────────────────────
# Hierarquia objetiva baseada em PokemonTCG official rarities. Substring-aware
# (case-insensitive) pra tolerar variacoes "Special Illustration Rare" vs "SIR".
CHASE_TIER_PATTERNS = {
    "TOP": [
        "special illustration rare", "sir", "illustration rare",
        "special art rare", "sar", "hyper rare", "ultra hyper rare",
        "secret rare", "rara secreta",
    ],
    "MID": [
        "full art", "alt art", "alternate art", "alternative art",
        "rainbow rare", "gold rare", "trainer gallery", "double rare",
        "rara hiper", "ultra rare",
    ],
    "MODEST": [
        "holo rare", "reverse holo", "reverse foil", "promo",
        "rare holo",
    ],
    "BULK": [
        "common", "comum", "uncommon", "incomum",
    ],
}

def classify_chase_tier(rarity: str | None, card_number: str | None = None,
                         markup_tier: str | None = None) -> str:
    """Classifica chase tier baseado em rarity. Fallback para proxies se rarity
    indisponivel (XLSX raw pre v2.5 do scanner nao persiste rarity)."""
    if rarity:
        r = str(rarity).lower().strip()
        for tier, patterns in CHASE_TIER_PATTERNS.items():
            if any(p in r for p in patterns):
                return tier
        # Rarity present but unmapped → default MODEST (conservative)
        return "MODEST"

    # ── Fallback: proxies quando rarity nao foi capturada pelo scanner ──
    num = str(card_number or "").strip()
    # TG## = Trainer Gallery → MID (mas tb gera flag separado abaixo)
    if TRAINER_GALLERY_RE.match(num):
        return "MID"
    # Supranumerary (numero > total set) é sinal de SIR/SAR — TOP
    if "/" in num:
        try:
            n, total = num.split("/")[:2]
            if int("".join(c for c in n if c.isdigit())) > int("".join(c for c in total if c.isdigit())):
                return "TOP"
        except (ValueError, IndexError):
            pass
    # Markup tier non-VAT (+20%) frequentemente sinaliza cards top
    # (sellers profissionais cobrando margem em high-value items)
    mt = str(markup_tier or "").lower()
    if "non-vat" in mt or "+20%" in mt:
        return "MID"
    # Default sem evidencia: MODEST (conservador — nao penaliza muito,
    # mas tb nao aprova compra automatic)
    return "MODEST"

# ─── Decisao mecanica + Porque ────────────────────────────────────────────────
@dataclass
class DecisionConfig:
    min_net_margin: float = 0.25        # 25% net margin pra COMPRA
    min_lucro_liq: float = 50.0          # R$50 lucro minimo pra COMPRA
    revisar_min_net: float = 0.20        # 20-25% → REVISAR
    revisar_chase_modest_min_net: float = 0.30  # MODEST com >=30% → REVISAR
    hub_fee_rate: float = 0.0            # v2.12: margem BRUTA (sem taxa). Override --hub-fee 0.06

def classify_decision(row, cfg: DecisionConfig):
    """Aplica regra mecanica. Retorna (decisao, porque)."""
    chase = row.get("chase_tier", "MODEST")
    nm = row.get("net_margin", 0)
    profit = row.get("lucro_liq", 0)
    val = str(row.get("validation_status", "")).upper()
    is_tg = bool(row.get("trainer_gallery_potential_fp", False))

    # v2.3 Layer 5 (bug-hunt 2026-05-18): alpha-suffix detect (153a, 022a).
    # Cards com sufixo alfanumérico no collector number são tipicamente
    # promos/league (1st/2nd/3rd/4th Place, Prerelease) cegas pra pokemontcg.io.
    # Operador validou em 2026-05-18: Lusamine sm5/153a — preço scanner $160
    # vs preço real $13.77 (1st Place variant). REVISAR forçado.
    card_num = str(row.get("card_number") or "").strip()
    if ALPHA_SUFFIX_RE.match(card_num) and not TRAINER_GALLERY_RE.match(card_num):
        return ("REVISAR",
                f"Alpha suffix '{card_num}' — provável promo/League variant "
                f"(1st/2nd/3rd/4th Place ou Prerelease); pokemontcg.io cega "
                f"pra essas, valide via Link TCG antes")

    # v2.2 Layer 3 (bug-hunt 2026-05-18): unsupported sets → REVISAR forçado.
    # Mesmo que Layer 1 (strict set match) tenha aceito o pricing, alguns
    # sets têm cobertura ruim no pokemontcg.io OU divergem de forma
    # silenciosa. Operador valida manual antes de comprar.
    set_code = _extract_set_code_from_label(row.get("set_code"))
    if set_code in UNSUPPORTED_SETS:
        return "REVISAR", f"Set {set_code} sem cobertura confiável em pokemontcg.io — valide TCG manual"

    if pd.isna(nm) or pd.isna(profit):
        return "NAO", "Dados insuficientes (margem/lucro ausentes)"

    # NAO (rejeitos definitivos)
    if is_tg:
        return "NAO", "TG##/GG## gallery potencial FP (pokemontcg.io infla 5-10x)"
    if val in ("STALE",):
        return "NAO", "Validation STALE — preço inseguro"
    if chase == "BULK":
        return "NAO", f"Chase BULK + net {nm:.0%} (bulk sem liquidez mesmo barato)"
    if nm < cfg.revisar_min_net:
        return "NAO", f"Net margin {nm:.0%} < {cfg.revisar_min_net:.0%} (abaixo do piso)"

    # REVISAR (zona cinza)
    if nm < cfg.min_net_margin:
        return "REVISAR", f"Net {nm:.0%} entre {cfg.revisar_min_net:.0%}-{cfg.min_net_margin:.0%} (borderline)"
    if profit < cfg.min_lucro_liq:
        return "REVISAR", f"Net {nm:.0%} OK mas lucro R${profit:.0f} < R${cfg.min_lucro_liq:.0f}"
    if chase == "MODEST":
        if nm >= cfg.revisar_chase_modest_min_net:
            return "REVISAR", f"MODEST + net alta {nm:.0%} (vale checar liquidez do set)"
        return "NAO", f"MODEST + net {nm:.0%} (precisa >={cfg.revisar_chase_modest_min_net:.0%} pra MODEST)"
    if val == "MARKUP_TIER_ANOMALOUS" or val == "ANOMALOUS_MARKUP":
        return "REVISAR", f"Markup anomalo (>45%); chase {chase} + net {nm:.0%}"
    if val not in ("VALIDATED_REAL", "VALIDATED_MARKUP"):
        return "REVISAR", f"Validation {val or 'ausente'}; chase {chase} + net {nm:.0%}"

    # v2.24 (2026-06-26): sinal "Variante Baixa Confiança" do scanner (v2.18 holo/
    # unlimited não-holo + v2.24 reverse-outlier, ex. Lileep ex12-56 reverse $37.50
    # vs normal $0.55 = 68×). Uma linha que sobreviveu a todos os filtros acima
    # SERIA COMPRA limpa — mas a âncora de variante é suspeita de estar inflada.
    # Downgrade pra REVISAR ("validar manual"): nunca apresentada como COMPRA limpa.
    # Aplicado por ÚLTIMO, então NÃO promove um NAO (margem/TG/STALE seguem NAO);
    # só rebaixa o que viraria COMPRA. Margem/preço/bucket INALTERADOS — sinal-only.
    if _is_truthy_flag(row.get("variant_low_confidence")):
        return ("REVISAR",
                "Variante Baixa Confiança — preço de referência casou numa variante "
                "reverse/holo possivelmente inflada (não-holo); valide a variante via "
                "Link TCG antes de comprar")

    # COMPRA
    return "COMPRA", f"Chase {chase} + net {nm:.0%} + lucro R${profit:.0f}"

# ─── Fundamental Score objetivo (0-100) ──────────────────────────────────────
def fundamental_score(row) -> int:
    """Score derivado de 4 metricas objetivas (operador-pedido em 2026-05-16).
    Nao usa heuristicas subjetivas (iconicidade, meta competitivo, etc)."""
    score = 0
    # Chase tier (peso 40)
    chase_pts = {"TOP": 40, "MID": 25, "MODEST": 10, "BULK": 0}
    score += chase_pts.get(row.get("chase_tier", "MODEST"), 0)
    # Net margin (peso 20)
    nm = row.get("net_margin", 0)
    if nm >= 0.40: score += 20
    elif nm >= 0.30: score += 15
    elif nm >= 0.25: score += 10
    elif nm >= 0.20: score += 5
    # Lucro absoluto (peso 25)
    p = row.get("lucro_liq", 0)
    if p >= 200: score += 25
    elif p >= 100: score += 18
    elif p >= 50: score += 12
    elif p >= 25: score += 6
    # Validation status (peso 15)
    val = str(row.get("validation_status", "")).upper()
    if val == "VALIDATED_REAL": score += 15
    elif val == "VALIDATED_MARKUP": score += 10
    elif val == "ANOMALOUS_MARKUP": score += 3
    return min(score, 100)

# ─── Column aliases (preserva v1.5 fixes) ────────────────────────────────────
COLUMN_ALIASES = {
    "card_name": ["card_name","name","card","Card","Name","Card Name","Carta"],
    "card_number": ["card_number","number","Number","Card Number","card_no","No","Numero","Nº","No."],
    "set_code": ["set_code","set","Set","expansion","Expansion","set_id"],
    "rarity": ["rarity","Rarity","Raridade"],
    "variant": ["variant","Variant","version","Version","foil","Foil","Variante"],
    "condition": ["condition","Condition","cond","Cond","Condicao","Condição"],
    "language": ["language","Language","lang","Lang","Idioma"],
    "seller": ["seller","Seller","seller_name"],
    "live_brl": ["live_brl","LIVE R$","LIVE R$ (real)","live_price_brl"],
    "markup_pct": ["markup_pct","Markup %","markup_percent","markup"],
    "markup_tier": ["markup_tier","Markup Tier","tier"],
    "validation_status": ["validation_status","Validation Status","valid_status","status"],
    "reference_price_brl": ["reference_price_brl","TCG Market (BRL)","TCG R$","TCG Market (R$)"],
    # Preço de referência TCGPlayer na MOEDA ORIGINAL (USD). Usado só na tabela
    # de entrega no chat (colunas "CT US$"/"TCG US$"). O raw do scanner grava
    # "TCG Market (USD)"; pode faltar em provider != pokemontcg → fallback FX.
    "reference_price_usd": ["reference_price_usd","TCG Market (USD)","TCG USD","TCG US$","tcg_market_usd"],
    "net_margin": ["net_margin","Net Margin % REAL","Net Margin %","Net REAL"],
    "lucro_liq": ["lucro_liq","Lucro R$ REAL","Lucro REAL","Lucro Liq (R$)","Net Profit (R$)"],
    "link_ct": ["link_ct","Link CT","Link CardTrader","CardTrader URL","CardTrader Link","Link"],
    "quantity": ["quantity","Quantity","qty","estoque","Qtd"],
    # v2.7.1 postprocess (2026-05-18): URL TCGPlayer da carta exata
    # matched no pokemontcg.io. Operador valida variante (ex Lusamine 1st Place
    # vs normal) antes de comprar. None aceitável (não-pokemontcg providers).
    "link_tcg": ["link_tcg","Link TCG","tcg_url","TCG URL","TCGPlayer URL","TCGPlayer Link"],
    # v2.24 (2026-06-26): sinal "Variante Baixa Confiança" do scanner ("Sim"/"").
    # Lê o flag de baixa confiança de variante (v2.18 holo/unlimited não-holo +
    # v2.24 reverse-outlier) pra forçar REVISAR — fim do "clean COMPRA" em linhas
    # com âncora de variante suspeita.
    "variant_low_confidence": ["variant_low_confidence","Variante Baixa Confiança","Variante Baixa Confianca","low_confidence_variant"],
}


def _is_truthy_flag(v) -> bool:
    """v2.24: interpreta o valor da coluna "Variante Baixa Confiança" do XLSX.
    O scanner grava "Sim"/"" (string); aceita também bool/1/true por robustez.
    NaN / vazio → False."""
    if v is None:
        return False
    try:
        if isinstance(v, float) and pd.isna(v):
            return False
    except (TypeError, ValueError):
        pass
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("sim", "true", "1", "yes", "y")

def detect_column(df, logical):
    aliases = COLUMN_ALIASES.get(logical, [logical])
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for a in aliases:
        if a in df.columns: return a
        if a.lower() in cols_lower: return cols_lower[a.lower()]
    return None

def normalize_columns(df):
    rename = {}
    for logical in COLUMN_ALIASES:
        actual = detect_column(df, logical)
        if actual: rename[actual] = logical
    return df.rename(columns=rename)

# ─── Hyperlink helper (preservado de v1.5) ────────────────────────────────────
HYPERLINK_FONT = Font(color="0563C1", underline="single")

def apply_card_hyperlinks(ws, df):
    """Aplica hyperlink ativo em Carta → Link CT, e na própria célula Link TCG.

    v2.7.1 (postprocess 2026-05-18): Link TCG (URL TCGPlayer) ganha
    hyperlink no próprio texto (operador clica diretamente). Distinção
    de design vs Link CT: Carta vira link clicável apontando pro CT
    (workflow de compra primário); Link TCG é texto URL clicável
    (workflow de validação de variante secundário).
    """
    cols = list(df.columns)
    # Carta → Link CT
    if "Carta" in cols and "Link CT" in cols:
        carta_idx = cols.index("Carta") + 1
        link_idx = cols.index("Link CT") + 1
        for ri in range(2, len(df) + 2):
            url = ws.cell(row=ri, column=link_idx).value
            if isinstance(url, str) and url.startswith("http"):
                c = ws.cell(row=ri, column=carta_idx)
                c.hyperlink = url
                c.font = HYPERLINK_FONT
    # Link TCG → célula própria (URL clicável)
    if "Link TCG" in cols:
        tcg_idx = cols.index("Link TCG") + 1
        for ri in range(2, len(df) + 2):
            cell = ws.cell(row=ri, column=tcg_idx)
            v = cell.value
            if isinstance(v, str) and v.startswith("http"):
                cell.hyperlink = v
                cell.font = HYPERLINK_FONT
    # Preço CT → Link CT (clicar no preço abre a página do card no CT pra conferir)
    if "Preço CT (R$)" in cols and "Link CT" in cols:
        price_idx = cols.index("Preço CT (R$)") + 1
        link_idx = cols.index("Link CT") + 1
        for ri in range(2, len(df) + 2):
            url = ws.cell(row=ri, column=link_idx).value
            if isinstance(url, str) and url.startswith("http"):
                c = ws.cell(row=ri, column=price_idx)
                c.hyperlink = url
                c.font = HYPERLINK_FONT
    # TCG (R$) → Link TCG (clicar no preço abre a página TCGplayer pra conferir)
    if "TCG (R$)" in cols and "Link TCG" in cols:
        tprice_idx = cols.index("TCG (R$)") + 1
        link_idx = cols.index("Link TCG") + 1
        for ri in range(2, len(df) + 2):
            url = ws.cell(row=ri, column=link_idx).value
            if isinstance(url, str) and url.startswith("http"):
                c = ws.cell(row=ri, column=tprice_idx)
                c.hyperlink = url
                c.font = HYPERLINK_FONT

# ─── Main pipeline ────────────────────────────────────────────────────────────
DECISAO_FILL = {
    "COMPRA":  PatternFill("solid", fgColor="C6EFCE"),
    "REVISAR": PatternFill("solid", fgColor="FFEB9C"),
    "NAO":     PatternFill("solid", fgColor="F4CCCC"),
}
CHASE_FILL = {
    "TOP":    PatternFill("solid", fgColor="A9D08E"),
    "MID":    PatternFill("solid", fgColor="FFE699"),
    "MODEST": PatternFill("solid", fgColor="F8CBAD"),
    "BULK":   PatternFill("solid", fgColor="D9D9D9"),
}

def _recompute_margin_with_fee(df: pd.DataFrame, drift_threshold_pp: float = 0.005,
                               hub_fee_rate: float = HUB_FEE_RATE) -> pd.DataFrame:
    """v2.1 #2 / v2.12: defensivamente recomputa net_margin + lucro_liq.

    v2.12 (2026-06-06): a fórmula passou a ser BRUTA por default
    (`hub_fee_rate` default 0.0 → custo = live_brl). Antes embutia `× 1.06`.
    Se o XLSX vier de scanner antigo com a taxa já aplicada, o drift é
    detectado e a margem é reescrita para a base bruta (mantendo o relatório
    coerente com a decisão do operador).
      1. Se `live_brl` + `reference_price_brl` válidos → recomputa
         `recomputed_net = (tcg - live*(1+hub_fee_rate)) / tcg`
      2. Compara com `net_margin` do input
      3. Se drift > 0.5pp → log WARNING, sobrescreve com recomputado
      4. Marca `margin_source = "input" | "recomputed"`

    Não toca quando colunas faltam (input incompleto — deixa subir o "Dados
    insuficientes" do classify_decision).
    """
    if "live_brl" not in df.columns or "reference_price_brl" not in df.columns:
        df["margin_source"] = "input"
        return df

    live = pd.to_numeric(df["live_brl"], errors="coerce")
    tcg = pd.to_numeric(df["reference_price_brl"], errors="coerce")
    valid = live.notna() & tcg.notna() & (tcg > 0) & (live > 0)

    custo = live * (1.0 + hub_fee_rate)
    recomputed_net = (tcg - custo) / tcg
    recomputed_lucro = tcg - custo

    if "net_margin" not in df.columns:
        df["net_margin"] = recomputed_net
        df["lucro_liq"] = recomputed_lucro
        df["margin_source"] = "recomputed"
        print(f"[postprocess v2.1] net_margin AUSENTE no input — recomputado em {valid.sum()} rows.")
        return df

    input_net = pd.to_numeric(df["net_margin"], errors="coerce")
    drift = (input_net - recomputed_net).abs()
    drifting = valid & drift.gt(drift_threshold_pp)

    source = pd.Series(["input"] * len(df), index=df.index)
    if drifting.any():
        n_drift = int(drifting.sum())
        max_drift_pp = float(drift[drifting].max()) * 100
        basis = "BRUTA (sem taxa)" if hub_fee_rate <= 0 else f"+{hub_fee_rate:.0%} hub fee"
        print(
            f"[postprocess v2.12] WARNING: net_margin do input diverge da fórmula "
            f"{basis} em {n_drift}/{int(valid.sum())} rows (drift máx {max_drift_pp:.2f}pp). "
            f"Sobrescrevendo com recomputado."
        )
        # Substitui apenas onde driftando
        df.loc[drifting, "net_margin"] = recomputed_net[drifting]
        df.loc[drifting, "lucro_liq"] = recomputed_lucro[drifting]
        source[drifting] = "recomputed"

    df["margin_source"] = source.values
    return df


def enrich_df(raw_df: pd.DataFrame, hub_fee_rate: float = HUB_FEE_RATE) -> pd.DataFrame:
    df = normalize_columns(raw_df).copy()
    # v2.22 (2026-06-22): fallback de near-miss. Listings precificados ABAIXO do
    # threshold (não validados per-blueprint) chegam SEM "LIVE R$ (real)" nem
    # "Net Margin % REAL" — só com os campos de SCAN ("Scan R$ (raw)" /
    # "Net Margin % (scan)"). Sem este fallback, a linha near-miss vinha na
    # tabela de entrega com Margem%/CT US$ vazios. Preenche `live_brl`/
    # `net_margin` a partir dos campos de scan SÓ onde os REAL faltam (não toca
    # nas linhas validadas COMPRA/REVISAR — essas já têm o valor real).
    scan_brl_col = next((c for c in raw_df.columns if str(c).strip() == "Scan R$ (raw)"), None)
    scan_net_col = next((c for c in raw_df.columns if str(c).strip() == "Net Margin % (scan)"), None)
    if scan_brl_col is not None:
        if "live_brl" not in df.columns:
            df["live_brl"] = pd.to_numeric(raw_df[scan_brl_col], errors="coerce")
        else:
            df["live_brl"] = pd.to_numeric(df["live_brl"], errors="coerce").fillna(
                pd.to_numeric(raw_df[scan_brl_col], errors="coerce")
            )
    if scan_net_col is not None:
        scan_net = pd.to_numeric(raw_df[scan_net_col], errors="coerce")
        if "net_margin" not in df.columns:
            df["net_margin"] = scan_net
        else:
            df["net_margin"] = pd.to_numeric(df["net_margin"], errors="coerce").fillna(scan_net)
    # v2.1 #2 fix: defensive recompute ANTES de classify_decision consumir.
    # v2.12: hub_fee_rate default 0.0 → margem BRUTA.
    df = _recompute_margin_with_fee(df, hub_fee_rate=hub_fee_rate)
    # TG##/GG## gallery flag (Trainer + Galarian Gallery — mesma inflação pokemontcg.io)
    if "card_number" in df.columns:
        df["trainer_gallery_potential_fp"] = df["card_number"].astype(str).str.match(r"^(?:TG|GG)\d+", case=False, na=False)
    else:
        df["trainer_gallery_potential_fp"] = False
    # Chase Tier
    df["chase_tier"] = df.apply(
        lambda r: classify_chase_tier(
            r.get("rarity"), r.get("card_number"), r.get("markup_tier")
        ), axis=1
    )
    # Fundamental Score (derived, after chase_tier)
    df["fundamental_score"] = df.apply(fundamental_score, axis=1)
    return df

def _combine_name_number(df: pd.DataFrame) -> pd.DataFrame:
    """Junta card_name + card_number numa célula só, formato MYP-style.

    PR-K (2026-05-29) — alinhamento com MYP scanner:
      - Com total mapeado: "Plusle (193/197)" (parens + total do set)
      - Sem total mapeado: "Plusle (193)" (parens só, sem total)

    ⚠️ CONTRATO IMPORTANTE (PR-L 2026-05-30): chamado em build_deals_sheet
    e build_all_listings_sheet ANTES do rename. Sheets derivadas (Top 50
    Margin / Validate Manually / TCG Suspect) recebem o df JÁ TRANSFORMADO
    — card_name é "Plusle (193/182)", card_number=193 fica preservado
    separado. Filtros nessas sheets usam Nº ou Set crus, NÃO Carta texto;
    se mudar essa convenção verifique todos os build_*_sheet PR-K.
      - Sem número: "Plusle" (só nome)

    set_code é extraído da coluna 'set_code' (string como "Stellar Crown (scr)"
    ou direto "scr"). Robusto a número float ('182.0'→'182'), ausente,
    alfanumérico (TG12, SV161, 153a) e total ausente.
    """
    if "card_name" not in df.columns or "card_number" not in df.columns:
        return df
    df = df.copy()
    has_set = "set_code" in df.columns

    def _combine(row):
        name = row["card_name"]
        num = row["card_number"]
        if num is None or (isinstance(num, float) and pd.isna(num)):
            return name
        s = str(num).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if not s or s.lower() == "nan":
            return name
        # Resolve total via CT_SET_TOTAL se set_code mapeado
        total = None
        if has_set:
            code = _extract_set_code_from_label(row.get("set_code"))
            total = CT_SET_TOTAL.get(code)
        if total is not None:
            return f"{name} ({s}/{total})"
        return f"{name} ({s})"

    df["card_name"] = df.apply(_combine, axis=1)
    return df

def build_deals_sheet(df: pd.DataFrame, cfg: DecisionConfig) -> pd.DataFrame:
    """Sheet principal: deals com Decisao mecanica (COMPRA + REVISAR, ordenado por lucro)."""
    df = df.copy()
    results = df.apply(lambda r: classify_decision(r, cfg), axis=1)
    df["decisao"] = [r[0] for r in results]
    df["porque"] = [r[1] for r in results]
    # Filter COMPRA + REVISAR (NAO fica em All Listings)
    deals = df[df["decisao"].isin(["COMPRA", "REVISAR"])].copy()
    if "lucro_liq" in deals.columns:
        deals = deals.sort_values("lucro_liq", ascending=False)
    # Renomeia pro display
    display_cols = ["decisao", "porque", "chase_tier", "fundamental_score",
                     "dh_score",  # 2ª opinião DoubleHolo (só se --doubleholo; senão ausente)
                     "set_code", "card_name", "card_number", "language",
                     "live_brl", "reference_price_brl", "net_margin", "lucro_liq",
                     "validation_status", "seller", "link_ct", "link_tcg"]
    display_cols = [c for c in display_cols if c in deals.columns]
    deals = deals[display_cols]
    deals = _combine_name_number(deals)
    rename_map = {
        "decisao": "Decisão", "porque": "Porque", "chase_tier": "Chase Tier",
        "fundamental_score": "Score", "dh_score": "DH", "set_code": "Set", "card_name": "Carta",
        "card_number": "Nº", "language": "Idioma", "live_brl": "Preço CT (R$)",
        "reference_price_brl": "TCG (R$)", "net_margin": "Net %",
        "lucro_liq": "Lucro Líq (R$)", "validation_status": "Validação",
        "seller": "Seller", "link_ct": "Link CT", "link_tcg": "Link TCG",
    }
    return deals.rename(columns=rename_map)

def build_all_listings_sheet(df: pd.DataFrame, cfg: DecisionConfig) -> pd.DataFrame:
    """Sheet completa: TUDO com Decisao + Chase + Score (incluindo NAO)."""
    df = df.copy()
    results = df.apply(lambda r: classify_decision(r, cfg), axis=1)
    df["decisao"] = [r[0] for r in results]
    df["porque"] = [r[1] for r in results]
    display_cols = ["decisao", "porque", "chase_tier", "fundamental_score",
                     "dh_score",  # 2ª opinião DoubleHolo (só se --doubleholo; senão ausente)
                     "set_code", "card_name", "card_number", "language",
                     "live_brl", "reference_price_brl", "net_margin", "lucro_liq",
                     "validation_status", "seller", "link_ct", "link_tcg"]
    display_cols = [c for c in display_cols if c in df.columns]
    df = df[display_cols]
    df = _combine_name_number(df)
    rename_map = {
        "decisao": "Decisão", "porque": "Porque", "chase_tier": "Chase Tier",
        "fundamental_score": "Score", "dh_score": "DH", "set_code": "Set", "card_name": "Carta",
        "card_number": "Nº", "language": "Idioma", "live_brl": "Preço CT (R$)",
        "reference_price_brl": "TCG (R$)", "net_margin": "Net %",
        "lucro_liq": "Lucro Líq (R$)", "validation_status": "Validação",
        "seller": "Seller", "link_ct": "Link CT", "link_tcg": "Link TCG",
    }
    return df.rename(columns=rename_map)

def build_top50_margin_sheet(all_listings: pd.DataFrame) -> pd.DataFrame:
    """PR-K: Top 50 listings por Net % desc (ranking puro do universo All Listings).
    Recebe DataFrame JÁ na forma final (mesmas colunas das outras sheets).
    Aceita ranking puro: qualquer Decisão (inclusive NÃO) — operador vê o teto."""
    if all_listings.empty or "Net %" not in all_listings.columns:
        return all_listings.head(0)
    return all_listings.sort_values("Net %", ascending=False).head(50)


def build_validate_manually_sheet(all_listings: pd.DataFrame) -> pd.DataFrame:
    """PR-K: listings que pedem revisão manual via heurísticas:
      - Carta TG##/GG## (Trainer + Galarian Gallery, ^(?:TG|GG)\\d+) no Nº
      - Validação ∈ {STALE, PRICE_CHANGED, API_ERROR}
      - Set vintage suspect (lc, ba-20, ba-22)
      - Net % > 200% (provável FP)
    """
    if all_listings.empty:
        return all_listings
    df = all_listings.copy()
    mask = pd.Series(False, index=df.index)
    # TG##/GG## gallery
    if "Nº" in df.columns:
        mask = mask | df["Nº"].astype(str).str.match(r"^(?:TG|GG)\d+", case=False, na=False)
    # Validation flags
    if "Validação" in df.columns:
        val_up = df["Validação"].astype(str).str.upper()
        mask = mask | val_up.isin({"STALE", "PRICE_CHANGED", "API_ERROR"})
    # Vintage suspect set
    if "Set" in df.columns:
        codes = df["Set"].astype(str).apply(_extract_set_code_from_label)
        mask = mask | codes.isin(VINTAGE_SUSPECT_SETS)
    # Net% extremo
    if "Net %" in df.columns:
        nm = pd.to_numeric(df["Net %"], errors="coerce")
        mask = mask | nm.gt(EXTREME_NET_PCT)
    return df[mask].sort_values("Net %", ascending=False) if "Net %" in df.columns else df[mask]


def build_tcg_suspect_sheet(all_listings: pd.DataFrame) -> pd.DataFrame:
    """PR-K: listings cujo preço TCG parece inflado (provável per-expansion bug ou
    vintage reverseHolofoil inflation):
      - TCG (R$) > 10× mediana TCG do mesmo Set (proxy de inflação per-blueprint)
      - Set vintage não-coberto bem por pokemontcg.io (mesmo critério)
    """
    if all_listings.empty or "TCG (R$)" not in all_listings.columns:
        return all_listings.head(0)
    df = all_listings.copy()
    tcg = pd.to_numeric(df["TCG (R$)"], errors="coerce")
    mask = pd.Series(False, index=df.index)
    # 10× mediana por Set
    if "Set" in df.columns:
        med = tcg.groupby(df["Set"]).transform("median")
        mask = mask | (tcg > med * 10)
        # Vintage set heuristic
        codes = df["Set"].astype(str).apply(_extract_set_code_from_label)
        mask = mask | codes.isin(VINTAGE_SUSPECT_SETS)
    return df[mask].sort_values("Net %", ascending=False) if "Net %" in df.columns else df[mask]


def build_summary(df: pd.DataFrame, cfg: DecisionConfig) -> pd.DataFrame:
    decisions = df.apply(lambda r: classify_decision(r, cfg), axis=1)
    decisao_col = pd.Series([r[0] for r in decisions])
    total = len(df)
    n_compra = (decisao_col == "COMPRA").sum()
    n_revisar = (decisao_col == "REVISAR").sum()
    n_nao = (decisao_col == "NAO").sum()
    # v2.1 #4: deals exportados = só COMPRA + REVISAR (sheet "Deals" filtra NAO)
    n_deals_exportados = int(n_compra + n_revisar)
    lucro_compra = df.loc[decisao_col == "COMPRA", "lucro_liq"].sum() if "lucro_liq" in df.columns else 0
    rows = [
        ("Total listings escaneados", total),
        ("Total deals exportados (COMPRA + REVISAR)", n_deals_exportados),
        ("COMPRA", f"{n_compra} ({n_compra/total*100:.1f}%)" if total else 0),
        ("REVISAR (zona cinza)", f"{n_revisar} ({n_revisar/total*100:.1f}%)" if total else 0),
        ("NÃO", f"{n_nao} ({n_nao/total*100:.1f}%)" if total else 0),
        ("Lucro líquido potencial (COMPRA)", f"R$ {lucro_compra:,.0f}"),
        ("", ""),
        ("Threshold COMPRA — net margin", f"≥ {cfg.min_net_margin:.0%}"),
        ("Threshold COMPRA — lucro líquido", f"≥ R$ {cfg.min_lucro_liq:.0f}"),
        ("Threshold COMPRA — chase tier", "≥ MID"),
        ("Threshold REVISAR — net margin", f"≥ {cfg.revisar_min_net:.0%}"),
        ("Threshold MODEST → REVISAR", f"net ≥ {cfg.revisar_chase_modest_min_net:.0%}"),
        ("", ""),
        ("Math: custo total",
         "preço_CT (margem BRUTA — sem taxa; operador soma fees por fora)"
         if cfg.hub_fee_rate <= 0
         else f"preço_CT × {1 + cfg.hub_fee_rate:.2f} (hub fee {cfg.hub_fee_rate:.0%})"),
        ("Math: lucro líquido", "TCG_BRL − custo_total"),
        ("Math: net margin", "lucro_líquido / TCG_BRL"),
    ]
    return pd.DataFrame(rows, columns=["Métrica", "Valor"])

def style_sheet(ws, df, decisao_col=None, chase_col=None):
    # Header style
    for ci in range(1, len(df.columns) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    # Color rows by Decisao
    if decisao_col and decisao_col in df.columns:
        ci = list(df.columns).index(decisao_col) + 1
        for ri in range(2, len(df) + 2):
            cell = ws.cell(row=ri, column=ci)
            if cell.value in DECISAO_FILL:
                cell.fill = DECISAO_FILL[cell.value]
                cell.font = Font(bold=True)
    # Color Chase Tier column
    if chase_col and chase_col in df.columns:
        ci = list(df.columns).index(chase_col) + 1
        for ri in range(2, len(df) + 2):
            cell = ws.cell(row=ri, column=ci)
            if cell.value in CHASE_FILL:
                cell.fill = CHASE_FILL[cell.value]
                cell.font = Font(bold=True)
    # Column widths
    for ci, cn in enumerate(df.columns, 1):
        try:
            vals = [str(v) for v in df.iloc[:50, ci-1].tolist()]
        except Exception:
            vals = []
        ml = max([len(str(cn))] + [len(v) for v in vals]) if vals else len(str(cn))
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 50)
    # Hyperlinks na coluna Carta
    apply_card_hyperlinks(ws, df)

def _safe_val(val):
    """Módulo-level (PR-L 2026-05-30, CT-M4): NaN/None → None (openpyxl
    escreve célula vazia, evita 'nan' literal). Aplicado em TODAS as
    sheets (Deals/All Listings/PR-K/Summary) pra uniformidade defensiva."""
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


# ─── Tabela de ENTREGA no chat (markdown, links clicáveis) ───────────────────
# Formato aprovado pelo operador 2026-06-09 (paridade com o scanner COMC):
# entrega = tabela markdown no chat, NÃO planilha. CSV/XLSX seguem com colunas
# separadas + URLs cruas; só esta tabela compõe Carta (nome+número) e Links.
#
#   | # | Margem % | CT US$ | TCG US$ | Dif | Carta | Set | Raridade | Cond | Qtd | Flag | Links |
#
# Coluna Links = "[oferta](url_ct) · [TCG](url_tcg)" — o link da oferta aponta
# pra página do CardTrader (de preferência per-blueprint, preço final com markup
# — lembrar dos ~76% de falsos positivos sem validação per-blueprint); o link TCG
# é o workflow canônico de validação manual do operador.
#
# Coluna Flag = sinaliza, por linha, o que a regra MECÂNICA (classify_decision)
# concluiu: COMPRA → "" (limpo); REVISAR → "validar manual" (zona cinza /
# suspeito de margem inflada — TG##, alpha-suffix promo/league, set sem cobertura
# confiável, markup anômalo, etc.). Reusa a MESMA classificação do XLSX, não
# duplica regra. NÃO é recomendação de compra — é o flag de cautela que o
# operador usa pra decidir o que conferir no Link TCG antes.
#
# Margem é a net_margin já enriquecida = margem BRUTA (regra 2026-06-06, sem Hub
# fee). Esta função NÃO altera threshold/filtro/classificação — só APRESENTA os
# deals que classify_decision marcou COMPRA/REVISAR (TODOS eles, até top_n).
_DELIVERY_HEADERS = [
    "#", "Margem %", "CT US$", "TCG US$", "Dif",
    "Carta", "Set", "Raridade", "Cond", "Qtd", "Flag", "Links",
]


def _fmt_usd(v) -> str:
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return ""


def _fmt_pct(v) -> str:
    """net_margin chega em fração (0.30 = 30%) → '30%'."""
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return f"{float(v) * 100:.0f}%"
    except (TypeError, ValueError):
        return ""


def _md_links_cell(link_ct, link_tcg) -> str:
    """'[oferta](url_ct) · [TCG](url_tcg)' — só inclui o que existir."""
    parts = []
    ct = "" if link_ct is None else str(link_ct).strip()
    tcg = "" if link_tcg is None else str(link_tcg).strip()
    if ct.startswith("http"):
        parts.append(f"[oferta]({ct})")
    if tcg.startswith("http"):
        parts.append(f"[TCG]({tcg})")
    return " · ".join(parts)


def _md_escape(s) -> str:
    """Pipe quebra a tabela markdown → vira '/'. None → ''."""
    if s is None:
        return ""
    try:
        if isinstance(s, float) and pd.isna(s):
            return ""
    except (TypeError, ValueError):
        pass
    return str(s).replace("|", "/").strip()


def _set_name_from_label(set_label) -> str:
    """'Paradox Rift (par)' → 'Paradox Rift'; sem parênteses → o próprio label.
    Usado como nome do set p/ o fallback-por-nome do resolver tcgcsv."""
    if not set_label:
        return ""
    s = str(set_label).strip()
    return s.rsplit(" (", 1)[0].strip() if " (" in s else s


def attach_product_ids(df: pd.DataFrame, resolver,
                       set_col: str = "set_code", num_col: str = "card_number",
                       variant_col: str = "variant",  # canonical pós-normalize_columns (era "Variant" → nunca casava)
                       url_col: str = "link_tcg",
                       resolve_mask=None) -> int:
    """Adiciona a coluna `tcg_product_id` (str ou None) por linha — chave de join DH.

    Ordem (menos invasiva 1º): (1) se `link_tcg` já é tcgplayer.com/product/<id>
    (sets via tcgcsv — Fix(1) do scanner) usa esse id direto; (2) senão resolve
    OFFLINE via tcgcsv por (set CT, número, variante priceada) — Fix(2). Linha
    sem resolução UNÍVOCA → None (DH = '—', honesto). NÃO toca preço/margem; o
    productId é só identidade p/ link/join. `resolver=None` → só o passo (1).

    `resolve_mask` (follow-up #4): sequência booleana alinhada às linhas de `df`;
    quando passada, o passo (2) (resolver OFFLINE, que faz I/O via tcgcsv) só roda
    nas linhas marcadas True — tipicamente só as que viram deal (COMPRA/REVISAR).
    Num `--all-sets` (milhares de linhas, a maioria NÃO) isso evita o I/O nas
    linhas que nunca aparecem na entrega. O passo (1) (productId direto do link,
    SEM I/O) continua valendo p/ TODAS as linhas. `None` = resolve todas (antigo).
    """
    have_url = url_col in df.columns
    mask = list(resolve_mask) if resolve_mask is not None else None
    pids = []
    n = 0
    for pos, (_, row) in enumerate(df.iterrows()):
        pid = doubleholo_join.extract_product_id(row.get(url_col)) if have_url else None
        allow_resolve = mask is None or (pos < len(mask) and bool(mask[pos]))
        if pid is None and resolver is not None and allow_resolve:
            label = row.get(set_col)
            try:
                pid = resolver.resolve(
                    _extract_set_code_from_label(label),
                    _set_name_from_label(label),
                    row.get(num_col),
                    row.get(variant_col) if variant_col in df.columns else None,
                )
            except Exception:  # noqa: BLE001 — resolução é best-effort; falha → '—'
                pid = None
        pids.append(pid)
        if pid is not None:
            n += 1
    df["tcg_product_id"] = pids
    return n


def _delivery_resolve_mask(df: pd.DataFrame, cfg: DecisionConfig, top_md: int):
    """Máscara booleana (alinhada às linhas de `df`) das linhas que aparecem na
    ENTREGA markdown COM coluna DH — o conjunto p/ o qual vale rodar o resolver
    OFFLINE (Fix(2), que faz I/O). ESPELHA a seleção de `build_delivery_markdown`:
    os deals (COMPRA/REVISAR) quando existirem; senão (near-miss) os `top_md`
    candidatos por margem. Sem cobrir o near-miss, quando NÃO há deal o resolver
    rodaria em ZERO linhas e a tabela near-miss entregue perderia a coluna DH."""
    decisao = df.apply(lambda r: classify_decision(r, cfg)[0], axis=1)
    is_deal = decisao.isin(["COMPRA", "REVISAR"])
    if bool(is_deal.any()):
        return is_deal
    mask = pd.Series(False, index=df.index)
    if "net_margin" in df.columns and len(df):
        top_idx = df["net_margin"].sort_values(ascending=False).head(top_md).index
        mask.loc[top_idx] = True
    else:
        mask.loc[:] = True  # sem margem p/ ordenar → não arrisca, resolve todas
    return mask


def _fmt_dh(v) -> str:
    """Nota DH (0-100) → string; None/NaN → '—' (sem dado Double Holo, honesto)."""
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "—"
        return f"{int(round(float(v)))}"
    except (TypeError, ValueError):
        return "—"


def build_delivery_markdown(
    df: pd.DataFrame,
    cfg: DecisionConfig,
    fx_usd_brl: float | None = None,
    top_n: int = 50,
    show_dh: bool = False,
) -> str:
    """Monta a tabela markdown de entrega (chat-first) a partir do df ENRIQUECIDO.

    Espera `df` já passado por `enrich_df` (colunas lógicas + net_margin/chase_tier).
    Reaproveita `classify_decision` p/ filtrar COMPRA+REVISAR (mesma classificação
    do XLSX — NÃO duplica regra), `_combine_name_number` p/ a coluna 'Carta'.

    USD:
      - "TCG US$" = `reference_price_usd` (nativo do raw, TCG Market USD).
      - "CT US$"  = `live_usd` se existir, senão `live_brl / fx_usd_brl`.
      - "Dif"     = TCG US$ − CT US$ (lacuna bruta em dólar).
    `fx_usd_brl` vem do Stats sheet do raw (usd_brl_rate). Sem FX nem live_usd,
    a célula CT US$ fica vazia (não inventa câmbio).
    """
    work = df.copy()
    decisions = work.apply(lambda r: classify_decision(r, cfg), axis=1)
    work["decisao"] = [d[0] for d in decisions]
    deals = work[work["decisao"].isin(["COMPRA", "REVISAR"])].copy()
    # Fallback near-miss: sem COMPRA/REVISAR, a entrega NÃO vira um beco
    # "nenhum deal" (que historicamente levava a montar tabela à mão, fora do
    # padrão). Em vez disso, mostra os candidatos mais próximos por margem, no
    # MESMO formato canônico, marcados "abaixo do limiar". Garante que a entrega
    # seja SEMPRE a tabela da ferramenta (modelo MYP + coluna Links combinada).
    near_miss = deals.empty
    if near_miss:
        deals = work.copy()
    if "net_margin" in deals.columns:
        deals = deals.sort_values("net_margin", ascending=False)
    deals = deals.head(top_n)
    deals = _combine_name_number(deals)  # 'card_name' vira "Nome (NNN/Total)"

    title = (
        f"### CardTrader — entrega (top {len(deals)} por margem · "
        f"margem BRUTA, threshold {cfg.min_net_margin:.0%})"
    )
    if deals.empty:
        # Sob o contrato de entrega v2.22, o scanner persiste TODO listing
        # precificado no XLSX (mesmo abaixo do threshold). Logo, df vazio aqui
        # significa MESMO "0 precificado" (set sem cobertura TCG / 0 listing
        # passou os filtros NM/EN/preço), não "precificou mas nada bateu o
        # threshold" — esse caso vira a tabela near-miss acima.
        return title + (
            "\n\n_(0 listing precificado — nada a entregar. Nenhuma carta passou "
            "os filtros (NM/EN/≥preço) com preço de referência TCG. Não confundir "
            "com 'precificou mas 0 acima do threshold' — esse caso mostra a tabela "
            "near-miss.)_"
        )

    # Coluna DH = 2ª opinião Double Holo (EXTRA, condicional à --doubleholo).
    # Inserida logo após "Margem %" (perto da margem), sem remover/reordenar as
    # colunas existentes nem colapsar a coluna Links — contrato de 2 links/linha
    # intacto. Só aparece se a coluna `dh_score` foi anexada (flag passada).
    show_dh = show_dh and "dh_score" in deals.columns
    headers = list(_DELIVERY_HEADERS)
    if show_dh:
        headers.insert(2, "DH")  # após "#" (0) e "Margem %" (1)
    header = "| " + " | ".join(headers) + " |"
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    lines = [title]
    if near_miss:
        lines.append(
            "\n_⚠️ Nenhum deal acima do limiar — mostrando os candidatos mais "
            "próximos por margem (todos ABAIXO do limiar, só referência)._"
        )
    lines += ["", header, sep]

    def _ct_usd(row):
        # live_usd direto se o raw trouxer; senão converte BRL via FX.
        for k in ("live_usd", "ct_price_usd"):
            if k in row and pd.notna(row.get(k)):
                return float(row[k])
        live_brl = row.get("live_brl")
        if fx_usd_brl and pd.notna(live_brl) and float(fx_usd_brl) > 0:
            return float(live_brl) / float(fx_usd_brl)
        return None

    for rank, (_, row) in enumerate(deals.iterrows(), 1):
        ct_usd = _ct_usd(row)
        tcg_usd = row.get("reference_price_usd")
        try:
            tcg_usd = float(tcg_usd) if pd.notna(tcg_usd) else None
        except (TypeError, ValueError):
            tcg_usd = None
        dif = (tcg_usd - ct_usd) if (ct_usd is not None and tcg_usd is not None) else None
        # Flag por linha: near-miss → "abaixo do limiar"; REVISAR (zona cinza /
        # suspeito de margem inflada) → "validar manual"; COMPRA → célula limpa.
        # Mesma classificação do XLSX.
        if near_miss:
            flag = "abaixo do limiar"
        else:
            flag = "validar manual" if str(row.get("decisao")) == "REVISAR" else ""
        cells = [
            str(rank),
            _fmt_pct(row.get("net_margin")),
            _fmt_usd(ct_usd),
            _fmt_usd(tcg_usd),
            _fmt_usd(dif),
            _md_escape(row.get("card_name")),
            _md_escape(row.get("set_code")),
            _md_escape(row.get("rarity")),
            _md_escape(row.get("condition")),
            _md_escape(row.get("quantity")),
            flag,
            _md_links_cell(row.get("link_ct"), row.get("link_tcg")),
        ]
        if show_dh:
            cells.insert(2, _fmt_dh(row.get("dh_score")))
        lines.append("| " + " | ".join(cells) + " |")
    if show_dh:
        lines.append(
            "\n_DH = 2ª opinião Double Holo 0-100 (50=neutro), não entra na "
            "margem/decisão; '—' = sem dado._"
        )
    return "\n".join(lines)


def _read_fx_usd_brl(input_path: Path) -> float | None:
    """Lê usd_brl_rate da aba Stats do raw XLSX do scanner (pra converter CT US$).

    Retorna None silenciosamente se a aba/linha não existir (input de origem
    diferente, raw antigo). Sem FX, a coluna CT US$ fica vazia — não inventa."""
    try:
        stats = pd.read_excel(input_path, sheet_name="Stats", header=None)
    except Exception:
        return None
    for _, r in stats.iterrows():
        if str(r.iloc[0]).strip().lower() == "usd_brl_rate":
            try:
                return float(r.iloc[1])
            except (TypeError, ValueError):
                return None
    return None


def write_report(df: pd.DataFrame, cfg: DecisionConfig, output_path: Path,
                 fx_usd_brl: float | None = None, top_md: int = 50,
                 dh_signals: dict | None = None, pid_resolver=None) -> str:
    df = enrich_df(df, hub_fee_rate=cfg.hub_fee_rate)
    # Caminho 1 DoubleHolo: anexa a coluna `dh_score` (2ª opinião). Só quando
    # --doubleholo foi passado; sem a flag a coluna não existe e a saída é
    # idêntica. NÃO toca margem/decisão.
    if dh_signals is not None:
        # 1º resolve o productId TCGplayer por linha (link tcgcsv direto via
        # Fix(1) OU ponte offline tcgcsv via Fix(2)); 2º casa a nota DH por ele.
        # Follow-up #4: o resolver OFFLINE (I/O) só roda nas linhas da ENTREGA
        # markdown — os deals (COMPRA/REVISAR) ou, sem nenhum, os candidatos
        # near-miss (ver _delivery_resolve_mask). Num --all-sets a maioria é NÃO e
        # nunca entra na entrega, então evita o I/O nelas. O Fix(1) (productId do
        # link direto, SEM I/O) segue valendo p/ TODAS as linhas — logo a coluna DH
        # do XLSX "All Listings" tem cobertura best-effort (deal/near-miss + linhas
        # com link tcgplayer.com/product direto); linha NÃO-deal sem link de produto
        # mostra "—" (honesto: DH é 2ª opinião de deal, não de listing rejeitado).
        deliver_mask = _delivery_resolve_mask(df, cfg, top_md)
        resolved = attach_product_ids(df, pid_resolver, resolve_mask=deliver_mask)
        matched = doubleholo_join.attach_scores_df(
            df, dh_signals, url_col="link_tcg", pid_col="tcg_product_id")
        # Follow-up #5: "casaram" (productId no índice) ≠ "com nota DH" (dh_score
        # não-None) — separa as duas contagens p/ a telemetria não inflar.
        scored = int(df["dh_score"].notna().sum())
        print(f"[DH] productId resolvido em {resolved}/{len(df)} linhas; "
              f"{matched} casaram com o DoubleHolo, {scored} com nota DH "
              f"(sem match ou sem nota → '—').")
    wb = Workbook(); wb.remove(wb.active)

    deals = build_deals_sheet(df, cfg)
    ws = wb.create_sheet("Deals")
    if deals.empty:
        ws.cell(row=1, column=1, value="Nenhum deal (COMPRA ou REVISAR) encontrado.")
    else:
        for ri, row in enumerate([deals.columns.tolist()] + deals.values.tolist(), 1):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=_safe_val(val) if ri > 1 else val)
        style_sheet(ws, deals, decisao_col="Decisão", chase_col="Chase Tier")

    all_l = build_all_listings_sheet(df, cfg)
    ws = wb.create_sheet("All Listings")
    if all_l.empty:
        ws.cell(row=1, column=1, value="Vazio.")
    else:
        for ri, row in enumerate([all_l.columns.tolist()] + all_l.values.tolist(), 1):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=_safe_val(val) if ri > 1 else val)
        style_sheet(ws, all_l, decisao_col="Decisão", chase_col="Chase Tier")

    # ─── PR-K: 3 sheets MYP-style adicionais ────────────────────────────────
    # Derivadas de All Listings — preservam mesmas colunas, só mudam filtro/sort.
    def _write_styled(sheet_name: str, sheet_df: pd.DataFrame):
        ws = wb.create_sheet(sheet_name)
        if sheet_df.empty:
            ws.cell(row=1, column=1, value="Nenhum listing nesta categoria.")
            return
        for ri, row in enumerate([sheet_df.columns.tolist()] + sheet_df.values.tolist(), 1):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=_safe_val(val) if ri > 1 else val)
        style_sheet(ws, sheet_df, decisao_col="Decisão", chase_col="Chase Tier")

    _write_styled("Top 50 Margin", build_top50_margin_sheet(all_l))
    _write_styled("Validate Manually", build_validate_manually_sheet(all_l))
    _write_styled("TCG Suspect", build_tcg_suspect_sheet(all_l))

    summary = build_summary(df, cfg)
    ws = wb.create_sheet("Summary")
    for ri, row in enumerate([summary.columns.tolist()] + summary.values.tolist(), 1):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=_safe_val(val) if ri > 1 else val)
    # Style summary
    for ci in range(1, len(summary.columns) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
    for ci, cn in enumerate(summary.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 40

    # Garante o diretório-alvo antes de salvar (outputs/ é gitignored → ausente
    # num clone limpo): sem isto o wb.save() quebra com FileNotFoundError.
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"OK: {output_path}")

    # ─── Entrega no chat: tabela markdown (links clicáveis) ──────────────────
    # Formato aprovado 2026-06-09. A entrega ao operador é a TABELA no chat;
    # o .md sidecar é só conveniência (mesmo conteúdo). XLSX segue cru/colunar.
    md = build_delivery_markdown(df, cfg, fx_usd_brl=fx_usd_brl, top_n=top_md,
                                 show_dh=dh_signals is not None)
    md_path = output_path.with_suffix(".md")
    try:
        md_path.write_text(md + "\n", encoding="utf-8")
        print(f"OK: {md_path} (tabela de entrega — copie pro chat)")
    except OSError as e:
        print(f"[postprocess] aviso: não gravou {md_path} ({e}); tabela só no stdout.")
    print("\n" + md + "\n")
    return md

def main():
    p = argparse.ArgumentParser(description="CardTrader postprocess v2.0 — simplificado")
    p.add_argument("--input", "-i", required=True, help="XLSX raw do scanner")
    p.add_argument("--output", "-o", required=True, help="XLSX relatorio destino")
    p.add_argument("--min-net-margin", type=float, default=0.25,
                   help="Net margin minimo pra COMPRA (default 0.25)")
    p.add_argument("--min-lucro", type=float, default=50.0,
                   help="Lucro liquido R$ minimo pra COMPRA (default 50)")
    p.add_argument("--revisar-min-net", type=float, default=0.20,
                   help="Net margin minimo pra zona REVISAR (default 0.20)")
    p.add_argument("--revisar-modest-min", type=float, default=0.30,
                   help="MODEST so vai REVISAR se net >= X (default 0.30)")
    p.add_argument("--hub-fee", type=float, default=0.0,
                   help=("v2.12: DEFAULT 0.0 — margem BRUTA (custo = preço do "
                         "site, SEM taxa). Operador soma Hub fee/frete/cartão/IOF "
                         "por fora. Passe 0.06 pra reembutir os 6%% históricos."))
    p.add_argument("--top-md", type=int, default=50,
                   help=("Quantas linhas na tabela de entrega markdown do chat "
                         "(default 50). XLSX sempre traz todos os deals."))
    p.add_argument("--doubleholo", default=None, metavar="JSON",
                   help=("Caminho do JSON canônico do DoubleHolo (saída de "
                         "doubleholo_signals.py ingest --json). Adiciona a coluna "
                         "DH (2ª opinião 0-100, 50=neutro) por productId TCGplayer. "
                         "NÃO entra na margem/decisão. Sem a flag, saída idêntica."))
    p.add_argument("--no-pid-resolve", action="store_true",
                   help=("Desliga a resolução OFFLINE de productId via tcgcsv "
                         "(Fix(2)) usada p/ casar DH nas linhas via pokemontcg.io. "
                         "Só tem efeito junto com --doubleholo; sem ela, só casam "
                         "linhas cujo Link TCG já é tcgplayer.com/product/<id>."))
    args = p.parse_args()

    # Paridade com o scanner: aceita `6` ou `0.06` (auto-converte percentual).
    if args.hub_fee > 1.0:
        print(f"[postprocess v2.12] --hub-fee {args.hub_fee} > 1.0 parece "
              f"percentual; convertendo para fração: {args.hub_fee/100}")
        args.hub_fee = args.hub_fee / 100.0

    cfg = DecisionConfig(
        min_net_margin=args.min_net_margin,
        min_lucro_liq=args.min_lucro,
        revisar_min_net=args.revisar_min_net,
        revisar_chase_modest_min_net=args.revisar_modest_min,
        hub_fee_rate=args.hub_fee,
    )
    df = pd.read_excel(args.input)
    print(f"Carregado: {args.input} | {len(df)} rows | cols: {len(df.columns)}")
    # FX usd_brl da aba Stats do raw → converte CT (BRL) p/ "CT US$" na tabela.
    fx = _read_fx_usd_brl(Path(args.input))
    if fx:
        print(f"FX usd_brl_rate (Stats) = {fx:.4f} — usado p/ coluna CT US$.")
    else:
        print("FX usd_brl_rate ausente no input — coluna CT US$ ficará vazia.")
    dh_signals = None
    pid_resolver = None
    if args.doubleholo:
        # A coluna DH é um EXTRA não-essencial: arquivo faltando/malformado NÃO
        # pode derrubar a entrega dos deals (espelha run_outlook.py, que degrada
        # com "--doubleholo ignorado"). Falhou → segue sem DH, saída idêntica.
        try:
            dh_signals = doubleholo_join.load_signals(args.doubleholo)
            print(f"DoubleHolo: {len(dh_signals)} registros com productId TCGplayer "
                  f"carregados de {args.doubleholo} — coluna DH ativa.")
        except Exception as e:  # noqa: BLE001 — DH é opcional; não aborta os deals
            dh_signals = None
            print(f"[DH] aviso: --doubleholo ignorado ({e}); entrega segue sem a "
                  f"coluna DH.")
        if dh_signals is not None and not args.no_pid_resolve:
            # Resolver OFFLINE de productId p/ linhas via pokemontcg.io (Fix(2)).
            # Import lazy: só paga o custo (e o import do scanner) quando há DH.
            try:
                import tcgcsv_productid
                pid_resolver = tcgcsv_productid.ProductIdResolver()
            except Exception as e:  # noqa: BLE001 — sem resolver → cai p/ Fix(1) só
                print(f"[DH] aviso: resolver de productId indisponível ({e}); "
                      f"só linhas com link tcgplayer.com/product casam.")
    write_report(df, cfg, Path(args.output), fx_usd_brl=fx, top_md=args.top_md,
                 dh_signals=dh_signals, pid_resolver=pid_resolver)

if __name__ == "__main__":
    main()
