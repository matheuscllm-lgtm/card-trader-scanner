#!/usr/bin/env python3
"""
test_checkpoint_crash_recovery.py — v2.6 regression (2026-05-17)

Verifica que o fluxo crash-recovery do scanner v2.6 funciona:
  1. CheckpointWriter grava JSONL append-only durante scan parcial
  2. Linha final pode ficar truncada (simula crash mid-write)
  3. recover_from_checkpoint.py parseia, ignora linha bad, reconstrói
     XLSX equivalente com opportunities dos sets que terminaram

NÃO faz network call. Usa CheckpointWriter direto + fixture sintética.
Pra integration test E2E com scanner real, ver smoke_checkpoint.ps1.

Roda:
    .venv/Scripts/python.exe scripts/test_checkpoint_crash_recovery.py

Exit code:
    0 = todos os asserts passaram
    1 = pelo menos um assert falhou
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

os.environ.setdefault("PYTHONIOENCODING", "utf-8")
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

from cardtrader_scanner import (  # noqa: E402
    CheckpointWriter,
    Listing,
    Opportunity,
)


def _make_listing(blueprint_id: int, set_code: str, name: str, price_brl: float) -> Listing:
    return Listing(
        product_id=blueprint_id * 1000,
        blueprint_id=blueprint_id,
        card_name=name,
        set_code=set_code,
        set_name=f"Set {set_code.upper()}",
        collector_number="42",
        condition="Near Mint",
        language="en",
        price_cents=int(price_brl * 100),
        price_currency="BRL",
        price_brl=price_brl,
        quantity=1,
        foil=False,
        graded=False,
        seller_username="testseller",
        seller_can_sell_via_hub=True,
        seller_user_type="zero_fee",
        cardtrader_url=f"https://www.cardtrader.com/cards/{blueprint_id}",
        rarity="Rare",
    )


def _make_opp(blueprint_id: int, set_code: str, name: str, ct_brl: float, tcg_usd: float) -> Opportunity:
    l = _make_listing(blueprint_id, set_code, name, ct_brl)
    tcg_brl = tcg_usd * 5.05  # FX 2026-05 approx
    custo = ct_brl * 1.06
    margin = (tcg_brl - custo) / tcg_brl
    return Opportunity(
        listing=l,
        tcg_market_usd=tcg_usd,
        tcg_market_brl=tcg_brl,
        ct_price_brl=ct_brl,
        margin_pct=margin,
        margin_brl=tcg_brl - custo,
        estimated_shipping_brl=0.0,
        net_margin_pct=margin,
    )


class TestRunner:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.tmpdir: Path = Path(tempfile.mkdtemp(prefix="ct_checkpoint_test_"))

    def assertion(self, label: str, ok: bool, detail: str = ""):
        if ok:
            self.passed += 1
            print(f"  [PASS] {label}")
        else:
            self.failed += 1
            print(f"  [FAIL] {label}  -- {detail}")

    def cleanup(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    # ─────────────────────────────────────────────────────────────────
    def test_writer_writes_header_opps_setcomplete(self):
        print("\n[1] CheckpointWriter writes header + 3 opps + set_complete")
        path = self.tmpdir / "test1.checkpoint.jsonl"
        cw = CheckpointWriter(path, every_n=1)
        cw.write_header({"sets": ["a", "b"], "threshold": 0.3}, total_sets=2)

        opp_a1 = _make_opp(1001, "sfa", "Mawile", 50.0, 30.0)
        opp_a2 = _make_opp(1002, "sfa", "Charizard", 200.0, 100.0)
        cw.write_opportunity(opp_a1)
        cw.write_opportunity(opp_a2)
        cw.write_set_complete("sfa", "Surging Sparks",
                               {"blueprints": 100, "filtered": 50, "priced": 30, "opps_found": 2},
                               elapsed_s=42.5)

        opp_b1 = _make_opp(2001, "scr", "Bulbasaur", 10.0, 8.0)
        cw.write_opportunity(opp_b1)
        cw.write_set_complete("scr", "Stellar Crown",
                               {"blueprints": 80, "filtered": 40, "priced": 20, "opps_found": 1},
                               elapsed_s=33.1)
        cw.write_scan_complete(total_opps=3, total_elapsed_s=75.6)
        cw.close()

        # Re-read e valida
        lines = path.read_text(encoding="utf-8").strip().split("\n")
        self.assertion("file has 7 lines (header + 3 opps + 2 setcomplete + scancomplete)",
                       len(lines) == 7,
                       detail=f"got {len(lines)} lines")

        types = [json.loads(l)["_type"] for l in lines]
        expected = ["scan_header", "opportunity", "opportunity", "set_complete",
                    "opportunity", "set_complete", "scan_complete"]
        self.assertion("line type order is correct",
                       types == expected,
                       detail=f"got {types}")

        # Per-opp denormalização
        opp_line_1 = json.loads(lines[1])
        self.assertion("opportunity line has top-level set_code",
                       opp_line_1.get("set_code") == "sfa",
                       detail=f"set_code={opp_line_1.get('set_code')}")
        self.assertion("opportunity line has top-level blueprint_id",
                       opp_line_1.get("blueprint_id") == 1001,
                       detail=f"blueprint_id={opp_line_1.get('blueprint_id')}")
        self.assertion("opportunity line has top-level name",
                       opp_line_1.get("name") == "Mawile",
                       detail=f"name={opp_line_1.get('name')}")
        # Nested listing presente
        self.assertion("opportunity nested listing has price_brl",
                       opp_line_1["listing"]["price_brl"] == 50.0)
        # Campos críticos
        self.assertion("opportunity has tcg_market_usd",
                       "tcg_market_usd" in opp_line_1)
        self.assertion("opportunity has margin_pct",
                       "margin_pct" in opp_line_1)

    def test_recovery_from_clean_checkpoint(self):
        print("\n[2] Recovery from clean checkpoint produces XLSX")
        path = self.tmpdir / "test2.checkpoint.jsonl"
        cw = CheckpointWriter(path, every_n=1)
        cw.write_header({"threshold": 0.25}, total_sets=2)
        cw.write_opportunity(_make_opp(3001, "sfa", "Pikachu", 40.0, 25.0))
        cw.write_opportunity(_make_opp(3002, "sfa", "Mewtwo", 100.0, 80.0))
        cw.write_set_complete("sfa", "Surging Sparks", {"opps_found": 2}, 10.0)
        cw.write_opportunity(_make_opp(4001, "scr", "Eevee", 60.0, 40.0))
        cw.write_set_complete("scr", "Stellar Crown", {"opps_found": 1}, 8.0)
        cw.write_scan_complete(total_opps=3, total_elapsed_s=20.0)
        cw.close()

        out_xlsx = self.tmpdir / "test2.recovered.xlsx"
        result = subprocess.run(
            [sys.executable, str(REPO_ROOT / "scripts" / "recover_from_checkpoint.py"),
             "--checkpoint", str(path),
             "--output", str(out_xlsx)],
            capture_output=True, text=True, env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        self.assertion("recover_from_checkpoint exit code 0",
                       result.returncode == 0,
                       detail=f"stderr={result.stderr[:200]}")
        self.assertion("recovered XLSX file exists",
                       out_xlsx.exists(),
                       detail=str(out_xlsx))
        self.assertion("recovered XLSX has size > 0",
                       out_xlsx.exists() and out_xlsx.stat().st_size > 0)
        # Output mentions correct opp count
        stdout_lower = result.stdout.lower() + result.stderr.lower()
        self.assertion("log mentions '3 opportunities'",
                       "3 opportunities" in stdout_lower,
                       detail=f"log fragment: {(result.stderr + result.stdout)[-200:]}")

    def test_recovery_preserves_fx_from_header(self):
        """v2.14 (candidato 3): FX gravado no scan_header é reconstruído na
        célula Stats `usd_brl_rate` do XLSX recuperado. Sem isso, ficava 0.0 e
        a tabela de entrega do postprocess (coluna CT US$) ficava vazia."""
        print("\n[5] Recovery reconstructs usd_brl_rate from header FX")
        from openpyxl import load_workbook  # local import — só onde precisa
        path = self.tmpdir / "test5.checkpoint.jsonl"
        cw = CheckpointWriter(path, every_n=1)
        cw.write_header({"threshold": 0.3}, total_sets=1,
                        usd_brl=5.4321, eur_brl=5.9876)
        cw.write_opportunity(_make_opp(7001, "sfa", "Lugia", 70.0, 50.0))
        cw.write_set_complete("sfa", "Surging Sparks", {"opps_found": 1}, 9.0)
        cw.write_scan_complete(total_opps=1, total_elapsed_s=9.0)
        cw.close()

        out_xlsx = self.tmpdir / "test5.recovered.xlsx"
        result = subprocess.run(
            [sys.executable, str(REPO_ROOT / "scripts" / "recover_from_checkpoint.py"),
             "--checkpoint", str(path), "--output", str(out_xlsx)],
            capture_output=True, text=True, env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        self.assertion("recover exit 0", result.returncode == 0,
                       detail=f"stderr={result.stderr[:200]}")
        # Lê a célula Stats usd_brl_rate do XLSX recuperado.
        wb = load_workbook(out_xlsx, read_only=True)
        found_rate = None
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                if row and str(row[0]).strip().lower() == "usd_brl_rate":
                    found_rate = row[1]
                    break
            if found_rate is not None:
                break
        wb.close()
        self.assertion("recovered XLSX has usd_brl_rate cell", found_rate is not None,
                       detail="usd_brl_rate não encontrado em nenhuma sheet")
        self.assertion("usd_brl_rate == 5.4321 (from header, not 0.0)",
                       found_rate is not None and abs(float(found_rate) - 5.4321) < 1e-6,
                       detail=f"got {found_rate}")

    def test_recovery_with_partial_last_line(self):
        print("\n[3] Recovery handles truncated last line (simulated crash)")
        path = self.tmpdir / "test3.checkpoint.jsonl"
        cw = CheckpointWriter(path, every_n=1)
        cw.write_header({"threshold": 0.3}, total_sets=3)
        cw.write_opportunity(_make_opp(5001, "sfa", "Snorlax", 80.0, 60.0))
        cw.write_set_complete("sfa", "Surging Sparks", {"opps_found": 1}, 12.0)
        cw.write_opportunity(_make_opp(5002, "scr", "Gengar", 90.0, 70.0))
        cw.close()
        # Simula crash: appenda linha PARCIAL (sem newline final + JSON truncado)
        with open(path, "a", encoding="utf-8") as f:
            f.write('{"_type": "opportunity", "listing": {"product_id": 9999, "blueprint')
            # NÃO escreve \n nem fecha o JSON. Esta linha vai falhar parse.
        # Set de scr nunca completou + scan_complete NUNCA escrito.

        out_xlsx = self.tmpdir / "test3.recovered.xlsx"
        result = subprocess.run(
            [sys.executable, str(REPO_ROOT / "scripts" / "recover_from_checkpoint.py"),
             "--checkpoint", str(path),
             "--output", str(out_xlsx)],
            capture_output=True, text=True, env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        self.assertion("recovery survives truncated last line (exit 0)",
                       result.returncode == 0,
                       detail=f"stderr={result.stderr[:300]}")
        # Não temos scan_complete → warning
        out_log = result.stdout + result.stderr
        # Captured subprocess stderr pode vir mojibake'd no Windows console (cp1252)
        # — verificamos ASCII fragments que sobrevivem ao re-encoding.
        self.assertion("warns scan did not finish cleanly",
                       "scan_complete" in out_log.lower() or "Recovery" in out_log,
                       detail=f"log tail: {out_log[-400:]}")
        self.assertion("XLSX generated despite partial line",
                       out_xlsx.exists() and out_xlsx.stat().st_size > 0)
        # 2 opps válidas (Snorlax + Gengar) — Gengar yielded, line escrita, set_complete não.
        # A linha truncada não conta. Recovery deve achar 2.
        self.assertion("log mentions '2 opportunities' recovered",
                       "2 opportunities" in out_log.lower(),
                       detail=f"log: {out_log[-300:]}")

    def test_recovery_empty_checkpoint_graceful(self):
        print("\n[4] Recovery from empty file is graceful (no crash)")
        path = self.tmpdir / "test4.checkpoint.jsonl"
        path.write_text("", encoding="utf-8")  # arquivo vazio (caso edge: crash antes do header)
        out_xlsx = self.tmpdir / "test4.recovered.xlsx"
        result = subprocess.run(
            [sys.executable, str(REPO_ROOT / "scripts" / "recover_from_checkpoint.py"),
             "--checkpoint", str(path),
             "--output", str(out_xlsx)],
            capture_output=True, text=True, env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        self.assertion("empty checkpoint → exit 0 (graceful)",
                       result.returncode == 0,
                       detail=f"stderr={result.stderr[:200]}")
        # XLSX é gerado mesmo com 0 opps (planilha vazia)
        self.assertion("XLSX gerado mesmo sem opps",
                       out_xlsx.exists())


def main() -> int:
    print("=" * 70)
    print("  test_checkpoint_crash_recovery.py — scanner v2.6 regression")
    print("=" * 70)
    runner = TestRunner()
    try:
        runner.test_writer_writes_header_opps_setcomplete()
        runner.test_recovery_from_clean_checkpoint()
        runner.test_recovery_preserves_fx_from_header()
        runner.test_recovery_with_partial_last_line()
        runner.test_recovery_empty_checkpoint_graceful()
    finally:
        print("\n" + "=" * 70)
        print(f"  RESULT: {runner.passed} passed / {runner.failed} failed")
        print("=" * 70)
        runner.cleanup()

    return 0 if runner.failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
